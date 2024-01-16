from ast import keyword
from calendar import month_abbr
from threading import currentThread
from unicodedata import name
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.tokens import default_token_generator
from django.urls import reverse

from django.db.models import Q
from datetime import date, datetime, timedelta

from psutil import users

# from drf_api_logger import API_LOGGER_SIGNAL
from .models import *
import random
import requests

# import bs4sta
import bs4
import datetime
import os
import math

# import os.path
from .contactus import contact_function
from .serializers import *
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from negbuy import settings
from paytmchecksum import PaytmChecksum
import uuid

# Nilabh Sahu
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from . import Checksum
from django.contrib.sites.shortcuts import get_current_site
from django.utils.encoding import force_bytes, force_text
from .utils import TokenGenerator, generate_token
from django.views.decorators.csrf import csrf_exempt
from rest_framework.views import APIView
import pandas as pd
from django.shortcuts import redirect, render
from django.utils.http import urlsafe_base64_decode
from django.utils.http import urlsafe_base64_encode
import re
import geopy.distance
from geopy.geocoders import Nominatim
import csv
from rest_framework.decorators import api_view
from rest_framework.response import Response
import calendar
from rest_framework.renderers import JSONRenderer


from deep_translator import GoogleTranslator

#######################

from django.utils import timezone




def api_logger(url1, user_id):
    csv_file_path = os.path.join(settings.BASE_DIR, "apilogs.csv")

    with open(csv_file_path, "r") as file:
        serial = len(list(csv.reader(file)))

    user = userDB.objects.filter(user_id=user_id)
    new_data = [
        serial,
        url1,
        user_id,
        user[0].first_name,
        user[0].phone,
        timezone.localtime().strftime("%I:%M %p"),
        timezone.localtime().strftime("%d %b %Y"),
    ]

    with open(csv_file_path, mode="a", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(new_data)


@api_view(["POST"])
def check_phone_status(request):
    try:
        phone = request.data["phone"]
        get_user_phone_data = userDB.objects.filter(phone=phone)

        if get_user_phone_data:
            return_obj = True
        else:
            return_obj = False

        return Response({"status": True, "message": "success", "data": return_obj})
    except Exception as e:
        return Response({"status": False, "message": str(e), "data": {}})


@api_view(["POST"])  # Finalized on 27/07/2022 - Nilabh
def login(request):
    user_id = request.headers["User-id"]
    phone = request.data["phone"]
    if userDB.objects.filter(user_id=user_id).exists():
        usr = userDB.objects.get(user_id=user_id)
        if usr.role == "Buyer":
            if (
                usr.address_line1
                and usr.city
                and usr.state
                and usr.postal_code
                and usr.country is not None
            ):
                usr.address_verified = True
            else:
                usr.address_verified = False
            response = {
                "status": False,
                "user_id": usr.user_id,
                "phone": usr.phone,
                "first_name": usr.first_name,
                "last_name": usr.last_name,
                "userType": usr.role,
                "imageUrl": usr.profile_picture.url,
                "Authentication": usr.auth,
                "address_verified": usr.address_verified,
                "language": "",
                "user_bio": 0 if usr.first_name is None else 1,
            }
        elif usr.role == "Seller" and usr.phone == phone:
            usr.auth = False
            response = {
                "status": False,
                "Authentication": usr.auth,
                "message": "This number already exists as seller",
            }
    else:
        userDB.objects.create(user_id=user_id, phone=phone)
        usr = userDB.objects.get(user_id=user_id)
        response = {
            "status": True,
            "user_id": user_id,
            "phone": phone,
            "Authentication": usr.auth,
            "address_verified": usr.address_verified,
            "first_name": "",
            "last_name": "",
        }

    api_logger("login api", user_id)
    return Response(response, status=200)


@api_view(["POST"])
def seller_login(request):
    phone = request.data["phone"]
    password = request.data["password"]

    try:
        usr = userDB.objects.get(role="Seller", phone=phone, password=password)
        if (
            usr.address_line1
            and usr.city
            and usr.state
            and usr.postal_code
            and usr.country is not None
        ):
            usr.address_verified = True
        else:
            usr.address_verified = False
        try:
            usrBank = bankDetail.objects.get(user=usr)
            accountName = usrBank.accountName
            accountNumber = usrBank.accountNumber
            accountIfsc = usrBank.accountIfsc
        except:
            accountName = None
            accountNumber = None
            accountIfsc = None
        response = {
            "status": True,

            "user_id": usr.user_id,
            "phone": usr.phone,
            "password": usr.password,
            "seller_name": usr.seller_name,
            "date_of_birth": usr.date_of_birth,
            "address_verified": usr.address_verified,
            "company": usr.company,
            "gst_number": usr.gst_number,
            "account_name": accountName,
            "account_number": accountNumber,
            "account_ifsc": accountIfsc,
        }
        api_logger("seller login api", usr.user_id)

    except Exception as e:
        response = {
            "status": False,
            "message": "Incorrect phone or password",
            "error_msg": str(e),
        }

    finally:
        return Response(response, status=200)


@api_view(["POST"])
def seller_signup(request):
    user_id = request.headers["User-id"]
    password = request.data["password"]
    phone = request.data["phone"]

    try:
        usr = userDB.objects.get(Q(user_id=user_id) | Q(phone=phone))
        sameField = "User Id " if usr.user_id == user_id else "Phone number "
        response = {"status": False, "message": sameField + "already exists"}

    except:
        userDB.objects.create(
            user_id=user_id, password=password, phone=phone, role="Seller"
        )
        response = {
            "status": True,
            "user_id": user_id,
            "phone": phone,
            "password": password,
            "seller_name": "",
        }

        api_logger("seller sign up api", user_id)

    finally:
        return Response(response, status=200)


# def addProduct(request, user):
#     price_choice = request.data['price_choice']

#     category = request.data['category']
#     try:
#         category_record = productCategory.objects.get(name=category)
#     except:
#         category_record = productCategory.objects.create(name=category)

#     if price_choice == 'Add Price':
#         product_record = product.objects.create(
#             username=user,
#             name=request.data['name'],
#             category_id=category_record,
#             brand=request.data['brand'],
#             main_image=request.data['main_image'],
#             keyword=request.data['keywords'],
#             color=request.data['colors'],
#             size=request.data['size'],
#             details=request.data['details'],
#             price_choice=price_choice,
#             price=request.data['price'],
#             mrp=request.data['mrp'],
#             sale_price=request.data['sale_price'],
#             sale_startdate=request.data['sale_startdate'],
#             sale_enddate=request.data['sale_enddate'],
#             manufacturing_time=request.data['manufacturing_time'],
#             maximum_order_quantity=request.data['maximum_order_quantity'],
#             # terms = terms_record,
#             weight=request.data['weight'],
#             transportation_port=request.data['transportation_port'],
#             packing_details=request.data['packing_details'],
#             packing_address=request.data['packing_address']
#         )
#         product_record.save()
#         images = dict((request.data).lists())['image']
#         for image in images:
#             productImages.objects.create(product=product_record, image=image)

#     elif price_choice == 'Price according to quantity':
#         product_record = product.objects.create(
#             username=user,
#             name=request.data['name'],
#             category_id=category_record,
#             brand=request.data['brand'],
#             main_image=request.data['main_image'],
#             keyword=request.data['keywords'],
#             color=request.data['colors'],
#             size=request.data['size'],
#             details=request.data['details'],
#             price_choice=price_choice,
#             quantity_price=request.data['quantity_price'],
#             # terms = terms_record,
#             weight=request.data['weight'],
#             transportation_port=request.data['transportation_port'],
#             packing_details=request.data['packing_details'],
#             packing_address=request.data['packing_address']
#         )
#         product_record.save()
#         images = dict((request.data).lists())['image']
#         for image in images:
#             productImages.objects.create(product=product_record, image=image)


# @api_view(['POST'])
# def product_upload_api(request):
#     user_id = request.headers['User-id']
#     try:
#         user = userDB.objects.get(user_id=user_id, role='Seller')
#         addProduct(request, user)
#         return Response({'success': 'Product Added'}, status=200)
#     except Exception as e:
#         return Response({'status': 'Error', 'error_msg': str(e)})


# def getProductObject(product):
#     try:
#         inventory = product.inventory_id.quantity
#     except:
#         inventory = ''

#     try:
#         productCategory = product.category_id.name
#     except:
#         productCategory = ''

#     try:
#         prodImages = productImages.objects.filter(product=product)
#         imageURL = []
#         for prodImage in prodImages:
#             imageURL.append(prodImage.image.url)
#     except:
#         imageURL = []

#     detailsList = []
#     list = [str.strip() for str in product.details.split(',')]
#     for eachDetail in list:
#         keyValue = eachDetail.split(':')
#         detailsObject = {
#             'title': keyValue[0].strip(),
#             'description': keyValue[1].strip(),
#         }
#         detailsList.append(detailsObject)

#     object = {
#         'id': product.id,
#         'name': product.name,
#         # 'sku': product.sku,
#         # 'category': product.category_id.name,
#         # 'inventory': inventory,
#         'main_image': str(product.main_image),
#         # 'image': imageURL,
#         # 'featured_products': product.featured_products,
#         # 'best_selling_products': product.best_selling_products,
#         # 'hot_selling_products': product.hot_selling_products,
#         # 'fast_dispatch': product.fast_dispatch,
#         # 'ready_to_ship': product.ready_to_ship,
#         # 'customized_product': product.customized_product,
#         # 'brand': product.brand,
#         # 'keyword': [str.strip() for str in product.keyword.split(',')],
#         #'color': [str.strip() for str in product.color.split(',')],
#         # 'size': [str.strip() for str in product.size.split(',')],
#         # 'details': detailsList,
#         # 'price_choice': product.price_choice,
#         # 'price': product.price,
#         # 'mrp': product.mrp,
#         # 'sale_price': product.sale_price,
#         # 'sale_startdate': product.sale_startdate,
#         # 'sale_enddate': product.sale_enddate,
#         # 'manufacturing_time': product.manufacturing_time,
#         # 'quantity_price': product.quantity_price,
#         # 'maximum_order_quantity': product.maximum_order_quantity,
#         # 'weight': product.weight,
#         # 'transportation_port': product.transportation_port,
#         # 'packing_details': product.packing_details,
#         # 'packing_address': product.packing_address,
#         # 'status': product.status,
#         # 'created_at': product.created_at,
#         # 'modified_at': product.modified_at,
#         # 'deleted_at': product.deleted_at,
#         # 'rating': {
#         #     'rate': 3.0,
#         #     'count': 430
#         # }
#     }
#     return object


# @api_view(['POST'])
# def product_info(request):
#     try:
#         review_list = []
#         product_id = request.data['product_id']
#         searched = request.data['searched']
#         product_info = product.objects.get(id=product_id)
#         product_seller = product_info.username.username
#         usr = userDB.objects.get(username=product_seller)
#         now=datetime.now()
#         weekday = now.strftime("%a")
#         date = now.strftime("%b" " %d")
#         year = now.strftime("%Y")
#         product_detail = product_detail_db.objects.filter(
#             product__id=product_id)
#         product_reviews = review_db.objects.filter(product__id=product_id)
#         product_serialized = ProductSerializer(product_info).data
#         try:
#             vv= sellerAnalytics.objects.get(seller=usr, date=date)
#             vv.checked_count +=1
#             if searched.lower() == 'true':
#                 vv.search_count +=1
#             vv.save()
#         except Exception as e:
#             rr = sellerAnalytics.objects.create(seller =usr, date=date)
#             rr.checked_count =1
#             rr.day = weekday
#             if searched.lower() == 'true':
#                 rr.search_count =1
#             rr.save()
#         finally:
#         # try:
#         #     review_serialized = ReviewSerializer(
#         #         product_reviews, many=True).data
#         #     for i in review_serialized:
#         #         review_list.append(i['rating'])

#         #     review_dict = {
#         #         'rating': round(sum(review_list)/len(review_list)),
#         #         'count': len(review_list)
#         #     }

#         # except Exception as e:
#         #     review_dict = {
#         #         'rating': 0,
#         #         'count': 0
#         #     }

#         # product_serialized['reviews'] = review_dict

#             obj = {
#                 'status': 'True',
#                 'message': 'success',
#                 'data': product_serialized,
#             }

#             return Response(obj)
#     except Exception as e:
#         return Response({'error_msg': str(e)}, status=500)


# @api_view(['GET'])
# def featured_product_api(request):
#     featured_products = product.objects.filter(featured_products=True).order_by('?')
#     featured_serialized = ProductSerializer(featured_products, many=True).data

#     # reviews = review_db.objects.filter(
#     #     id__in=featured_products.values_list('id', flat=True))
#     # reviews_serialized = ReviewSerializer(reviews, many=True).data

#     return Response({
#         'Message': 'OK',
#         'data': featured_serialized
#     })


# @api_view(['GET'])
# def fast_dispatch_api(request):
#     fast_dispatch_products = product.objects.filter(fast_dispatch=True).order_by('?')

#     fast_dispatch_serialized = ProductSerializer(
#         fast_dispatch_products, many=True).data

#     return Response({
#         'status': 200,
#         'message': 'success',
#         'data': fast_dispatch_serialized
#     })


# @api_view(['GET'])
# def ready_to_ship_api(request):
#     ready_to_ship_products = product.objects.filter(ready_to_ship=True).order_by('?')

#     ready_to_ship_products_serialized = ProductSerializer(
#         ready_to_ship_products, many=True).data

#     return Response({
#         'status': 200,
#         'message': 'success',
#         'data': ready_to_ship_products_serialized
#     })


# @api_view(['GET'])
# def customized_product_api(request):
#     customized_products = product.objects.filter(customized_product=True).order_by('?')
#     customized_products_serialized = ProductSerializer(
#         customized_products, many=True).data

#     return Response({
#         'status': 200,
#         'message': 'success',
#         'data': customized_products_serialized
#     })


# @api_view(['GET'])
# def new_arrivals_api(request):
#     new_arrivals_products = product.objects.all().order_by('-id')[:10]

#     newArrival_products_serialized = ProductSerializer(
#         new_arrivals_products, many=True).data
#     return Response({
#         'status': 200,
#         'message': 'success',
#         'data': newArrival_products_serialized
#     })


# @api_view(['GET'])
# def top_selling_api(request):
#     top_selling_products = list(product.objects.all())
#     random_top_selling_products = random.sample(top_selling_products, 10)

#     topSelling_products_serialized = ProductSerializer(
#         random_top_selling_products, many=True).data
#     return Response({
#         'status': 200,
#         'message': 'success',
#         'data': topSelling_products_serialized
#     })


# @api_view(['POST'])
# def add_to_cart(request):
#     user_id = request.headers['User-id']

#     try:
#         usr = userDB.objects.get(user_id=user_id)
#         product_id = int(request.data['product_id'])
#         quantity = int(request.data['quantity'])

#         try:
#             record = cart.objects.get(user_id=usr.id, product_id=product_id)
#             record.quantity = quantity
#             record.save()
#         except:
#             cart.objects.create(
#                 user_id=usr.id,
#                 product_id=product_id,
#                 quantity=quantity,
#             )
#         return Response({'status': 'success'}, status=200)

#     except Exception as e:
#         return Response({'status': 'error', 'error_msg': str(e)}, status=401)


# @api_view(['POST'])
# def remove_from_cart(request):
#     user_id = request.headers['User-id']

#     try:
#         usr = userDB.objects.get(user_id=user_id)
#         product_id = int(request.data['product_id'])
#         removed_quantity = int(request.data['removed_quantity'])
#         item_quantity = cart.objects.get(
#             user_id=usr.id, product_id=product_id).quantity
#         new_quantity = item_quantity - removed_quantity

#         if new_quantity <= 0:
#             cart.objects.get(user_id=usr.id, product_id=product_id).delete()
#             return Response({'status': 'success', 'msg': 'item removed from cart'})
#         else:
#             record = cart.objects.get(user_id=usr.id, product_id=product_id)
#             record.quantity = new_quantity
#             record.save()
#             return Response({'status': 'success', 'msg': 'item quantity decreased to ' + str(new_quantity)})

#     except Exception as e:
#         return Response({'status': 'error', 'error_msg': str(e)})


# @api_view(['GET'])
# def my_cart(request):
#     status = 200
#     cart_items = []
#     user_id = request.headers['User-id']
#     try:
#         usr = userDB.objects.get(user_id=user_id)
#         all_items = cart.objects.filter(user_id=usr.id)
#         for each_item in all_items:
#             object = getProductObject(each_item.product)
#             object['quantity'] = each_item.quantity
#             cart_items.append(object)
#     except:
#         status = 401

#     return Response(cart_items, status=status)


# @api_view(['GET','POST'])
# def user_products_api(request):
#     user_id = request.headers['User-id']
#     offset =request.data['offset']
#     limit = request.data['limit']
#     response = []
#     # try:
#     user = userDB.objects.get(user_id=user_id, role='Seller')
#     if user:
#         user_products = product.objects.filter(username=user).order_by('-id')
#         user_products_serialized = ProductSerializer(user_products, many=True)
#         return Response({
#             'data': user_products_serialized.data[int(offset):int(limit)+int(offset)],
#             'total':len(user_products_serialized.data)
#         })
#     else:
#         return Response(response, status=200)
#     # except Exception as e:
#     #     return Response({'status': 'error', 'error_msg': str(e)}, status=401)


def getGstObject(gst_number, divData):
    object = {
        'gst_number': gst_number,
        'tradeName': divData[0].input.get('value'),
        'legalName': divData[1].input.get('value'),
        'type': divData[2].input.get('value'),
        'regDate': divData[3].input.get('value'),
        'constBusiness': divData[4].input.get('value'),
        'businessNature': divData[5].input.get('value'),
        'principalPlace': divData[6].input.get('value'),
    }
    return object


# @api_view(['POST'])
# def verify_gst(request):
#     gst_number = request.data.get('gstNo')


# @api_view(["POST"])
# def verify_gst(request):
#     user_id = request.headers["User-id"]
#     gst_number = request.data["gstNo"]
#     # following has to be changed when gst is actially verified
#     usr = userDB.objects.get(user_id=user_id, role="Seller")
#     usr.gst_number = request.data["gstNo"]
#     usr.save()

#     try:
#         addr = "https://irisgst.com/gstin-filing-detail/?gstinno=" + gst_number
#         response = requests.get(addr)
#         htmlPage = bs4.BeautifulSoup(response.text, "html.parser")
#         divData = htmlPage.find_all('div', {'class': 'form-group'})
#         response = getGstObject(gst_number, divData)
#         usr.gst_number = response['gst_number']
#         usr.save(update_fields=['gst_number'])
#         usr.save()
#         return Response(
#             {
#                 "status": "True",
#                 "gst_number": gst_number,
#                 "trade": "testing",
#                 "legalName": "testing negbuy",
#             },
#             status=200,
#         )
#     except:
#         gst_pattern = "^(?:(?:[0-4]\d|[5][0-3])|[6-9][0-9])[A-Z]{5}\d{4}[A-Z]{1}[A-Z\d]{1}[Z]{1}[A-Z\d]{1}$"

#         if re.match(gst_pattern, gst_number):
#             check = gst_number[-1]
#             gst = gst_number[:-1]
#             l = [int(c) if c.isdigit() else ord(c) - 55 for c in gst]
#             l = [val * (ind % 2 + 1) for (ind, val) in list(enumerate(l))]
#             l = [(int(x / 36) + x % 36) for x in l]
#             csum = 36 - sum(l) % 36
#             csum = str(csum) if (csum < 10) else chr(csum + 55)
#             if check == csum:
#                 return Response({"msg": "Valid GST Number", "status": True}, status=200)
#             else:
#                 return Response({"msg": "Invalid GST Number", "status": False})
#         else:
#             # return Response({'msg':'Invalid GST Number'})
#             return Response(
#                 {
#                     "status": "True",
#                     "gst_number": request.data["gstNo"],
#                     "trade": "testing",
#                     "legalName": "testing signup",
#                 },
#                 status=200,
#             )


@api_view(["POST"])
def bank_details(request):
    user_id = request.headers["User-id"]

    try:
        usr = userDB.objects.get(user_id=user_id, role="Seller")
        accName = request.data["accName"]
        accNo = request.data["accNo"]
        ifsc = request.data["ifsc"]

        bankDetail.objects.create(
            user_id=usr.id, accountName=accName, accountNumber=accNo, accountIfsc=ifsc
        )
        return Response({"status": "success"}, status=200)

    except:
        return Response({"status": "User does not exist"}, status=401)


@api_view(["POST"])
def seller_details(request):
    user_id = request.headers["User-id"]

    try:
        usr = userDB.objects.get(user_id=user_id, role="Seller")
        usr.seller_name = request.data["seller_name"]
        usr.date_of_birth = request.data["date_of_birth"]
        usr.email = request.data["email"]
        usr.company = request.data["company"]
        usr.address_line1 = request.data["address"]
        usr.document_verification = request.FILES["document_verification"]
        usr.save(
            update_fields=[
                "seller_name",
                "date_of_birth",
                "email",
                "company",
                "document_verification",
            ]
        )
        return Response({"status": "success"}, status=200)

    except Exception as e:
        return Response({"status": "error", "error_msg": str(e)}, status=401)


# @api_view(['POST'])
# def search_category(request):
#     response = []

#     category_selected = request.data['category_selected']
#     raw_string = request.data['category']
#     file_path = 'static/categories/'+str(category_selected).strip()+'.txt'

#     if (not (raw_string and raw_string.strip())):
#         keywords = raw_string.replace('>', '').replace(
#             '& ', '').replace('and ', '').strip()

#     if os.path.exists(file_path):
#         if len(keywords) == 0:
#             with open(file_path) as file:
#                 catLines = file.readlines()
#                 for line in catLines:
#                     response.append(line.strip())
#             return Response({
#                 'status': True,
#                 'message': 'Success',
#                 'data': response
#             })

#         else:
#             file_path = 'static/categories/' + \
#                 str(category_selected).strip()+'.txt'
#             with open(file_path) as file:
#                 for line in file:
#                     if keywords.lower() in line.lower().replace('& ', ''):
#                         response.append(line.strip())
#                 if len(response) == 0:
#                     keywords = keywords.replace(',', '')
#                     keywordList = keywords.split(' ')
#                     with open(file_path) as file:
#                         for line in file:
#                             if any(keyword.lower() in line.lower() for keyword in keywordList):
#                                 response.append(line.strip())
#             return Response(response, status=200)
#     else:
#         return Response({
#             'status': True,
#             'message': 'Success',
#             'data': 'Invalid Category'
#         })


# @api_view(['GET'])
# def get_ports(request):
#     all_ports = port.objects.filter(country='Brazil')
#     port_serializer = PortSerializer(all_ports, many=True)
#     return Response({
#         'status': True,
#         'message': 'Success',
#         'data': port_serializer.data
#     })


# @api_view(['GET'])
# def get_categories(request):
#     try:
#         # /home/vf2586e813kg/api.negbuy/static
#         path = "/home/negbswof/api.negbuy/static/categories"
#         # path = "/Users/habibahmed/Desktop/NegBuy/negbuy/static/categories"
#         dir_list = os.listdir(path)
#         mylst = map(lambda each: each.strip(".txt"), dir_list)

#         # for item in mylst:
#         #     productCategory.objects.create(
#         #         name=item,
#         #         desc="Lorem Ipsum"
#         #     )

#         return Response({
#             'status': True,
#             'message': 'Success',
#             'data': mylst
#         })
#     except Exception as e:
#         return Response({
#             'status': False,
#             'message': e,
#             'data': ''
#         })


# @api_view(['GET'])
# def get_orders(request):
#     user_id = request.headers['User-id']
#     all_orders_list = list()
#     try:
#         usr = userDB.objects.get(user_id=user_id, role='Seller')
#         # get orders details
#         all_orders = orders.objects.filter(user=usr)
#         for each_order in all_orders:
#             obj = {
#                 'order_date': each_order.created_at,
#                 'order_number': each_order.order_number,
#                 'product_name': each_order.product_info.name,
#                 'product_quantity': each_order.order_quantity,
#                 'ship_date': each_order.shipping_date,
#                 'delivery_date': each_order.delivery_date,
#                 'order_status': each_order.status
#             }

#             all_orders_list.append(obj)
#         return Response({
#             'status': True,
#             'message': 'Success',
#             'data': all_orders_list
#         })
#     except Exception as e:
#         return Response({
#             'status': False,
#             'message': e,
#             'data': ''
#         })


# @api_view(['POST'])
# def contactus_function(request):
#     try:
#         name = request.data['name']
#         contact_num = request.data['number']
#         req_email = request.data['email']
#         message = request.data['message']
#         full_message = "Please contact "+name+" " + \
#             contact_num+" "+req_email+" for "+message
#         cont_res = contact_data.objects.create(message=full_message)
#         if cont_res:
#             return Response({
#                 'status': True,
#                 'message': 'Success',
#                 'data': "Message sent"
#             })
#         else:
#             return Response({
#                 'status': False,
#                 'message': 'Error',
#                 'data': "Please enter details again"
#             })
#     except:
#         return Response({
#             'status': False,
#             'message': 'Error',
#             'data': "Please enter details again"
#         })


# @api_view(['POST'])
# def delete_product(request):
#     try:
#         user = request.headers['User-id']
#         product_id = request.data['product_id']

#         user_details = userDB.objects.get(user_id=user)
#         product_details = product.objects.get(
#             id=product_id, username=user_details).delete()

#         return Response({
#             'status': True,
#             'message': 'Success',
#             'data': "Deleted successfully"
#         })
#     except Exception as e:
#         return Response({
#             'status': False,
#             'message': 'Error',
#             'data': str(e)
#         })


# @api_view(['POST'])
# def product_detail(request):
#     try:
#         product_id = request.data['product_id']
#         image_list = []

#         product_details = product.objects.filter(id=product_id)
#         image_details = productImages.objects.filter(product__id=product_id)
#         product_db_details = product_detail_db.objects.filter(
#             product__id=product_id)

#         serializer = ProductSerializer(product_details, many=True)
#         productDetails_serializer = ProductDetailSerializer(
#             product_db_details, many=True)
#         img_serializer = ImageSerializer(image_details, many=True)
#         for dic in img_serializer.data:
#             for key in dic:
#                 if key == 'image':
#                     image_list.append(dic[key])

#         data_dict = {
#             'product details': serializer.data,
#             'image detail': productDetails_serializer.data,
#             'images': image_list
#         }

#         return Response({
#             'status': True,
#             'message': 'Success',
#             'data': data_dict
#         })

#     except Exception as e:
#         return Response({
#             'status': 'False',
#             'message': 'Error',
#             'data': str(e)
#         })


# @api_view(['GET'])
# def best_selling(request):
#     try:
#         best_selling = product.objects.filter(best_selling_products=True).order_by('?')[:10]
#         bestSelling_serialized = ProductSerializer(
#             best_selling, many=True).data

#         return Response({
#             'status': 'True',
#             'message': 'Success',
#             'data': bestSelling_serialized
#         })

#     except Exception as e:
#         return Response({
#             'status': 'Error',
#             'message': e,
#             'data': ''
#         })


# @api_view(['GET'])
# def hot_selling(request):
#     try:
#         hot_selling = product.objects.filter(hot_selling_products=True).order_by('?')[:10]
#         hotSelling_serialized = ProductSerializer(hot_selling, many=True).data
#         return Response({
#             'status': 'True',
#             'message': 'Success',
#             'data': hotSelling_serialized
#         })

#     except Exception as e:
#         return Response({
#             'status': 'Error',
#             'message': e,
#             'data': ''
#         })

# # =================================================================================#


# # new api of read file product_list.json.....


# @ api_view(['GET'])
# def read_json(request):
#     # remove on live version
#     # all_products = productCategory.objects.all().delete()
#     # all_inventory = productInventory.objects.all().delete()
#     user_data_all = userDB.objects.all()
#     category_data = productCategory.objects.all()

#     bool_list = [True, False]
#     with open('product_lists.json', 'r') as f:
#         jsondata = f.read()
#         obj = json.loads(jsondata)
#         user_id = ['Neyyjn9DCncSDQDWtJNEa57u6el2',
#                    'M1TOJ78YaQcGJZ6fsuMoxCB0ad93', '2utKoA5gheWtqe5ZRBNqdFamtQr1']
#         # user_id = ['Neyyjn9DCncSDQDWtJNEa57u6el2']
#         for pd in obj:

#             # inventory_data = Inventory.objects.create(quantity=20)
#             # inventory_data.save()

#             user_data = random.choice(user_data_all)
#             category_obj = random.choice(category_data)
#             # inventory_obj = inventory_data
#             name = str(pd['name'])
#             sku = str(pd['sku'])
#             main_image = str(pd['main_image'])
#             hot_selling_products = random.choice(bool_list)
#             best_selling_products = random.choice(bool_list)
#             featured_products = random.choice(bool_list)
#             fast_dispatch = random.choice(bool_list)
#             ready_to_ship = random.choice(bool_list)
#             customized_product = random.choice(bool_list)
#             brand = str(pd['brand'])
#             keyword = str(pd['keyword'])
#             color = str(pd['color'])
#             size = str(pd['size'])
#             details = str(pd['details'])
#             price_choice = str(pd['price_choice'])
#             price = str(pd['price'])
#             mrp = str(pd['mrp'])
#             sale_price = str(pd['sale_price'])
#             sale_startdate = str(pd['sale_startdate'])
#             sale_enddate = str(pd['sale_enddate'])
#             manufacturing_time = str(pd['manufacturing_time'])
#             quantity_price = str(pd['quantity_price'])
#             maximum_order_quantity = str(pd['maximum_order_quantity'])
#             weight = str(pd['weight'])
#             transportation_port = str(pd['transportation_port'])
#             packing_details = str(pd['packing_details'])
#             packing_address = str(pd['packing_address'])
#             status = str(pd['status'])
#             created_at = str(pd['created_at'])
#             modified_at = str(pd['modified_at'])
#             deleted_at = str(pd['deleted_at'])

#             product_obj = product.objects.create(
#                 username=user_data,
#                 name=name,
#                 sku=sku,
#                 main_image=main_image,
#                 category_id=category_obj,
#                 # inventory_id=inventory_obj,
#                 hot_selling_products=hot_selling_products,
#                 best_selling_products=best_selling_products,
#                 featured_products=True,
#                 fast_dispatch=fast_dispatch,
#                 ready_to_ship=ready_to_ship,
#                 customized_product=customized_product,
#                 brand=brand,
#                 keyword=keyword,
#                 color=color,
#                 size=size,
#                 details=details,
#                 price_choice=price_choice,
#                 price=0 if price == 'None' else float(price),
#                 mrp=0 if mrp == 'None' else float(mrp),
#                 sale_price=sale_price,
#                 sale_startdate=sale_startdate,
#                 sale_enddate=sale_enddate,
#                 manufacturing_time=manufacturing_time,
#                 quantity_price=quantity_price,
#                 maximum_order_quantity=maximum_order_quantity,
#                 weight=weight,
#                 transportation_port=transportation_port,
#                 packing_details=packing_details,
#                 packing_address=packing_address,
#                 status=status,
#                 created_at=created_at,
#                 modified_at=modified_at,
#                 deleted_at=deleted_at
#             )
#             product_obj.save()
#             Inventory.objects.create(product=product_obj, quantity=0)

#     return Response({
#         'status': 'success',
#         'message': 'Added products',
#         'data': obj
#     })


# # ---------------------------- Seller.Negbuy -------------------------------------- #

# # =========== DMS to DEC ============ #

# def dms2dec(dms_str):

#     dms_str = re.sub(r'\s', '', dms_str)

#     sign = -1 if re.search('[swSW]', dms_str) else 1

#     numbers = [*filter(len, re.split('\D+', dms_str, maxsplit=4))]

#     degree = numbers[0]
#     minute = numbers[1] if len(numbers) >= 2 else '0'
#     second = numbers[2] if len(numbers) >= 3 else '0'
#     frac_seconds = numbers[3] if len(numbers) >= 4 else '0'

#     second += "." + frac_seconds
#     return sign * (int(degree) + float(minute) / 60 + float(second) / 3600)


# # Funtion to Read and add Port Details from XLSX
# @ api_view(['POST'])
# def add_ports(request):
#     start_line = int(request.data['start_line'])
#     end_line = int(request.data['end_line'])
#     wb = load_workbook('portdata.xlsx')
#     ws = wb.active

#     for row in range(start_line, end_line+1):
#         for col in range(1, 5):
#             char = get_column_letter(col)
#             country = ws[char+str(row)].value
#             portname = ws[get_column_letter(col+1)+str(row)].value
#             lat = ws[get_column_letter(col+2)+str(row)].value
#             lon = ws[get_column_letter(col+3)+str(row)].value
#             try:
#                 dec_lat = dms2dec(lat)
#                 dec_lon = dms2dec(lon)
#             except Exception as e:
#                 print(e)
#             port.objects.create(
#                 name=portname,
#                 country=country,
#                 latitude=dec_lat,
#                 longitude=dec_lon
#             )
#             break

#     return Response({
#         'status': 'success',
#         'message': 'Added ports',
#     })


# # Api to fetch order with status Running
# @ api_view(['POST'])
# def my_orders(request):
#     user_id = request.headers['User-id']
#     running_orders = orders.objects.filter(status='Running', user__id=user_id)

#     order_serializer = OrderSerializer(running_orders, many=True)

#     return Response({
#         'status': 'success',
#         'message': 'Added ports',
#         'data': order_serializer.data
#     })


# # Api to fetch orders with status Completed
# @ api_view(['POST'])
# def order_history(request):
#     user_id = request.headers['User-id']
#     completed_orders = orders.objects.filter(
#         status='Completed', user__id=user_id)

#     order_serializer = OrderSerializer(completed_orders, many=True)

#     return Response({
#         'status': 'success',
#         'message': 'Added ports',
#         'data': order_serializer.data
#     })

# # Add order Note to a particular order


# @ api_view(['POST'])
# def order_note(request):

#     user_id = request.headers['User-id']
#     order_id = request.data['order_id']
#     note = request.data['note']

#     order_obj = orders.objects.get(order_id=order_id)
#     order_obj.order_note = note
#     order_obj.save()

#     return Response({
#         'status': 'success',
#         'message': 'Added ports',
#         'data': "Success"
#     })

# # Get order Details


# @ api_view(['GET'])
# def order_details(request):
#     user_id = request.headers['User-id']
#     order_id = request.data['order_id']

#     order = orders.objects.filter(id=order_id).values()
#     order_item = orders.objects.get(id=order_id)
#     product_detail = product.objects.filter(
#         name=order_item.product_info).values()
#     user_detail = userDB.objects.filter(user_id=user_id)

#     order_serialized = OrderSerializer(order, many=True)
#     product_serialized = ProductSerializer(product_detail, many=True)
#     user_serialized = UserSerializer(user_detail, many=True)

#     quantity = 0
#     price = 0
#     # Price calc
#     for item in order:
#         quantity = int(item.get('order_quantity'))

#     for item in product_detail:
#         price = item.get('price')

#     cart_total = quantity * price

#     data_dict = {
#         'Product Details': product_serialized.data,
#         'Order Details': order_serialized.data,
#         'Customer Details': user_serialized.data,
#         'Price Details': cart_total
#     }

#     return Response({
#         'status': 'success',
#         'message': 'Order Details',
#         'data': data_dict
#     })


# # Image uploading of Products from CSV file
# @ api_view(['GET'])
# def read_csv(request):
#     with open('images.csv', 'r') as csv_file:
#         csv_reader = csv.reader(csv_file)
#         for a in csv_reader:
#             if product.objects.filter(name=a[0]).exists():
#                 product_object = product.objects.get(name=a[0])
#                 productImages.objects.create(
#                     product=product_object,
#                     image=a[1]
#                 )
#     return Response({
#         'status': 'success'
#     })


# # Distance between Longitude and latitude
# @ api_view(['GET'])
# def port_distance(request):
#     a_lat = float(request.data['a_lat'])
#     a_lon = float(request.data['a_lon'])

#     geolocator = Nominatim(user_agent="geoapiExercises")

#     location = geolocator.reverse(str(a_lat)+","+str(a_lon))
#     address = location.raw['address']
#     country = address.get('country').lower()
#     dist = 0
#     cl_port = ''
#     coords_1 = (a_lat, a_lon)
#     qs = port.objects.filter(country__icontains=country).values()
#     for item in qs:
#         latitude_val = item.get('latitude')
#         longitude_val = item.get('longitude')
#         coords_2 = (latitude_val, longitude_val)
#         distance = geopy.distance.geodesic(coords_1, coords_2).km
#         if dist < distance:
#             dist = distance
#             cl_port = item.get('name')
#             data_dict = {
#                 'Distance': dist,
#                 'port': cl_port
#             }
#         return Response({
#             'status': 'success',
#             'message': 'Order Details',
#             'data': data_dict
#         })
#     return Response({
#         'status': 'success',
#         'message': 'Order Details',
#         'data': ''
#     })


# # Add reviews
# @ api_view(['POST'])
# def post_review(request):
#     user_id = request.headers['user-id'] # buyer
#     review_description = request.data['review_description'] # review_description
#     review_title = request.data['review_title']
#     product_id = request.data['product_id']
#     rating = request.data['rating']
#     try:
#         user = userDB.objects.get(user_id=user_id)
#         products = product.objects.get(id=product_id)

#         review_db.objects.create(
#             user=user,
#             product=products,
#             review_description=review_description,
#             rating=rating,
#             review_title=review_title
#         )

#         user_obj = userDB.objects.get(user_id=user_id)
#         userSerialized = UserNEWSerializer(user_obj)

#         data_dict = {
#             'review_title': review_title,
#             'review_description': review_description,
#             'rating': int(rating),
#             'user': userSerialized.data
#         }

#         return Response({
#             'status': 'success',
#             'message': 'Review posted',
#             'data': data_dict
#         })
#     except Exception as e:
#         return Response({
#             'status': 'success',
#             'message': str(e)
#         })


# # Get a list of reviews
# # @ api_view(['POST'])
# # def product_reviews(request):
# #     product_id = request.data['product_id']

# #     product_reviews = review_db.objects.filter(product__id=product_id)

# #     reviews_serialized = ReviewSerializer(product_reviews, many=True)

# #     return Response({
# #         'status': 'success',
# #         'message': 'Review posted',
# #         'data': reviews_serialized.data
# #     })


# def saveImage(img):
#     img.save('media/Product_images/' + img)
#     return 'Product_images/' + str(img)

# import ast
# from rest_framework import status

# @ api_view(['POST'])
# def upload_product(request):
#     user = request.headers['User-id']
#     main_images = request.FILES.getlist(
#         'main_image')  # main_image ['', '', '']
#     image_list = request.FILES.getlist('image')  # image ['','','']
#     jdata = dict((request.data))['data'][0]
#     #print(jdata)
#     data = json.loads(jdata)
#     #print(data['screen_no'])
#     booleanList = ['featured_products','best_selling_products','hot_selling_products']
#     try:
#         if data['screen_no'] == 1:
#             booleanList = ['featured_products','best_selling_products','hot_selling_products']
#             specialproduct = random.choice(booleanList)
#             if data['varients'] == []:
#                 user_obj = userDB.objects.get(user_id=user)
#                 category_obj = productCategory.objects.get(
#                     name=data['category'])
#                 data[specialproduct] = True
#                 product_obj = product.objects.create(
#                     username=user_obj,
#                     category_id=category_obj,
#                     featured_products = data['featured_products'],
#                     best_selling_products =data['best_selling_products'],
#                     hot_selling_products = data['hot_selling_products'],
#                     sub_category=data['sub_category'],
#                     name=data['name'],
#                     keyword=str(data['keyword']).translate(
#                         {ord(c): None for c in "'[]"}),
#                     details=str(data['details']).translate(
#                         {ord(c): None for c in "'[]"}),
#                     size=str(data['size']).translate(
#                         {ord(c): None for c in "'[]"}),
#                     color=str(data['color']).translate(
#                         {ord(c): None for c in "'[]"}),
#                     main_image=main_images[0],
#                     page_status=data['screen_no'],
#                 )

#                 product_obj.save()

#                 product_obj.brand=data['brand']
#                 product_obj.save()
#                 for img in image_list:
#                     productImages.objects.create(
#                         product=product_obj, image=img)
#                 return Response({
#                     'status': status.HTTP_200_OK,
#                     'message': 'product added',
#                     'data': product_obj.id
#                 })

#             else:
#                 user_obj = userDB.objects.get(user_id=user)
#                 category_obj = productCategory.objects.get(
#                     name=data['category'])
#                 data[specialproduct] = True
#                 product_obj = product.objects.create(
#                     username=user_obj,
#                     category_id=category_obj,
#                     featured_products = data['featured_products'],
#                     best_selling_products =data['best_selling_products'],
#                     hot_selling_products = data['hot_selling_products'],
#                     sub_category=data['sub_category'],
#                     name=data['name'],
#                     brand=data['brand'],
#                     keyword=str(data['keyword']).translate(
#                         {ord(c): None for c in "'[]"}),
#                     details=str(data['details']).translate(
#                         {ord(c): None for c in "'[]"}),
#                     varients=True,
#                     page_status=data['screen_no'],
#                 )
#                 product_obj.save()
#                 print(data)
#                 for i in data['varients']:
#                     color = i['color']
#                     size = str(i['size']).translate(
#                         {ord(c): None for c in "'[]"})

#                     product_obj.save()

#                     for m_img in main_images:
#                         if str(m_img) in i['main_image']:
#                             productImages.objects.create(
#                                 product=product_obj, image=m_img, main_image=True, size=size, color=color)
#                             break

#                     for img in image_list:
#                         if str(img) in i['images']:
#                             productImages.objects.create(
#                                 product=product_obj, image=img, main_image=False, size=product_obj.size, color=product_obj.color)

#                 return Response({
#                     'status':  status.HTTP_200_OK,
#                     'message': 'product added',
#                     'data': product_obj.id
#                 })

#         elif data['screen_no'] == 2:
#             product_obj = product.objects.get(id=data['product_id'])
#             product_obj.detailed_description = data['detailed_description']
#             product_obj.page_status = data['screen_no']
#             product_obj.save()
#             count = 0
#             for i in data['product_details']:
#                 product_detail_db.objects.create(
#                     product=product_obj,
#                     image=image_list[count], # ['','','']
#                     heading=i['title'],
#                     desc=i['description']
#                 )
#                 count += 1

#             return Response({
#                 'status': 'success',
#                 'message': 'product added',
#                 'data': product_obj.id
#             })

#         elif data['screen_no'] == 3:
#             product_obj = product.objects.get(id=data['product_id'])

#             if data['price_choice'] == 'Add Price':
#                 product_obj.price_choice = str(data['price_choice'])
#                 product_obj.price = data['price']
#                 product_obj.mrp = data['mrp']
#                 product_obj.sale_price = data['sale_price']
#                 product_obj.sale_startdate = data['sale_startdate']
#                 product_obj.sale_enddate = data['sale_enddate']
#                 product_obj.maximum_order_quantity = data['minimum_order_quantity']
#                 product_obj.manufacturing_time = data['manufacturing_time']
#                 product_obj.page_status = data['screen_no']
#                 product_obj.save()

#                 return Response({
#                     'status': 'success',
#                     'message': 'product added',
#                     'data': product_obj.id
#                 })

#             elif data['price_choice'] == 'Price according to quantity':
#                 product_obj.price_choice = str(data['price_choice'])
#                 product_obj.quantity_price = str(data['quantity_price']).translate(
#                     {ord(c): None for c in "'[]"})
#                 product_obj.page_status = data['screen_no']
#                 product_obj.save()

#                 return Response({
#                     'status': 'success',
#                     'message': 'product added',
#                     'data': product_obj.id
#                 })

#         elif data['screen_no'] == 4:
#             product_obj = product.objects.get(id=data['product_id'])
#             product_obj.weight = data['weight']
#             product_obj.transportation_port = data['transportation_port']
#             product_obj.dimensions=data['dimension']
#             product_obj.packing_details = data['packing_details']
#             product_obj.packing_address = data['package_address']
#             product_obj.page_status = data['screen_no']
#             product_obj.save()
#             if product_obj.varients:
#                 image_detail_obj = list(productImages.objects.filter(
#                     product=product_obj).values_list('color', 'size'))
#                 colors = []
#                 for item in set(image_detail_obj):
#                     inventory_obj = Inventory.objects.create(
#                         product=product_obj,
#                         quantity=0,
#                         color=item[0],
#                         size=item[1],
#                     )
#             else:
#                 inventory_obj = Inventory.objects.create(
#                     product=product_obj,
#                     quantity=0,
#                     color=product_obj.color,
#                     size=product_obj.size,
#                 )
#                 inventory_obj.save()

#             return Response({
#                 'status': 'success',
#                 'message': 'product added',
#                 'data': 'Product uploaded'
#             })
#     except Exception as e:
#         return Response({
#             'status': 'error',
#             'message': e,
#         })


# @ api_view(['POST'])
# def video_upload(request):
#     product_id = request.data['product_id']
#     video = request.FILES['video']
#     try:
#         product_obj = product.objects.get(id=product_id)
#         product_obj.video = video
#         product_obj.save()
#         url_VIDEO = 'https://api.negbuy.com' + product_obj.video.url
#         return Response({
#             'status': 'success',
#             'message': 'video added',
#             'data': url_VIDEO
#         })
#     except Exception as e:
#         return Response({
#             'status': 'error',
#             'message': str(e)
#         })

# import openpyxl

# # Deepa's Size API
# @api_view(['GET'])
# def size_api(request):
#     user_id = request.headers['User-id']
#     workbook = openpyxl.load_workbook('Dropdown_list.xlsx')
#     ws = workbook['Sheet1']

#     items = []
#     a = []
#     for row in ws.iter_rows(1, ):
#         row_cells = []
#         for cell in row:
#             row_cells.append(cell.value)
#         items.append(tuple(row_cells))

#     for i in range(1, len(items)):
#         s = items[i][1]
#         a.append(s)

#     return Response({
#         'status': True,
#         'message': 'Success',
#         'data': a
#     })


# @api_view(['POST'])
# def update_inventory(request):
#     data = request.data
#     inventory_obj = Inventory.objects.get(
#         product__id=data['product_id'], color=data['color'])
#     inventory_obj.quantity = data['quantity']
#     inventory_obj.size = ','.join(data['size'])
#     inventory_obj.save()

#     return Response({
#         'status': True,
#         'message': 'Success',
#     })


# @api_view(['GET'])
# def get_inventory(request):
#     user_id = request.headers['User-id']
#     user_products = product.objects.filter(
#         username__user_id=user_id).values_list('id')

#     return_list = []
#     for item in user_products:
#         print(item)
#         data_dict = {}
#         sum = 0
#         inventory_obj = list(Inventory.objects.filter(
#             product=item[0]).values('quantity'))
#         for q in inventory_obj:
#             sum = sum + q['quantity']

#         product_obj = list(product.objects.filter(
#             id=item[0]).values('name', 'price', 'id'))
#         product_obj[0]['quantity'] = sum
#         return_list.append(product_obj[0])

#     return Response({
#         'status': True,
#         'message': 'Success',
#         'data': return_list
#     })


# @api_view(['GET'])
# def db_rfq(request):
#     # user_id = request.headers['User-id']
#     all_rfq_list= list()
#     try:
#         # user_obj= userDB.objects.get(user_id=user_id)
#         all_rfq=rfq.objects.all()
#         for each_rfq in all_rfq:
#             obj= {
#                 'user': each_rfq.user_id,
#                 'requirement': each_rfq.requirement,
#                 'quantity' : each_rfq.quantity,
#                 'target_price': each_rfq.target_price
#             }
#             all_rfq_list.append(obj)
#         return Response({
#             'status': True,
#             'message': 'Success',
#             'data' : all_rfq_list
#         })

#     except Exception as e:
#         return Response({
#             'status': 'Error',
#             'message': e,
#         })


# @api_view(['POST'])
# def inventory_detail(request):
#     # try:
#     product_id = request.data['product_id']
#     inventory_list = Inventory.objects.filter(product=product_id)
#     product_obj = product.objects.get(id=product_id)

#     if product_obj.varients:
#         product_serialized = product_serializer(product_obj).data
#         inventory_serialized = InventorySerializer(
#             inventory_list, many=True).data
#         product_serialized['varients'] = inventory_serialized
#         if product_serialized['sku'] == None:
#             product_serialized['sku'] = ''
#         if product_serialized['size'] == None:
#             product_serialized['size'] = ''

#         return Response({
#             'status': True,
#             'message': 'Success',
#             'data': product_serialized
#         })
#     else:
#         product_serialized = product_serializer(product_obj).data
#         inventory_obj = list(Inventory.objects.filter(
#             product=product_id).values('size', 'quantity'))
#         product_serialized['size'] = inventory_obj[0]['size']
#         product_serialized['quantity'] = inventory_obj[0]['quantity']
#         image_obj = productImages.objects.filter(product=product_id)
#         product_images = ImageSerializer(image_obj, many=True).data
#         product_serialized['varients'] = []
#         product_serialized['images'] = product_images
#         if product_serialized['sku'] == None:
#             product_serialized['sku'] = ''
#         if product_serialized['size'] == None:
#             product_serialized['size'] = ''

#         return Response({
#             'status': True,
#             'message': 'Success',
#             'data': product_serialized
#         })


# # Search Product API - 23 June 2022

# @api_view(['GET'])
# def AllProducts(request):
#     prods = product.objects.all().order_by('id')[:]
#     all_products = [getProductObject(i) for i in prods]
#     return Response(all_products)

# @api_view(['GET'])
# def product_api(request):
#     queryset= product.objects.all()
#     prod_serialized = ProductSerializer_new(queryset, many=True)
#     data = prod_serialized.data
#     return Response(data)


# @api_view(['POST'])
# def Search(request):
#     q = request.POST['search']
#     offset =request.data['offset']
#     limit = request.data['limit']
#     now=datetime.now()
#     weekday = now.strftime("%a")
#     date = now.strftime("%b" " %d")
#     if q:
#         prods = product.objects.filter(name__icontains=q)
#         serializer = ProductSerializer(prods, many=True)
#         jj = serializer.data
#         for i in jj:
#             if sellerAnalytics.objects.filter(seller__id = i['username'], date=date).exists():
#                 usr = sellerAnalytics.objects.get(seller__id = i['username'], date=date)
#                 usr.search_count +=1
#                 usr.save()
#             else:
#                 print(i['username'])
#                 usr = userDB.objects.filter(id = i['username'])[0]
#                 rr = sellerAnalytics.objects.create(seller =usr, date=date)
#                 rr.day = weekday
#                 rr.date = date
#                 rr.search_count = 1
#                 rr.save()
#         if len(prods) <= 0:
#             return Response({
#                 'status': 'False',
#                 'message': 'Product not Found in Inventory',
#                 'data':serializer.data
#             })
#         else:
#             return Response({
#                 'data': serializer.data[int(offset):int(limit)+int(offset)],
#                 'total':len(list(serializer.data))
#                 })

# # Address Update API

# @api_view(['POST'])
# def address_update_api(request):
#     user_id = request.headers['User-id']
#     try:
#         usr = userDB.objects.get(user_id=user_id)
#         data = request.data
#         usr.address_line1 = data['address_line1']
#         usr.address_line2 = data['address_line2']
#         usr.city = data['city']
#         usr.state = data['state']
#         usr.postal_code = data['postal_code']
#         usr.country = data['country']
#         usr.countrycode = data['countrycode']
#         usr.save()
#         serializer = UserSerializer(usr)
#         return Response({'status':'Updated',
#         'message':'The Address is Updated!!!',
#         'data': serializer.data})

#     except:
#         return Response({
#             'status': 'Error',
#             'message': 'User Not Found',
#         })

# # Password Update API

# @api_view(['POST'])
# def password_update_api(request):
#     user_id = request.headers['User-id']
#     all_passwords = []
#     try:
#         usr = userDB.objects.get(user_id=user_id, role = 'Seller')
#         data = request.data
#         all_passwords.append(usr.password)
#         existing_pass = data['existing_password']
#         new_pass = data['new_password']
#         for every_password in all_passwords:
#             if new_pass == every_password: # edge case 1
#                 return Response({
#                     'status': 'Error',
#                     'message': "The password has been used before. Please enter a different password"
#                 })
#         if usr.password == existing_pass and new_pass not in all_passwords: # edge case 2
#             usr.password = new_pass
#         else:
#             return Response({
#                 'status': 'Error',
#                 'message': 'The existing password does not match the current password'
#                 })
#         all_passwords.append(usr.password)
#         print(all_passwords)
#         usr.save()
#         serializer = UserSerializer(usr)
#         return Response({
#             'status': 'True',
#             'message': 'Password Changed Successfully',
#             'data':serializer.data
#         })
#     except Exception:
#         return Response({
#             'status': 'Error',
#             'message': 'User Not Found',
#         })

# # Updating The Email API

# @api_view(['POST'])
# def email_update_api(request):
#     user_id = request.headers['User-id']
#     user = userDB.objects.get(user_id=user_id)
#     data = request.data
#     new_email = data['new_email']
#     if user.email == new_email:
#         return Response({
#         'status': 'Error',
#         'message': 'The email is already registered. Enter a different Email'
#     })
#     current_site = get_current_site(request)
#     mydict = {
#     'user':user,
#     'username':user.username,
#     'domain': current_site,
#     'uidb64':urlsafe_base64_encode(force_bytes(user_id)),
#     'new_email':urlsafe_base64_encode(force_bytes(new_email)),
#     'token':generate_token.make_token(user)
#     }

#     html_template = 'email.html'
#     html_content = render_to_string(html_template,mydict)
#     text_content = strip_tags(html_content)

#     email = EmailMultiAlternatives(subject="Verify your Email",
#         body=text_content,
#         from_email= settings.EMAIL_HOST_USER,
#         to = [new_email]
#     )
#     email.attach_alternative(html_content,"text/html")
#     email.send(fail_silently=False)


#     return Response({
#             "status":"True",
#             "message":"Mail has been sent"
#         })


# @api_view(['GET','POST'])
# def email_verify_api(request,uidb64,token,new_email):
#     if request.method == 'GET':
#         # print(uidb64)
#         # encoded_uid = urlsafe_b64decode(uidb64)
#         # print(encoded_uid)
#         email = force_text(urlsafe_base64_decode(new_email))
#         uid = force_text(urlsafe_base64_decode(uidb64))
#         user = userDB.objects.get(user_id=uid)
#         if user and generate_token.check_token(user,token):
#             user.email = email
#             user.save()
#             #print(user.email)
#             serialize = UserSerializer(user)
#             return redirect('https://negbuy.com/email-verified')


# # Distance between Longitude and latitude

# @api_view(['POST','GET'])
# def lat_long_location(request):
#     user_id = request.headers['User-id']
#     data = request.data
#     productID = data['product_id']
#     usr = userDB.objects.get(user_id=user_id)
#     buyer_city = usr.city
#     buyer_country = usr.country
#     prods = product.objects.get(id=productID)
#     product_seller = prods.username.username
#     seller = userDB.objects.get(username=product_seller)
#     seller_city = seller.city
#     seller_country = seller.country
#     international = False

#     if usr.role == 'Buyer':
#         geolocator = Nominatim(user_agent="negbuyapi", timeout=10)
#         buyer_location = geolocator.geocode(buyer_city)
#         curr_lat_buyer_port=0
#         curr_long_buyer_port=0
#         min_distance_buyer =1000000
#         curr_name_buyer = ''
#         coords_1_buyer = (buyer_location.latitude, buyer_location.longitude)
#         qs_buyer = port.objects.filter(country=buyer_country).values()

#         for item_buyer in qs_buyer:
#             name_val_buyer = item_buyer['name']
#             latitude_val_buyer = item_buyer['latitude']
#             longitude_val_buyer = item_buyer['longitude']
#             coords_2_buyer = (latitude_val_buyer, longitude_val_buyer)
#             distance_buyer = geopy.distance.geodesic(coords_1_buyer, coords_2_buyer).km
#             if min_distance_buyer > distance_buyer:
#                 min_distance_buyer=distance_buyer
#                 curr_lat_buyer_port = latitude_val_buyer
#                 curr_long_buyer_port =longitude_val_buyer
#                 curr_name_buyer=name_val_buyer

#         try:
#             seller_location = geolocator.geocode(seller_city)
#             curr_name_seller = ''
#             curr_lat_seller_port=0
#             curr_long_seller_port=0
#             min_dist_seller=1000000
#             coords_1_seller = (seller_location.latitude, seller_location.longitude)
#             qs_seller = port.objects.filter(country=seller_country).values()
#             for item_seller in qs_seller:
#                 name_val_seller = item_seller['name']
#                 latitude_val_seller = item_seller['latitude']
#                 longitude_val_seller = item_seller['longitude']
#                 coords_2_seller = (latitude_val_seller, longitude_val_seller)
#                 distance_seller = geopy.distance.geodesic(coords_1_seller, coords_2_seller).km
#                 if min_dist_seller > distance_seller:
#                     min_dist_seller=distance_seller
#                     curr_lat_seller_port = latitude_val_seller
#                     curr_long_seller_port =longitude_val_seller
#                     curr_name_seller=name_val_seller
#             if usr.country.lower() != 'india':
#                 international = True

#             data_diction = {
#                 'international':international,
#                 'buyer_location': {
#                     'address':usr.city,
#                     'lat':str(buyer_location.latitude),
#                     'long': str(buyer_location.longitude)
#                 },
#                 'buyer_port': {
#                     'address':curr_name_buyer,
#                     'lat':curr_lat_buyer_port,
#                     'long': curr_long_buyer_port
#                 },
#                 'seller_location': {
#                     'address':seller.city,
#                     'lat':str(seller_location.latitude),
#                     'long': str(seller_location.longitude)
#                 },
#                 'seller_port': {
#                     'address':curr_name_seller,
#                     'lat':curr_lat_seller_port,
#                     'long': curr_long_seller_port
#                 },
#                 'seller_address':{
#                     'address':seller.address_line1 +','+ seller.address_line2,
#                     'city':seller.city,
#                     'state':seller.state,
#                     'country':seller.country,
#                     'postal_code': seller.postal_code,
#                 }
#             }
#             return Response({
#                 "status":'True',
#                 'message':'Data is returned',
#                 'data':data_diction
#             })
#         except requests.exceptions.ConnectionError:
#             return (requests.exceptions.ConnectionError)
#     else:
#         return Response({
#                 "status":'False',
#                 'message':'User Not Found'
#             })

# # Paytm

# # import uuid
# # @api_view(['POST','GET'])
# # def payment(request):
# #     if request.method == "GET":
# #         return render(request, 'pay.html')
# #     if request.method == "POST":
# #         order_id = uuid.uuid4()
# #         data = request.data
# #         user_id = data['User-id']
# #         productID = data['product_id']
# #         usr = userDB.objects.get(user_id=user_id)
# #         prods = product.objects.get(id=productID)
# #         product_name = prods.name
# #         phone= usr.phone
# #         amount = prods.price
# #         address=usr.city

# #         m_id = settings.PAYTM_MERCHANT_ID
# #         m_key = settings.PAYTM_MERCHANT_KEY

# #         transaction = Orders_Id.objects.create(ORDER_ID = order_id, User_Id = user_id,
# #                                 Mobile= phone,Address=address,TXNAMOUNT= str(amount),
# #                             Product_Name =product_name, Username = usr.username)

# #         param_dict = {
# #             'MID': m_id,
# #             'INDUSTRY_TYPE_ID': settings.PAYTM_INDUSTRY_TYPE_ID,
# #             'WEBSITE': settings.PAYTM_WEBSITE,
# #             'CHANNEL_ID': settings.PAYTM_CHANNEL_ID,
# #             'CALLBACK_URL': settings.PAYTM_CALLBACK_URL,
# #             'MOBILE_NO': str(phone),
# #             'CUST_ID': str(usr.email),
# #             'ORDER_ID':str(order_id),
# #             'TXN_AMOUNT': str(amount),
# #         }
# #         checksum = Checksum.generate_checksum(param_dict= param_dict,merchant_key= m_key)
# #         param_dict['CHECKSUMHASH'] = checksum

# #         transaction.Checksum = checksum
# #         transaction.save()
# #     context = {
# #         'payment_url': settings.PAYTM_PAYMENT_GATEWAY_URL,
# #         'comany_name': settings.PAYTM_COMPANY_NAME,
# #         'data_dict': param_dict
# #     }
# #     print(context)
# #     return render(request, 'payment.html', context)
# #     #return Response({'data': param_dict})


# @api_view(['POST','GET'])
# def payment(request):
#     if request.method == "GET":
#         return render(request, 'pay.html')
#     if request.method == "POST":
#         order_id = uuid.uuid4()
#         data = request.data
#         user_id = data['User-id']
#         productID = data['product_id']
#         usr = userDB.objects.get(user_id=user_id)
#         prods = product.objects.get(id=productID)
#         product_name = prods.name
#         phone= usr.phone
#         amount = prods.price
#         address=usr.city

#         m_id = settings.PAYTM_MERCHANT_ID
#         m_key = settings.PAYTM_MERCHANT_KEY

#         transaction = Orders_Id.objects.create(ORDER_ID = order_id, User_Id = user_id,
#                                 Mobile= phone,Address=address,TXNAMOUNT= str(amount),
#                             Product_Name =product_name, Username = usr.username)

#         paytmParams = dict()

#         paytmParams["body"] = {
#             "requestType"   : "Payment",
#             "mid"           : m_id,
#             "websiteName"   : settings.PAYTM_WEBSITE,
#             "orderId"       : str(order_id),
#             "enablePaymentMode": [
#                 {"mode" : "UPI", "channels" : ["UPIPUSH","UPI","UPIPUSHEXPRESS"]},
#                 {"mode" : "BALANCE"},
#                 {"mode" : "PAYTM_DIGITAL_CREDIT"},
#                 {"mode" : "CREDIT_CARD", "channels" : ["VISA","MASTER","AMEX"]},
#                 {"mode" : "DEBIT_CARD", "channels" : ["VISA","MASTER","AMEX"]},
#                 {"mode" : "NET_BANKING", "channels" : ["SBI","PNB","HDFC","ICICI","AIRP",
#                                         "ALH","AXIS","PYTM","YES","VJYA","UBI",
#                                 "UNI","USFB","SIB","RATN","PSB","NKMB"]},
#                 ],
#             "callbackUrl"   : "http://localhost:8000/api/callback",
#             "txnAmount"     : {
#                 "value"     : str(amount),
#                 "currency"  : "INR",
#                 },
#             "userInfo"      : {
#                 "custId"    : str(user_id),
#             },
#         }

#         # Generate checksum by parameters we have in body
#         # Find your Merchant Key in your Paytm Dashboard at https://dashboard.paytm.com/next/apikeys
#         checksum = PaytmChecksum.generateSignature(json.dumps(paytmParams["body"]), m_key)

#         paytmParams["head"] = {
#             "signature"    : checksum,
#             "channelId" : settings.PAYTM_CHANNEL_ID
#         }

#         # # for Production
#         # url = f"https://securegw.paytm.in/theia/api/v1/initiateTransaction?mid={m_id}&orderId={order_id}"

#         post_data = json.dumps(paytmParams)

#         #for Staging
#         url = f"https://securegw-stage.paytm.in/theia/api/v1/initiateTransaction?mid={m_id}&orderId={order_id}"


#         response = requests.post(url, data = post_data, headers = {"Content-type": "application/json"}).json()
#         transaction.Checksum = response['head']['signature']
#         transaction.save()
#         #print(response)
#         param_dict = {
#             'mid':m_id,
#             'txnToken': response['body']['txnToken'],
#             'orderId': paytmParams['body']['orderId'],
#             'company_name': settings.PAYTM_COMPANY_NAME
#         }
#         print(response)
#     #return ({'data': response})
#     return render(request, 'demo.html', {'data':param_dict})


# # @csrf_exempt
# # @api_view(['POST','GET'])
# # def callback(request):
# #     form= request.POST
# #     response_dict={}
# #     for i in form.keys():
# #         response_dict[i]=  form[i]
# #         if i =='CHECKSUMHASH':
# #             checksum = form[i]
# #             verify = Checksum.verify_checksum(response_dict, settings.PAYTM_MERCHANT_KEY, checksum)
# #     transaction = Orders_Id.objects.get(ORDER_ID = response_dict['ORDERID'])
# #     verify = Checksum.verify_checksum(response_dict, settings.PAYTM_MERCHANT_KEY, checksum)
# #     if verify:
# #         if response_dict['RESPCODE'] == '01':
# #             transaction.payment_mode = response_dict['PAYMENTMODE']
# #             transaction.Payment_Status = True
# #             transaction.TXNID = response_dict['TXNID']
# #             transaction.bank_transaction_id =response_dict['BANKTXNID']
# #             transaction.CURRENCY = response_dict['CURRENCY']
# #             transaction.GATEWAYNAME = response_dict['GATEWAYNAME']
# #             transaction.RESPMSG = response_dict['RESPMSG']
# #             transaction.BANKNAME = response_dict['BANKNAME']
# #             transaction.MID = response_dict['MID']
# #             transaction.STATUS = response_dict['STATUS']
# #             transaction.TXNDATE = response_dict['TXNDATE']
# #             transaction.save()
# #         else:
# #             transaction.TXNID = response_dict['TXNID']
# #             transaction.RESPMSG = response_dict['RESPMSG']
# #             transaction.MID = response_dict['MID']
# #             transaction.STATUS = response_dict['STATUS']
# #             transaction.save()

# #     return redirect(f'https://negbuy.com/transaction/{transaction.TXNID}')

# @csrf_exempt
# @api_view(['POST','GET'])
# def callback(request):
#     form= request.POST
#     response_dict={}

#     for i in form.keys():
#         response_dict[i]=  form[i]
#         #print(response_dict)
#         if i =='CHECKSUMHASH':
#             checksum = form[i]
#             verify = PaytmChecksum.verifySignature(response_dict, settings.PAYTM_MERCHANT_KEY, checksum)
#     transaction = Orders_Id.objects.get(ORDER_ID = response_dict['ORDERID'])
#     verify = Checksum.verify_checksum(response_dict, settings.PAYTM_MERCHANT_KEY, checksum)
#     if verify:
#         if response_dict['RESPCODE'] != '01':
#             transaction.TXNID = response_dict['TXNID']
#             transaction.RESPMSG = response_dict['RESPMSG']
#             transaction.MID = response_dict['MID']
#             transaction.STATUS = response_dict['STATUS']
#             transaction.CURRENCY = response_dict['CURRENCY']
#             transaction.TXNAMOUNT = response_dict['TXNAMOUNT']
#             transaction.save()
#         else:
#             #transaction.payment_mode = response_dict['PAYMENTMODE']
#             transaction.Payment_Status = True
#             transaction.TXNID = response_dict['TXNID']
#             transaction.bank_transaction_id =response_dict['BANKTXNID']
#             transaction.CURRENCY = response_dict['CURRENCY']
#             transaction.GATEWAYNAME = response_dict['GATEWAYNAME']
#             transaction.RESPMSG = response_dict['RESPMSG']
#             transaction.TXNAMOUNT = response_dict['TXNAMOUNT']
#             transaction.BANKNAME = response_dict['BANKNAME']
#             transaction.MID = response_dict['MID']
#             transaction.STATUS = response_dict['STATUS']
#             transaction.TXNDATE = response_dict['TXNDATE']
#             transaction.save()


#     #return redirect(f'https://negbuy.com/transaction/{transaction.TXNID}')
#     return render(request, 'callback.html', {'data':response_dict})

# @api_view(['POST','GET'])
# def transaction_api(request):
#     data = request.data
#     transactionID = data['transactionID']
#     queryset= Orders_Id.objects.filter(TXNID = transactionID)
#     prod_serialized = TransactionSerializer(queryset, many=True)
#     data = prod_serialized.data
#     return Response({'data':data[0]})

# # rfq
# @api_view(['POST'])
# def rfq_create_api(request):
#     user_id = request.headers['User-id']
#     try:
#         usr = userDB.objects.get(user_id=user_id)
#         requirement =request.data['requirement']
#         target_price = request.data['target_price']
#         quantity = request.data['quantity']
#         delivery_expected_date = request.data['delivery_expected_date']
#         rfq_create= rfq.objects.create(user=usr,requirement =requirement,
#                             target_price=target_price,quantity=quantity,
#                             delivery_expected_date=delivery_expected_date)
#         rfq_create.save()
#         return Response({
#             'status':'True',
#             'message':'Your RFQ is Updated. Our Team will get to you Soon !!!'
#         })
#     except Exception as e:
#         return Response({
#             'status': 'Error',
#             'message': e,
#         })

# # Review/Rating

# @api_view(['POST'])
# def reviews(request):
#     try:
#         product_id = request.data['product_id']
#         offset =request.data['offset']
#         limit = request.data['limit']
#         product_reviews = review_db.objects.filter(product__id=product_id)
#         review_serialized = ReviewNEWSerializer(product_reviews, many=True).data
#         review_list = []
#         for i in review_serialized:
#             review_list.append(i['review_title'])
#         obj = {
#             'data': review_serialized[int(offset):int(limit)+int(offset)],
#             'total':len(review_list)
#         }
#         return Response(obj)
#     except Exception as e:
#         return Response({'error_msg': str(e)}, status=500)

# @api_view(['GET','POST'])
# def product_review_stats(request):
#     try:
#         user_id = request.headers['User-id']
#         offset =request.data['offset']
#         limit = request.data['limit']
#         usr=userDB.objects.get(user_id=user_id)
#         queryset= product.objects.filter(username= usr)
#         if usr.role =='Buyer':
#             obj = {
#             'Status': 'False',
#             'Message': 'Feature is available only for Seller'
#         }
#         else:
#             prod_serialized = NEWProductSerializer(queryset, many=True)
#             for i in prod_serialized.data:
#                 product_reviews = review_db.objects.filter(product__id=i['id'])[0:2]
#                 review_serialized = ReviewNEWSerializer(product_reviews, many=True).data
#                 i['reviews'] = review_serialized
#             data = prod_serialized.data

#             obj = {
#                 'data': data[int(offset):int(limit)+int(offset)],
#                 'total': len(data)
#             }
#         return Response(obj)
#     except Exception as e:
#         return Response({'error_msg': str(e)}, status=500)

# @api_view(['GET','POST'])
# def review_stats(request):
#     product_id = request.data['product_id']
#     try:
#         prod = product.objects.get(id = product_id)
#         product_reviews = review_db.objects.filter(product=prod)
#         review_serialized = ReviewStatsSerializer(product_reviews, many=True).data
#         review_list = [0]
#         count_1 = 0
#         count_2 = 0
#         count_3 = 0
#         count_4 = 0
#         count_5 = 0
#         length_review_list = 0
#         avg_review_list = 0
#         for i in review_serialized:
#             if review_list[0] == 0:
#                 review_list[0] = int(i['rating'])
#             else:
#                 review_list.append(int(i['rating']))
#             if i['rating'] ==1:
#                 count_1+= 1
#             if i['rating'] ==2:
#                 count_2+= 1
#             if i['rating'] ==3:
#                 count_3+= 1
#             if i['rating'] ==4:
#                 count_4+= 1
#             if i['rating'] ==5:
#                 count_5+= 1
#             if review_list[0]==0:
#                 length_review_list = 0
#             else:
#                 length_review_list = len(review_list)
#             if length_review_list == 0:
#                 avg_review_list = 0
#             else:
#                 avg_review_list= round(sum(review_list)/length_review_list)
#             review_dict = {
#                 'ratings': sum(review_list),
#                 'reviews': length_review_list,
#                 'average_rating':avg_review_list,
#             }
#             rating_stats_dict = {
#                 'rating_1': count_1,
#                 'rating_2': count_2,
#                 'rating_3': count_3,
#                 'rating_4': count_4,
#                 'rating_5': count_5,
#             }

#             i['statistics'] = review_dict
#             i['rating_stats'] = rating_stats_dict

#             if i['rating'] !=0:
#                 del i['rating']
#                 return Response({"data":i})
#         return Response({"data": {
#                                 "statistics": {
#                                     "ratings": 0,
#                                     "reviews": 0,
#                                     "average_rating": 0
#                                 },
#                                 "rating_stats": {
#                                     "rating_1": 0,
#                                     "rating_2": 0,
#                                     "rating_3": 0,
#                                     "rating_4": 0,
#                                     "rating_5": 0
#                                 }}})
#     except Exception as e:
#         return Response({'error_msg': str(e)}, status=500)


# class ExportImportExcel(APIView):

#     def get(self, request):
#         product_objs = product.objects.all()
#         serializer = ProductSerializer_new(product_objs, many=True)
#         df = pd.DataFrame(serializer.data)
#         print(df)
#         df.to_csv("media/excel/new.csv", encoding = "UTF-8", index = False)
#         return Response({'status': 200})


#     def post(self, request):
#         with open('E:/Projects/latest negbuy/data/Abhinandan.csv', encoding ='unicode_escape') as csvfile:
#             reader = csv.DictReader(csvfile)
#             for row in reader:
#                 print(row['name'])
#                 prod_obj = product.objects.create(name = row['name'])
#                 user_data = userDB.objects.get(user_id = '1')
#                 #print(row['category_id'])
#                 category_data = productCategory.objects.get(name = row['category_id'])
#                 prod_obj.username = user_data
#                 prod_obj.sku = row['sku']
#                 prod_obj.category_id = category_data
#                 prod_obj.main_image = row['main_image']
#                 img = productImages.objects.create(product = prod_obj)
#                 img.image = row['main_image']
#                 img.main_image = True
#                 img.color = str(row['color']).split(',')[0]
#                 img.size = str(row['size']).split(',')[0]
#                 img.save()
#                 prod_obj.featured_products = row['featured_products']
#                 prod_obj.best_selling_products = row['best_selling_products']
#                 prod_obj.hot_selling_products=row['hot_selling_products']
#                 prod_obj.fast_dispatch=row['fast_dispatch']
#                 prod_obj.ready_to_ship=row['ready_to_ship']
#                 prod_obj.customized_product=row['customized_product']
#                 prod_obj.brand = row['brand']
#                 prod_obj.keyword = row['keyword']
#                 prod_obj.color = row['color']
#                 prod_obj.size = row['size']
#                 details_obj = str(row['details(Heading:Value)']).split(',')
#                 try:
#                     for each_value in details_obj:
#                         details = each_value.split(':')
#                         #print(details)
#                         prod_obj.details = details
#                         prod_details = product_detail_db.objects.create(product = prod_obj)
#                         prod_details.heading = details[0]
#                         prod_details.desc = details[1]
#                         prod_details.save()
#                 except:
#                     pass
#                 prod_obj.price_choice = row['price_choice']
#                 prod_obj.price = row['mrp']
#                 prod_obj.quantity_price = row['quantity_price(Qty:Price:Timetaken)']
#                 prod_obj.weight = row['weight']
#                 prod_obj.transportation_port = row['transportation_port']
#                 prod_obj.packing_details= row['packing_details']
#                 prod_obj.packing_address = row['packing_address']
#                 prod_obj.status = row['status']
#                 prod_obj.detailed_description = row['details(Heading:Value)']

#                 prod_obj.save()
#             return Response({'message':'Your Data is uploaded successfully !!'})


def privacy_policy(request):
    return render(request, "privacy-policy.html")


def refund_cancellation(request):
    return render(request, "refund-cancellation-policy.html")


def terms_of_service(request):
    return render(request, "terms-of-service.html")


def terms_of_use(request):
    return render(request, "terms.html")


def payment_flow(request):
    return render(request, "payment-flow.html")


# @ api_view(['POST'])
# def categorized_product(request):
#     try:
#         category = request.data['category']

#         product_list = product.objects.filter(category_id__name=category)
#         serializer = ProductSerializer(product_list, many=True)

#         return Response({
#             'status': True,
#             'message': 'Success',
#             'data': serializer.data,
#         })

#     except Exception as e:
#         return Response({
#             'status': 'Error',
#             'message': e,
#             'data': ''
#         })

# @api_view(['GET'])
# def news(request):
#     queryset= newsPost.objects.all()[::-1]
#     serialized_data = newsPostSerializer(queryset, many=True)
#     data = serialized_data.data
#     return Response(data[:3])

# @api_view(['GET'])
# def categories(request):
#     query =productCategory.objects.all()
#     serialized_data = CategorySerializer(query, many=True)
#     data = serialized_data.data
#     kk= []
#     for i in data:
#         kk.append(i['name'])
#     return Response(kk)

# @api_view(['GET'])
# def aboutUS(request):
#     queryset= teamSection.objects.all()
#     serialized_data = teamSerializer(queryset, many=True)
#     data = serialized_data.data
#     return Response(data)

# from django.db.models import Q
# @api_view(['POST'])
# def Searchproductlist(request):
#     q = request.POST['search']
#     limit = request.data['limit']
#     if q:
#         prods = product.objects.filter(Q(name__icontains=q) |
#                             Q(keyword__icontains ='#' + q)).order_by('?')
#         serializer = ProductSearchSerializer(prods, many=True)
#         if len(prods) <= 0:
#             return Response({
#                 'status': 'False',
#                 'message': 'Product not Found in Inventory',
#                 'data':serializer.data[:int(limit)],
#             })
#         else:
#             return Response({
#                 'data': serializer.data[:int(limit)],
#                 })
# import datetime

# @api_view(['POST'])
# def supplychain(request):
#     product_id = request.data['product_id']
#     units_sold = request.data['units_sold']
#     month = request.data['month']
#     year = request.data['year']
#     prod = product.objects.get(id =product_id)
#     product_seller = prod.username.username
#     seller = userDB.objects.get(username=product_seller)
#     revenue =float(prod.price) * int(units_sold)
#     trans_obj = transDB.objects.create(seller=seller, units_sold=units_sold)
#     trans_obj.total_revenue =revenue
#     trans_obj.month = month
#     trans_obj.save()
#     # tt = transDB.objects.filter(seller = seller, product = prod)
#     # if transDB.objects.filter(seller=seller).exists() and transDB.objects.filter(product =prod).exists():
#     return Response({
#     'month':month,
#     'year':year,
#     'revenue':revenue,
#     'units_sold':int(units_sold),
#     })

# from django.core import serializers

# @api_view(['GET','POST'])
# def seller_dashv1(request):
#     user_id = request.headers['User-id']
#     year = request.data['year']
#     month_list = [
#     {'month':'Jan',
#     'year':year,
#         'revenue':0,
# 	    'units_sold':0,
# 	    'average_revenue':0,
# 	},
#     {'month':'Feb',
#     'year':year,
#         'revenue':0,
# 	    'units_sold':0,
# 	    'average_revenue':0,
# 	},
#     {'month':'Mar',
#     'year':year,
#         'revenue':0,
# 	    'units_sold':0,
# 	    'average_revenue':0,
# 	},
#     {'month':'Apr',
#     'year':year,
#         'revenue':0,
# 	    'units_sold':0,
# 	    'average_revenue':0,
# 	},
#     {'month':'May',
#     'year':year,
#         'revenue':0,
# 	    'units_sold':0,
# 	    'average_revenue':0,
# 	},
#     {'month':'Jun',
#     'year':year,
#         'revenue':0,
# 	    'units_sold':0,
# 	    'average_revenue':0,
# 	},
#     {'month':'Jul',
#     'year':year,
#         'revenue':0,
# 	    'units_sold':0,
# 	    'average_revenue':0,
# 	},
#     {'month':'Aug',
#     'year':year,
#         'revenue':0,
# 	    'units_sold':0,
# 	    'average_revenue':0,
# 	},
#     {'month':'Sep',
#     'year':year,
#         'revenue':0,
# 	    'units_sold':0,
# 	    'average_revenue':0,
# 	},
#     {'month':'Oct',
#     'year':year,
#         'revenue':0,
# 	    'units_sold':0,
# 	    'average_revenue':0,
# 	},
#     {'month':'Nov',
#     'year':year,
#         'revenue':0,
# 	    'units_sold':0,
# 	    'average_revenue':0,
# 	},
#     {'month':'Dec',
#     'year':year,
#         'revenue':0,
# 	    'units_sold':0,
# 	    'average_revenue':0,
# 	}]
#     usr = userDB.objects.get(user_id = user_id)
#     data = serializers.serialize("json", transDB.objects.filter(seller = usr ,year__icontains = year))
#     ll =[]
#     dd = json.loads(data)
#     for i in dd:
#         ll.append(i['fields'])
#     for i in ll:
#         if i['month'] == 'Jan':
#             month_list[0].update(revenue= month_list[0].get('revenue')+float(i['total_revenue']))
#             month_list[0].update(units_sold= month_list[0].get('units_sold')+int(i['units_sold']))
#         elif i['month'] == 'Feb':
#             month_list[1].update(revenue= month_list[1].get('revenue')+float(i['total_revenue']))
#             month_list[1].update(units_sold= month_list[1].get('units_sold')+int(i['units_sold']))
#         elif i['month'] == 'Mar':
#             month_list[2].update(revenue= month_list[2].get('revenue')+float(i['total_revenue']))
#             month_list[2].update(units_sold= month_list[2].get('units_sold')+int(i['units_sold']))
#         elif i['month'] == 'Apr':
#             month_list[3].update(revenue= month_list[3].get('revenue')+float(i['total_revenue']))
#             month_list[3].update(units_sold= month_list[3].get('units_sold')+int(i['units_sold']))
#         elif i['month'] == 'May':
#             month_list[4].update(revenue= month_list[4].get('revenue')+float(i['total_revenue']))
#             month_list[4].update(units_sold= month_list[4].get('units_sold')+int(i['units_sold']))
#         elif i['month'] == 'Jun':
#             month_list[5].update(revenue= month_list[5].get('revenue')+float(i['total_revenue']))
#             month_list[5].update(units_sold= month_list[5].get('units_sold')+int(i['units_sold']))
#         elif i['month'] == 'Jul':
#             month_list[6].update(revenue= month_list[6].get('revenue')+float(i['total_revenue']))
#             month_list[6].update(units_sold= month_list[6].get('units_sold')+int(i['units_sold']))
#         elif i['month'] == 'Aug':
#             month_list[7].update(revenue= month_list[7].get('revenue')+float(i['total_revenue']))
#             month_list[7].update(units_sold= month_list[7].get('units_sold')+int(i['units_sold']))
#         elif i['month'] == 'Sep':
#             month_list[8].update(revenue= month_list[8].get('revenue')+float(i['total_revenue']))
#             month_list[8].update(units_sold= month_list[8].get('units_sold')+int(i['units_sold']))
#         elif i['month'] == 'Oct':
#             month_list[9].update(revenue= month_list[9].get('revenue')+float(i['total_revenue']))
#             month_list[9].update(units_sold= month_list[9].get('units_sold')+int(i['units_sold']))
#         elif i['month'] == 'Nov':
#             month_list[10].update(revenue= month_list[10].get('revenue')+float(i['total_revenue']))
#             month_list[10].update(units_sold= month_list[10].get('units_sold')+int(i['units_sold']))
#         elif i['month'] == "Dec":
#             month_list[11].update(revenue= month_list[11].get('revenue')+float(i['total_revenue']))
#             month_list[11].update(units_sold= month_list[11].get('units_sold')+int(i['units_sold']))

#     for i in month_list:
#         if i['units_sold'] >0:
#             i.update(average_revenue = float(float(i.get('revenue'))/int(i.get('units_sold'))))
#     return Response({'data':month_list})

# @api_view(['POST','GET'])
# def product_screen_time(request):
#     product_id = request.data['product_id']
#     time = request.data['time']
#     now=datetime.now()
#     weekday = now.strftime("%a")
#     date = now.strftime("%b" " %d")
#     year = now.strftime("%Y")
#     prod = product.objects.get(id =product_id)
#     product_seller = prod.username.username
#     usr = userDB.objects.get(username=product_seller)
#     try:
#         loi= TimeManagement.objects.get(seller=usr,prod =prod, date=date)
#         loi.time = int(loi.time) + int(time)
#         loi.save()
#     except:
#         iop = TimeManagement.objects.create(seller =usr, prod = prod, date=date, day=weekday, year=year)
#         iop.time = int(time)
#         iop.save()
#     finally:
#         return Response({'message':'Data is Saved !!'},status = status.HTTP_200_OK)

# @api_view(['POST'])
# def seller_dashv2(request):
#     user_id = request.headers['User-id']
#     month = request.data['month']
#     usr = userDB.objects.get(user_id = user_id)
#     data_year = serializers.serialize("json", transDB.objects.filter(seller = usr ,year__icontains = '2022'))
#     data = serializers.serialize("json", transDB.objects.filter(seller = usr ,month__icontains = month, year__icontains = '2022'))
#     jj = {
#         'total_products': 0,
#         'products_sold':0,
#         'products_returned':0
#     }
#     mm=[]
#     units_sold_list_total = []
#     ss = json.loads(data_year)
#     for k in ss:
#         mm.append(k['fields'])
#     for i in mm:
#         units_sold_list_total.append(int(i['units_sold']))
#     units_sold_list_total_sum = sum(units_sold_list_total)
#     ll =[]
#     units_sold_list = []
#     dd = json.loads(data)
#     for i in dd:
#         ll.append(i['fields'])
#     for i in ll:
#         units_sold_list.append(int(i['units_sold']))
#     jj.update(total_products = units_sold_list_total_sum)
#     jj.update(products_sold = sum(units_sold_list))
#     return Response({'data':jj})

# from datetime import datetime

# @api_view(['POST','GET'])
# def seller_analyticsV1(request):
#     user_id = request.headers['User-id']
#     usr = userDB.objects.get(user_id = user_id)
#     try:
#         jj = sellerAnalytics.objects.filter(seller = usr).order_by('-created_at')
#         serial = AnalyticSerializer(jj,many=True)
#         return Response({'data':serial.data[:7]})
#     except Exception as e:
#         return Response({'message':str(e)})


# @api_view(['POST','GET'])
# def seller_analyticsV2(request):
#     user_id = request.headers['User-id']
#     usr = userDB.objects.get(user_id = user_id)
#     data = serializers.serialize("json", transDB.objects.filter(seller = usr , year__icontains = '2022'))
#     oo = sellerAnalytics.objects.filter(seller = usr).order_by('-created_at')
#     serial = AnalyticSerializer(oo,many=True)
#     search_product_count = serial.data
#     sum_count = []
#     for i in search_product_count:
#         sum_count.append(i['search_count'])
#     jj = {
#         'wishlist_count':0,
#         'added_to_cart_count': 0,
#         'sold_product_count': 0,
#         'search_product_count': 0
#         }
#     ll =[]
#     units_sold_list = []
#     dd = json.loads(data)
#     for i in dd:
#         ll.append(i['fields'])
#     for i in ll:
#         units_sold_list.append(int(i['units_sold']))
#     units_sold_list_sum = sum(units_sold_list)
#     jj.update(sold_product_count = units_sold_list_sum)
#     jj.update(search_product_count = sum(sum_count))
#     return Response({'data':jj})

# from dateutil.relativedelta import relativedelta

# @api_view(['POST'])
# def seller_analyticsV3(request):
#     month_dict ={
#         'Jan':31,
#         'Feb':28,
#         'Mar':31,
#         'Apr':30,
#         'May':31,
#         'Jun':30,
#         'Jul':31,
#         'Aug':31,
#         'Sep':30,
#         'Oct':31,
#         'Nov':30,
#         'Dec':31
#     }
#     user_id = request.headers['User-id']
#     usr = userDB.objects.get(user_id = user_id)
#     now=datetime.now()
#     weekday = now.strftime("%a")
#     date1 = now.strftime("%b" " %d")
#     month_day = now.strftime("%b")
#     curr_date = now.strftime("%d")
#     last_week_date = (int(curr_date)-7)
#     past_14thDay_date = (int(curr_date)-14)
#     year = now.strftime("%Y")
#     last_month = datetime.now() - relativedelta(months=1)
#     text = format(last_month, '%b') # prints last month
#     queryset = TimeManagement.objects.filter(seller=usr).order_by('created_at')
#     serial = TimeSerializer(queryset, many=True).data
#     current_week = []
#     last_week= []
#     percent =0

#     for y in range(int(past_14thDay_date),int(last_week_date)):
#         for key, value in month_dict.items():
#             if int(curr_date) < 15:
#                 month_day = text
#                 if month_day == key:
#                     if y<=0:
#                         y = value +y
#                     #print(y)
#         for i in serial:
#             if (month_day +" " +str(y)) == i['date']:
#                 last_week.append(i['time'])
#             else:
#                 last_week.append(0)


#     for y in range(int(last_week_date),int(curr_date)):
#         for key, value in month_dict.items():
#             if int(curr_date) < 8:
#                 month_day = text
#                 if month_day == key:
#                     if y<=0:
#                         y = value +y
#                     print(y)
#         for i in serial:
#             if (month_day +" " +str(y)) == i['date']:
#                 current_week.append(i['time'])
#             else:
#                 current_week.append(0)

#     if sum(current_week) > sum(last_week) == True:
#         if sum(last_week) > 0:
#             percent += ((sum(current_week)-sum(last_week))/sum(last_week))*100
#         else:
#             percent += sum(current_week)-sum(last_week)
#     else:
#         if sum(last_week) > 0:
#             percent += ((sum(last_week)-sum(current_week))/sum(last_week))*100
#         else:
#             percent += ((sum(last_week)-sum(current_week))/100)

#     dict1 = {
#         'last_week': sum(last_week),
#         'current_week': sum(current_week),
#         'today_screen_time':0,
#         'increased': sum(current_week) > sum(last_week),
#         'percent': percent,
#         'time_in':'sec'
#     }
#     today_total = []
#     for i in serial:
#         if i['date'] == date1:
#             today_total.append(i['time'])
#         dict1.update(today_screen_time =sum(today_total))

#     return Response(dict1)


######################     ADMIN DASHBOARD    #########################

g_vars = [
    "admin_last_login",
    "admin_id",
    "old_sellers",
    "old_products",
    "old_verified_products",
    "old_un_verified_products",
]


def update(g_vars):
    with open("variables.txt", "w") as file:
        for i in g_vars:
            file.write(str(i) + "\n")


def get_variable(g_vars):
    with open("variables.txt", "r") as file:
        content = file.readlines()
        g_vars[0] = content[0]
        g_vars[1] = int(content[1])
        g_vars[2] = int(content[2])
        g_vars[3] = int(content[3])
        g_vars[4] = int(content[4])
        g_vars[5] = int(content[5])


@api_view(["POST"])
def admin_login(request):
    username = request.data["Username"]
    password = request.data["Password"]
    user = userDB.objects.filter(
        role="Admin", username=username, password=password
    ).values("username", "id")

    if user:
        global g_vars
        get_variable(g_vars)
        admin_id = user[0]["id"]
        admin_last_login = g_vars[0].split("\n")[0]
        g_vars[1] = admin_id
        g_vars[0] = admin_last_login
        update(g_vars)
        data_key = {"admin_id": admin_id}
        message = {
            "status": True,
            "data": data_key,
            "username": username,
            "msg": "login success",
        }

    else:
        message = {"status": False, "msg": "username or password did not match."}

    return Response(message)


@api_view(["POST"])
def admin_dashboard_analytics(request):
    data_key = int(request.data["admin_id"])

    global g_vars
    get_variable(g_vars)
    admin_last_login = g_vars[0].split("\n")[0]
    admin_last_login = datetime.datetime.strptime(
        admin_last_login, "%Y-%m-%d %H:%M:%S.%f%z"
    )
    admin_id = g_vars[1]
    old_sellers = g_vars[2]
    old_products = g_vars[3]
    old_verified_products = g_vars[4]
    old_un_verified_products = g_vars[5]

    if data_key == int(admin_id):
        try:
            new_sellers = len(
                userDB.objects.filter(role="Seller", created_at__gte=admin_last_login)
            )
            new_products = len(
                ProductDB.objects.filter(created_at__gte=admin_last_login)
            )
            verified_products = len(
                ProductDB.objects.filter(verification_status="verified")
            )
            un_verified_products = len(
                ProductDB.objects.filter(verification_status="under verification")
            )

            if old_sellers == 0:
                change_in_sellers = (new_sellers - old_sellers) * 100
                old_sellers = new_sellers
            else:
                change_in_sellers = round(
                    (((new_sellers - old_sellers) * 100) / old_sellers), 2
                )
                old_sellers = new_sellers

            if old_products == 0:
                change_in_products = (new_products - old_products) * 100
                old_products = new_products
            else:
                change_in_products = round(
                    (((new_products - old_products) * 100) / old_products), 2
                )
                old_products = new_products

            if old_verified_products == 0:
                change_in_verified_products = (
                    verified_products - old_verified_products
                ) * 100
                old_verified_products = verified_products
            else:
                change_in_verified_products = round(
                    (
                        ((verified_products - old_verified_products) * 100)
                        / old_verified_products
                    ),
                    2,
                )
                old_verified_products = verified_products

            if old_un_verified_products == 0:
                change_in_un_verified_products = (
                    un_verified_products - old_un_verified_products
                ) * 100
                old_un_verified_products = un_verified_products
            else:
                change_in_un_verified_products = round(
                    (
                        ((un_verified_products - old_un_verified_products) * 100)
                        / old_un_verified_products
                    ),
                    2,
                )
                old_un_verified_products = un_verified_products

            g_vars[0] = timezone.make_aware(datetime.datetime.now())
            g_vars[2] = old_sellers
            g_vars[3] = old_products
            g_vars[4] = old_verified_products
            g_vars[5] = old_un_verified_products

            update(g_vars)

            return Response(
                {
                    "status": True,
                    "new_sellers": new_sellers,
                    "new_products": new_products,
                    "verified_products": verified_products,
                    "un_verified_products": un_verified_products,
                    "change_in_sellers": f"{change_in_sellers}%",
                    "change_in_products": f"{change_in_products}%",
                    "change_in_verified_products": f"{change_in_verified_products}%",
                    "change_in_un_verified_products": f"{change_in_un_verified_products}%",
                    "admin_id": admin_id,
                }
            )
        except:
            return Response({"status": False, "msg": "please login again"})
    else:
        return Response({"msg": "non-autherised user"})


@api_view(["POST"])
def admin_dashboard_user(request):
    data_key = int(request.data["admin_id"])

    global g_vars
    get_variable(g_vars)
    admin_id = g_vars[1]

    if data_key == int(admin_id):
        try:
            un_veri_products = ProductDB.objects.filter(
                verification_status="under verification"
            ).order_by("-created_at")
            cor_sellers = []
            for i in un_veri_products:
                if i.seller_id not in cor_sellers:
                    cor_sellers.append(i.seller_id)

            if len(cor_sellers) > 4:
                cor_sellers = cor_sellers[:4]

            elif len(cor_sellers) < 4:
                sellers = userDB.objects.filter(role="Seller").order_by("-created_at")
                for k in range(len(sellers)):
                    if len(cor_sellers) == 4:
                        break
                    else:
                        if sellers[k] not in cor_sellers:
                            cor_sellers.append(sellers[k])

            output = []
            for j in cor_sellers:
                verified_products = ProductDB.objects.filter(
                    seller_id=j, verification_status="verified"
                )
                unverified_products = ProductDB.objects.filter(
                    seller_id=j, verification_status="under verification"
                )

                y = []
                for p in verified_products:
                    if p.category_id.name not in y:
                        y.append(p.category_id.name)
                for p in unverified_products:
                    if p.category_id.name not in y:
                        y.append(p.category_id.name)

                verified_products = len(verified_products)
                unverified_products = len(unverified_products)

                total = verified_products + unverified_products
                if total == 0:
                    total = 1

                if j.profile_picture:
                    profile_pic = (
                        "https://negbuy.com:8080"
                        + UserSerializer(j).data["profile_picture"]
                    )
                else:
                    profile_pic = False

                output.append(
                    {
                        "User-id": j.user_id,
                        "profile_picture": profile_pic,
                        "first_name": j.first_name,
                        "last_name": j.last_name,
                        "phone": j.phone,
                        "city": j.city,
                        "category": y,
                        "unverified_products": unverified_products,
                        "verified_products_percent": round(
                            ((verified_products / total) * 100), 2
                        ),
                        "unverified_products_percent": round(
                            ((unverified_products / total) * 100), 2
                        ),
                    }
                )

            # return Response({'status':True, 'message':'Success', 'data':output})
            return Response(
                {
                    "seller1": output[0],
                    "seller2": output[1],
                    "seller3": output[2],
                    "seller4": output[3],
                }
            )

        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def admin_seller_list(request):
    data_key = int(request.data["admin_id"])

    global g_vars
    get_variable(g_vars)
    admin_id = g_vars[1]

    if data_key == int(admin_id):
        try:
            sellers = userDB.objects.filter(role="Seller").order_by("-created_at")
            output = []
            for seller in sellers:
                s = userDB.objects.filter(id=seller.id)
                products = ProductDB.objects.filter(seller_id=seller)
                x = []
                y = []
                for p in products:
                    x.append(p.verification_status)
                    y.append(p.category_id)

                unverified_products = x.count("under verification")
                if unverified_products == 0:
                    verification_status = True
                else:
                    verification_status = False

                y = set(y)
                y = list(y)
                z = []
                cats = []

                for i in y:
                    z.append(productCategory.objects.filter(id=i.id))
                for j in range(len(z)):
                    cats.append(z[j][0].name)

                output_data = {
                    "User-id": s[0].user_id,
                    "profile_picture": "https://negbuy.com:8080"
                    + UserSerializer(s[0]).data["profile_picture"],
                    "first_name": s[0].first_name,
                    "last_name": s[0].last_name,
                    "phone": s[0].phone,
                    "city": s[0].city,
                    "category": cats,
                    "verified": verification_status,
                }
                output.append(output_data)

            output.append({"admin_id": admin_id})

            return Response(output)

        except:
            return Response({"msg": "some error occured"})

    else:
        return Response({"msg": "non-autherised user"})


@api_view(["POST"])
def admin_view_profile(request):
    data_key = int(request.data["admin_id"])
    User_id = request.headers["User-id"]

    global g_vars
    get_variable(g_vars)
    admin_id = g_vars[1]

    if data_key == int(admin_id):
        try:
            seller = userDB.objects.get(user_id=User_id)
            products_info = ProductDB.objects.filter(
                Q(seller_id=seller)
                & (
                    Q(verification_status="under verification")
                    | Q(verification_status="verified")
                )
            ).order_by("-created_at")
            total_products = len(products_info)

            y = []
            for c in products_info:
                y.append(c.category_id)

            y = set(y)
            y = list(y)
            z = []
            cats = []

            for i in y:
                z.append(productCategory.objects.filter(id=i.id))
            for j in range(len(z)):
                cats.append(z[j][0].name)

            product_list = []

            for pro in products_info:
                date = str(pro.created_at.date())
                main_image = (
                    "https://negbuy.com:8080"
                    + ProductColorVariationsSerializer(
                        ProductColorVariations.objects.get(
                            main_product_id=pro, main_variant=True
                        )
                    ).data["main_image"]
                )
                date_object = timezone.datetime.strptime(date, "%Y-%m-%d")
                uploaded_date = date_object.strftime("%d %B %Y")

                info = {
                    "product_id": pro.id,
                    "product_image": main_image,
                    "product_name": pro.product_title,
                    "category": (
                        productCategory.objects.filter(id=pro.category_id.id).values(
                            "name"
                        )
                    )[0]["name"],
                    "uploaded_on": uploaded_date,
                    "status": pro.verification_status,
                }
                product_list.append(info)

            output = {
                "user_id": seller.user_id,
                "profile_picture": "https://negbuy.com:8080"
                + UserSerializer(seller).data["profile_picture"],
                "first_name": seller.first_name,
                "last_name": seller.last_name,
                "phone": seller.phone,
                "email": seller.email,
                "address_line1": seller.address_line1,
                "address_line2": seller.address_line2,
                "city": seller.city,
                "postal_code": seller.postal_code,
                "state": seller.state,
                "category": cats,
                "total_products": total_products,
                "products": product_list,
                "admin_id": admin_id,
            }

            return Response(output)

        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def admin_view_product_details(request):
    data_key = int(request.data["admin_id"])
    product_id = int(request.data["product_id"])

    global g_vars
    get_variable(g_vars)
    admin_id = g_vars[1]

    if data_key == int(admin_id):
        try:
            product = ProductDB.objects.get(id=product_id)
            return Response(
                {
                    "status": True,
                    "message": "Success",
                    "data": {
                        "User-id": product.seller_id.user_id,
                        "username": product.seller_id.first_name
                        + " "
                        + product.seller_id.last_name
                        if product.seller_id.first_name and product.seller_id.last_name
                        else product.seller_id.first_name
                        if product.seller_id.first_name
                        else "Profile",
                        "product_id": product.id,
                        "product_title": product.product_title,
                        "main_sku_id": product.main_sku_id,
                        "category": product.category_id.name,
                        "subcategory": product.subcategory,
                        "brand": product.brand,
                        "keywords": product.keywords,
                        "address": product.packing_address,
                        "port": product.transportation_port,
                        "sale_start": product.sale_startdate,
                        "sale_end": product.sale_enddate,
                        "video": "https://negbuy.com:8080"
                        + ProductDBSerializer(product).data["video"]
                        if product.video
                        else None,
                        "detailed_desc": product.detailed_description,
                        "details": ProductDetails.objects.filter(
                            main_product_id=product
                        ).values("heading", "description"),
                        "product_img_desc": [
                            {
                                "image": "https://negbuy.com:8080"
                                + ProductImageDescriptionSerializer(i).data["image"],
                                "heading": i.heading,
                                "description": i.description,
                            }
                            for i in ProductImageDescription.objects.filter(
                                main_product_id=product
                            )
                        ],
                        "variants": [
                            {
                                "color": col.color,
                                "main_variant": col.main_variant,
                                "main_image": "https://negbuy.com:8080"
                                + ProductColorVariationsSerializer(col).data[
                                    "main_image"
                                ],
                                "extra_images": [
                                    "https://negbuy.com:8080"
                                    + ProductExtraImagesSerializer(j).data["image"]
                                    for j in ProductExtraImages.objects.filter(
                                        variant_id=col
                                    )
                                ],
                                "size_variants": [
                                    {
                                        "size": k.size,
                                        "mrp": k.mrp,
                                        "price": k.selling_price,
                                        "sale_price": k.sale_price,
                                        "weight": k.weight,
                                        "packing_details": k.packing_details,
                                        "dim_length": k.dim_length,
                                        "dim_width": k.dim_width,
                                        "dim_height": k.dim_height,
                                        "manufacturing_time": k.manufacturing_time,
                                        "max_order_quantity": k.max_order_quantity,
                                        "main_size": k.main_size,
                                        "stock": k.productinventorydb.stock
                                        if hasattr(k, "productinventorydb")
                                        else None,
                                        "sub_sku_id": k.subskuiddb.sub_sku_id
                                        if hasattr(k, "subskuiddb")
                                        else None,
                                    }
                                    for k in ProductSizeVariations.objects.filter(
                                        variant_id=col
                                    )
                                ],
                            }
                            for col in ProductColorVariations.objects.filter(
                                main_product_id=product
                            )
                        ],
                    },
                }
            )

        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def admin_product_update(request):
    data_key = int(request.data["admin_id"])
    product_id = request.data["product_id"]

    global g_vars
    get_variable(g_vars)
    admin_id = g_vars[1]

    if data_key == int(admin_id):
        try:
            product_title = request.data.get("title", None)
            category_name = request.data.get("category_name", None)
            subcategory = request.data.get("subcategory", None)
            keywords = request.data.get("keywords", None)
            packing_address = request.data.get("address", None)
            brand = request.data.get("brand", None)
            video = request.FILES.get("video", None)
            detailed_description = request.data.get("detailed_description", None)
            transportation_port = request.data.get("transportation_port", None)
            sale_startdate = request.data.get("sale_startdate", None)
            sale_enddate = request.data.get("sale_enddate", None)

            product = ProductDB.objects.get(id=product_id)
            if product_title:
                product.product_title = product_title
            if subcategory:
                product.subcategory = subcategory
            if keywords:
                product.keywords = keywords
            if packing_address:
                product.packing_address = packing_address
            if brand:
                product.brand = brand
            if detailed_description:
                product.detailed_description = detailed_description
            if transportation_port:
                product.transportation_port = transportation_port
            if sale_startdate:
                product.sale_startdate = sale_startdate
            if sale_enddate:
                product.sale_enddate = sale_enddate
            if video:
                if product.video:
                    storage, path = product.video.storage, product.video.path
                    storage.delete(path)
                product.video = video
            if category_name:
                category_id = (productCategory.objects.filter(name=category_name))[0]
                product.category_id = category_id
            product.save()

            return Response({"status": True, "message": "Success", "data": {}})

        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def admin_verify_product(request):
    data_key = int(request.data["admin_id"])
    product_id = request.data["product_id"]

    global g_vars
    get_variable(g_vars)
    admin_id = g_vars[1]

    if data_key == int(admin_id):
        try:
            prod = ProductDB.objects.get(id=product_id)
            prod.verification_status = "verified"
            prod.verify_reject_date = timezone.now()
            prod.save()
            ViewsTrackingDB.objects.create(main_product_id=prod)

            return Response({"status": True, "message": "success", "data": {}})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def admin_raise_issue(request):
    data_key = int(request.data["admin_id"])
    global g_vars
    get_variable(g_vars)
    admin_id = g_vars[1]

    if data_key == int(admin_id):
        try:
            product_id = int(request.data["product_id"])
            user_id = request.headers["User-id"]
            field_name = request.data["field_name"]
            remarks = request.data["remarks"]
            seller_id = userDB.objects.get(user_id=user_id)
            main_product_id = ProductDB.objects.get(id=product_id)
            ProductEditRemarks.objects.create(
                seller_id=seller_id,
                main_product_id=main_product_id,
                field_name=field_name,
                remarks=remarks,
            )

            return Response({"status": True, "message": "Success", "data": {}})

        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def admin_current_requests(request):
    data_key = int(request.data["admin_id"])
    global g_vars
    get_variable(g_vars)
    admin_id = g_vars[1]

    if data_key == int(admin_id):
        try:
            products = ProductDB.objects.filter(
                verification_status="under verification"
            ).order_by("-modified_at")
            verify_ready_products = []
            pending_products = []
            for product in products:
                issues = ProductEditRemarks.objects.filter(
                    main_product_id=product
                ).order_by("-modified_at")
                if issues:
                    statuses = []
                    for issue in issues:
                        statuses.append(issue.status)
                    if "Updated" in statuses:
                        col_var = ProductColorVariations.objects.get(
                            main_product_id=product, main_variant=True
                        )
                        ext_imgs = ProductExtraImages.objects.filter(
                            variant_id=col_var
                        )[:3]
                        extra_images = []
                        for ext_img in ext_imgs:
                            extra_images.append(
                                "https://negbuy.com:8080"
                                + ProductExtraImagesSerializer(ext_img).data["image"]
                            )

                        pending_products.append(
                            {
                                "verify": False,
                                "image": "https://negbuy.com:8080"
                                + ProductColorVariationsSerializer(col_var).data[
                                    "main_image"
                                ],
                                "product_id": product.id,
                                "category": product.category_id.name,
                                "extra_images": extra_images,
                                "product_name": product.product_title,
                                "sku_id": product.main_sku_id,
                                "created_on": product.created_at.strftime("%d %B %Y"),
                                "product_issues": [
                                    {
                                        "id": j.id,
                                        "field_name": j.field_name,
                                        "remarks": j.remarks,
                                        "status": j.status,
                                        "date": j.modified_at.strftime("%d %B %Y"),
                                    }
                                    if j.status == "Updated"
                                    else {
                                        "remarks": j.remarks,
                                        "status": j.status,
                                        "date": j.modified_at.strftime("%d %B %Y"),
                                    }
                                    if j.status == "Resolved"
                                    else {
                                        "remarks": j.remarks,
                                        "status": j.status,
                                        "date": j.modified_at.strftime("%d %B %Y"),
                                    }
                                    if j.status == "Unresolved"
                                    else {
                                        "remarks": j.remarks,
                                        "status": j.status,
                                        "date": j.created_at.strftime("%d %B %Y"),
                                    }
                                    for j in issues
                                ],
                                "product_status": product.verification_status,
                            }
                        )
                    elif "New" not in statuses and "Updated" not in statuses:
                        col_var = ProductColorVariations.objects.get(
                            main_product_id=product, main_variant=True
                        )
                        verify_ready_products.append(
                            {
                                "verify": True,
                                "image": "https://negbuy.com:8080"
                                + ProductColorVariationsSerializer(col_var).data[
                                    "main_image"
                                ],
                                "product_id": product.id,
                                "product_name": product.product_title,
                                "sku_id": product.main_sku_id,
                                "created_on": product.created_at.strftime("%d %B %Y"),
                                "product_issues": [
                                    {
                                        "remarks": j.remarks,
                                        "status": j.status,
                                        "date": j.modified_at.strftime("%d %B %Y"),
                                    }
                                    for j in issues
                                ],
                            }
                        )

            current_requests = []
            current_requests.extend(verify_ready_products)
            current_requests.extend(pending_products)
            return Response(
                {"status": True, "message": "success", "data": current_requests}
            )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def admin_pending_requests(request):
    data_key = int(request.data["admin_id"])
    global g_vars
    get_variable(g_vars)
    admin_id = g_vars[1]

    if data_key == int(admin_id):
        try:
            products = ProductDB.objects.filter(
                verification_status="under verification"
            ).order_by("-modified_at")
            pending_products = []
            for product in products:
                issues = ProductEditRemarks.objects.filter(
                    main_product_id=product
                ).order_by("-modified_at")
                if issues:
                    statuses = []
                    for issue in issues:
                        statuses.append(issue.status)
                    if "Updated" not in statuses and "New" in statuses:
                        col_var = ProductColorVariations.objects.get(
                            main_product_id=product, main_variant=True
                        )
                        pending_products.append(
                            {
                                "image": "https://negbuy.com:8080"
                                + ProductColorVariationsSerializer(col_var).data[
                                    "main_image"
                                ],
                                "product_id": product.id,
                                "category": product.category_id.name,
                                "product_name": product.product_title,
                                "sku_id": product.main_sku_id,
                                "created_on": product.created_at.strftime("%d %B %Y"),
                                "product_issues": [
                                    {
                                        "remarks": j.remarks,
                                        "status": j.status,
                                        "date": j.modified_at.strftime("%d %B %Y"),
                                    }
                                    for j in issues
                                ],
                                "product_status": product.verification_status,
                            }
                        )

            return Response(
                {"status": True, "message": "success", "data": pending_products}
            )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def admin_issues_history(request):
    data_key = int(request.data["admin_id"])
    global g_vars
    get_variable(g_vars)
    admin_id = g_vars[1]

    if data_key == int(admin_id):
        try:
            products = ProductDB.objects.filter(
                Q(verification_status="verified") | Q(verification_status="rejected")
            ).order_by("-verify_reject_date")
            product_history = []
            for product in products:
                issues = ProductEditRemarks.objects.filter(main_product_id=product)

                if product.verification_status == "verified":
                    if issues:
                        product_issues = [
                            {
                                "remarks": j.remarks,
                                "status": j.status,
                                "date": j.modified_at.strftime("%d %B %Y"),
                            }
                            for j in issues
                        ]
                        col_var = ProductColorVariations.objects.get(
                            main_product_id=product, main_variant=True
                        )
                        product_history.append(
                            {
                                "product_id": product.id,
                                "image": "https://negbuy.com:8080"
                                + ProductColorVariationsSerializer(col_var).data[
                                    "main_image"
                                ],
                                "product_name": product.product_title,
                                "sku_id": product.main_sku_id,
                                "created_on": product.created_at.strftime("%d %B %Y"),
                                "product_issues": product_issues,
                                "product_status": product.verification_status,
                                "verify_reject_date": product.verify_reject_date.strftime(
                                    "%d %B %Y"
                                ),
                            }
                        )
                else:
                    product_issues = [
                        {
                            "remarks": j.remarks,
                            "status": j.status,
                            "date": j.modified_at.strftime("%d %B %Y"),
                        }
                        for j in issues
                    ]
                    col_var = ProductColorVariations.objects.get(
                        main_product_id=product, main_variant=True
                    )
                    product_history.append(
                        {
                            "product_id": product.id,
                            "image": "https://negbuy.com:8080"
                            + ProductColorVariationsSerializer(col_var).data[
                                "main_image"
                            ],
                            "product_name": product.product_title,
                            "sku_id": product.main_sku_id,
                            "created_on": product.created_at.strftime("%d %B %Y"),
                            "product_issues": product_issues,
                            "product_status": product.verification_status,
                            "verify_reject_date": product.verify_reject_date.strftime(
                                "%d %B %Y"
                            ),
                        }
                    )

            return Response(
                {"status": True, "message": "success", "data": product_history}
            )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def admin_approve_updated_fields(request):
    data_key = int(request.data["admin_id"])
    global g_vars
    get_variable(g_vars)
    admin_id = g_vars[1]

    if data_key == int(admin_id):
        try:
            issue_id = int(request.data["issue_id"])
            ProductEditRemarks.objects.filter(id=issue_id).update(status="Resolved")
            return Response({"status": True, "message": "success", "data": {}})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def admin_re_raise_issue(request):
    data_key = int(request.data["admin_id"])
    global g_vars
    get_variable(g_vars)
    admin_id = g_vars[1]

    if data_key == int(admin_id):
        try:
            issue_id = int(request.data["issue_id"])
            ProductEditRemarks.objects.filter(id=issue_id).update(status="Unresolved")
            issue = ProductEditRemarks.objects.filter(id=issue_id).values(
                "main_product_id", "seller_id", "field_name", "remarks"
            )
            seller_id = userDB.objects.get(id=issue[0]["seller_id"])
            main_product_id = ProductDB.objects.get(id=issue[0]["main_product_id"])
            ProductEditRemarks.objects.create(
                main_product_id=main_product_id,
                seller_id=seller_id,
                field_name=issue[0]["field_name"],
                remarks=issue[0]["remarks"],
            )
            return Response({"status": True, "message": "success", "data": {}})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def admin_product_detailpage_issues(request):
    data_key = int(request.data["admin_id"])
    global g_vars
    get_variable(g_vars)
    admin_id = g_vars[1]

    if data_key == int(admin_id):
        try:
            product_id = int(request.data["product_id"])
            main_product_id = ProductDB.objects.get(id=product_id)
            issues = (
                ProductEditRemarks.objects.filter(main_product_id=main_product_id)
                .values("created_at", "remarks", "status")
                .order_by("-created_at")
            )

            for issue in issues:
                issue["created_at"] = issue["created_at"].strftime("%d %B %Y")

            return Response({"status": True, "message": "success", "data": issues})

        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def admin_reject_product(request):
    data_key = int(request.data["admin_id"])
    global g_vars
    get_variable(g_vars)
    admin_id = g_vars[1]

    if data_key == int(admin_id):
        try:
            product_id = int(request.data["product_id"])
            ProductDB.objects.filter(id=product_id).update(
                verification_status="rejected", verify_reject_date=timezone.now()
            )
            return Response({"status": True, "message": "success", "data": {}})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def admin_create_remark_notice(request):
    data_key = int(request.data["admin_id"])
    global g_vars
    get_variable(g_vars)
    admin_id = g_vars[1]

    if data_key == int(admin_id):
        # try:
        seller_id = int(request.data["seller_id"])
        subject = request.data["subject"]
        message = request.data["message"]
        sender = "admin"
        seller_id = userDB.objects.get(role="Seller", id=seller_id)
        SellerNoticeBoard.objects.create(
            seller_id=seller_id, subject=subject, message=message, sender=sender
        )

        return Response(status=200)


@api_view(["POST"])
def admin_pending_rfqs_page(request):
    data_key = int(request.data["admin_id"])
    global g_vars
    get_variable(g_vars)
    admin_id = g_vars[1]

    if data_key == int(admin_id):
        try:
            pending_rfqs = []
            for rf in rfq.objects.filter(rfq_status="pending"):
                pending_rfqs.append(
                    {
                        "rfq_id": rf.id,
                        "product_title": rf.requirement,
                        "delivery_date": rf.delivery_expected_date,
                        "quantity": rf.quantity,
                        "price": rf.target_price,
                    }
                )

            return Response(
                {"status": True, "message": "success", "data": pending_rfqs}
            )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def admin_pending_rfq_approve(request):
    data_key = int(request.data["admin_id"])
    global g_vars
    get_variable(g_vars)
    admin_id = g_vars[1]

    if data_key == int(admin_id):
        try:
            rfq_id = int(request.data["rfq_id"])
            rf = rfq.objects.get(id=rfq_id)
            rf.rfq_status = "accepted"
            rf.save()

            return Response({"status": True, "message": "success", "data": {}})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def admin_undo_approved_rfqs(request):
    data_key = int(request.data["admin_id"])
    global g_vars
    get_variable(g_vars)
    admin_id = g_vars[1]

    if data_key == int(admin_id):
        try:
            rfq_id = int(request.data["rfq_id"])
            rf = rfq.objects.get(id=rfq_id)
            rf.rfq_status = "pending"
            rf.save()

            return Response({"status": True, "message": "success", "data": {}})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def admin_pending_rfq_reject(request):
    data_key = int(request.data["admin_id"])
    global g_vars
    get_variable(g_vars)
    admin_id = g_vars[1]

    if data_key == int(admin_id):
        try:
            rfq_id = int(request.data["rfq_id"])
            rf = rfq.objects.get(id=rfq_id)
            rf.rfq_status = "rejected"
            reason = request.data.get("reason", None)
            if reason:
                rf.reason = reason
            rf.save()

            return Response({"status": True, "message": "success", "data": {}})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


########## Search API ###########
@api_view(["POST"])
def test_product_search(request):
    search_query = request.data["search_query"]

    import re

    search_query = re.sub(r"[^a-zA-Z0-9\s]", "", search_query)
    pattern = rf"\b{search_query}\b"

    results = ProductDB.objects.filter(
        Q(product_title__iregex=pattern)
        | Q(keywords__iregex=pattern)
        | Q(detailed_description__iregex=pattern)
    ).values("id", "product_title", "keywords", "detailed_description")

    return Response({"status": True, "data": results})


########## Search Suggestions API ###########
@api_view(["POST"])
def test_search_suggest(request):
    search_query = request.data["search_query"]

    import re

    search_query = re.sub(r"[^a-zA-Z0-9\s]", "", search_query)
    pattern = rf"\b{search_query}\b"

    results = ProductDB.objects.filter(Q(product_title__iregex=pattern)).values(
        "product_title", "category_id__name"
    )

    return Response({"status": True, "data": results})
    # search_query = request.data["search_query"]
    # search_quer = re.sub(r"[^a-zA-Z0-9\s]", "", search_query)
    # search_quer = "#" + search_quer
    # print(search_quer)
    # pattern = rf"^{search_quer}.*"
    # results = ProductDB.objects.filter(keywords__iregex=pattern).values("keywords")
    # output = []
    # for i in results:
    #    output.extend(i["keywords"].split(","))

    # pattern = re.compile(rf"{search_query}", re.IGNORECASE)

    # matches = []

    # for j in output:
    #    match = pattern.search(j)

    #    if match:
    #        matches.append(j)

    # return Response({"msg": True, "keys": matches})


#########################  Add a product  #################


@api_view(["POST"])
def seller_add_update_ProductDB_object(request):
    user_id = request.headers["User-id"]
    if userDB.objects.filter(user_id=user_id):
        product_id = request.data.get("product_id", None)
        product_title = request.data.get("title", None)
        category_name = request.data.get("category_name", None)
        subcategory = request.data.get("subcategory", None)
        keywords = request.data.get("keywords", None)
        packing_address = request.data.get("address", None)
        brand = request.data.get("brand", None)
        video = request.FILES.get("video", None)
        detailed_description = request.data.get("detailed_description", None)
        transportation_port = request.data.get("transportation_port", None)
        price_choice = request.data.get("price_choice", None)
        sale_startdate = request.data.get("sale_startdate", None)
        sale_enddate = request.data.get("sale_enddate", None)
        gst_data = request.data.get("gst", None)

        seller_id = userDB.objects.get(user_id=user_id)
        response_obj = {}
        if product_id:
            try:
                product = ProductDB.objects.get(id=product_id)
                if product_title:
                    product.product_title = product_title
                if subcategory:
                    product.subcategory = subcategory
                if keywords:
                    product.keywords = str(keywords).replace("[", "").replace("]", "")
                if packing_address:
                    product.packing_address = packing_address
                if brand:
                    product.brand = brand
                if detailed_description:
                    product.detailed_description = detailed_description
                if transportation_port:
                    product.transportation_port = transportation_port
                if price_choice:
                    product.price_choice = price_choice
                if sale_startdate:
                    product.sale_startdate = sale_startdate
                if sale_enddate:
                    product.sale_enddate = sale_enddate
                if gst_data:
                    product.gst = gst_data
                if video:
                    if product.video:
                        storage, path = product.video.storage, product.video.path
                        storage.delete(path)
                    product.video = video
                if category_name:
                    category_id = (productCategory.objects.filter(name=category_name))[
                        0
                    ]
                    product.category_id = category_id

                product.verification_status = "under verification"
                product.save()

                api_logger("seller add a product", user_id)

                response_obj["product_id"] = product_id
                response_obj["title"] = product_title
                response_obj["owner_name"] = str(seller_id.first_name) + str(
                    seller_id.last_name
                )
                response_obj["status"] = "under verification"

                return Response(
                    {"status": True, "message": "success", "data": response_obj}
                )

            except Exception as e:
                return Response({"status": "Error", "message": str(e), "data": {}})

        else:
            try:
                category_id = productCategory.objects.get(name=category_name)
                productdb_obj = ProductDB.objects.create(
                    seller_id=seller_id,
                    product_title=product_title,
                    category_id=category_id,
                    subcategory=subcategory,
                    keywords=str(keywords).replace("[", "").replace("]", ""),
                    packing_address=packing_address,
                    brand=brand,
                    video=video,
                    detailed_description=detailed_description,
                    transportation_port=transportation_port,
                    price_choice=price_choice,
                    sale_startdate=sale_startdate,
                    sale_enddate=sale_enddate,
                    gst=gst_data,
                )
                productdb_obj.save()

                response_obj["product_id"] = productdb_obj.id
                response_obj["title"] = product_title
                response_obj["owner_name"] = str(seller_id.first_name) + str(
                    seller_id.last_name
                )
                response_obj["status"] = "under verification"

                return Response(
                    {"status": True, "msg": "Success", "data": response_obj},
                    status=200,
                )
            except Exception as e:
                return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def seller_add_update_ProductDetails(request):
    user_id = request.headers["User-id"]
    if userDB.objects.filter(user_id=user_id).exists():
        try:
            product_id = int(request.data["product_id"])
            product_details = request.data.get(
                "product_details", None
            )  # get a list of dicts
            if product_details:
                product_details = list(eval(product_details))
                product_id = ProductDB.objects.get(id=product_id)
                pro_dets = ProductDetails.objects.filter(main_product_id=product_id)
                if pro_dets:
                    for pro_det in pro_dets:
                        pro_det.delete()

                for i in product_details:
                    ProductDetails.objects.create(
                        main_product_id=product_id,
                        heading=i["heading"],
                        description=i["description"],
                    )

                return Response({"status": True, "message": "success", "data": {}})
            else:
                return Response({"status": True, "message": "success", "data": {}})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def seller_add_update_ProductImageDescription_object(request):
    user_id = request.headers["User-id"]
    if userDB.objects.filter(user_id=user_id).exists():
        try:
            product_id = int(request.data["product_id"])
            instance_id = request.data.get("instance_id", None)
            image = request.FILES.get("image", None)
            heading = request.data.get("heading", None)
            description = request.data.get("description", None)
            delete = bool(request.data.get("delete", False))
            product_id = ProductDB.objects.get(id=product_id)
            if instance_id:
                inst = ProductImageDescription.objects.filter(
                    id=int(instance_id)
                ).first()
                if delete:
                    inst.delete()
                    return Response({"status": True, "msg": "Success", "data": {}})
                else:
                    if image:
                        if inst.image:
                            storage, path = inst.image.storage, inst.image.path
                            storage.delete(path)
                        inst.image = image
                        inst.save()
                    if heading:
                        inst.heading = heading
                        inst.save()
                    if description:
                        inst.description = description
                        inst.save()
                    return Response({"status": True, "msg": "Success", "data": {}})
            else:
                img_desc_obj = ProductImageDescription.objects.create(
                    main_product_id=product_id,
                    image=image,
                    heading=heading,
                    description=description,
                )
                return Response(
                    {"status": True, "message": "Success", "data": img_desc_obj.id}
                )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def seller_add_update_ProductColorVariations_object(request):
    user_id = request.headers["User-id"]
    if userDB.objects.filter(user_id=user_id).exists():
        try:
            product_id = int(request.data["product_id"])
            variant_id = request.data.get("variant_id", None)
            main_variant = bool(request.data.get("main_variant", False))
            color = request.data.get("color", None)
            main_image = request.FILES.get("main_image", None)
            delete = bool(request.data.get("delete", False))
            product_id = ProductDB.objects.get(id=product_id)

            if variant_id:
                color_variant = ProductColorVariations.objects.filter(
                    id=int(variant_id)
                ).first()
                if delete:
                    color_variant.delete()
                    return Response(
                        {
                            "variant_id": variant_id,
                            "msg": "Color Variant deleted Successfully",
                        },
                        status=200,
                    )
                else:
                    if main_image:
                        if color_variant.main_image:
                            storage, path = (
                                color_variant.main_image.storage,
                                color_variant.main_image.path,
                            )
                            storage.delete(path)
                        color_variant.main_image = main_image
                        color_variant.save()
                    if color:
                        color_variant.color = color
                        color_variant.save()

                    return Response({"status": True, "msg": "Success", "data": {}})
            else:
                col_var_obj = ProductColorVariations.objects.create(
                    main_product_id=product_id,
                    main_variant=main_variant,
                    color=color,
                    main_image=main_image,
                )
                return Response(
                    {"status": True, "msg": "Success", "data": col_var_obj.id}
                )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def seller_add_update_ProductExtraImages(request):
    user_id = request.headers["User-id"]
    if userDB.objects.filter(user_id=user_id).exists():
        try:
            variant_id = int(request.data["variant_id"])
            image_id = request.data.get("image_id", None)
            extra_image = request.FILES["extra_image"]
            variant_id = ProductColorVariations.objects.filter(id=variant_id).first()

            if image_id:
                image_id = ProductExtraImages.objects.filter(id=int(image_id)).first()
                storage, path = image_id.image.storage, image_id.image.path
                storage.delete(path)
                image_id.image = extra_image
                image_id.save()
                return Response({"status": True, "msg": "Success", "data": {}})
            else:
                extra_img_obj = ProductExtraImages.objects.create(
                    variant_id=variant_id, image=extra_image
                )
                return Response(
                    {"status": True, "msg": "Success", "data": extra_img_obj.id}
                )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def seller_add_update_ProductSizeVariations(request):
    user_id = request.headers["User-id"]
    if userDB.objects.filter(user_id=user_id).exists():
        try:
            variant_id = int(request.data["variant_id"])
            size_id = request.data.get("size_id", None)
            size = request.data.get("size", None)
            mrp = request.data.get("mrp", None)
            selling_price = request.data.get("selling_price", None)
            sale_price = request.data.get("sale_price", None)
            weight = request.data.get("weight", None)
            packing_details = request.data.get("packing_details", None)
            dim_length = request.data.get("dim_length", None)
            dim_width = request.data.get("dim_width", None)
            dim_height = request.data.get("dim_height", None)
            manufacturing_time = request.data.get("manufacturing_time", None)
            max_order_quantity = request.data.get("max_order_quantity", None)
            main_size = bool(request.data.get("main_size", False))
            stock = request.data.get("stock", None)
            delete = request.data.get("delete", False)

            variant_id = ProductColorVariations.objects.get(id=variant_id)

            response_obj = {}

            if size_id:
                size_variant = ProductSizeVariations.objects.filter(id=int(size_id))

                if delete:
                    size_variant.delete()
                    return Response({"status": True, "msg": "Success", "data": {}})
                else:
                    if size:
                        size_variant.update(size=size)
                    if mrp:
                        size_variant.update(mrp=mrp)
                    if selling_price:
                        size_variant.update(selling_price=selling_price)
                    if sale_price:
                        size_variant.update(sale_price=sale_price)
                    if weight:
                        size_variant.update(weight=weight)
                    if packing_details:
                        size_variant.update(packing_details=packing_details)
                    if dim_length:
                        size_variant.update(dim_length=dim_length)
                    if dim_width:
                        size_variant.update(dim_width=dim_width)
                    if dim_height:
                        size_variant.update(dim_height=dim_height)
                    if manufacturing_time:
                        size_variant.update(manufacturing_time=manufacturing_time)
                    if max_order_quantity:
                        size_variant.update(max_order_quantity=max_order_quantity)
                    if stock:
                        stk = ProductInventoryDB.objects.get(product_id=size_variant[0])
                        stk.stock = stock
                        stk.save()

                    response_obj["size_id"] = size_id
                    response_obj["selling_price"] = selling_price
                    return Response(
                        {"status": True, "msg": "Success", "data": response_obj}
                    )
            else:
                size_var_obj = ProductSizeVariations.objects.create(
                    variant_id=variant_id,
                    size=size,
                    mrp=mrp,
                    selling_price=selling_price,
                    sale_price=sale_price,
                    weight=weight,
                    packing_details=packing_details,
                    dim_length=dim_length,
                    dim_width=dim_width,
                    dim_height=dim_height,
                    manufacturing_time=manufacturing_time,
                    max_order_quantity=max_order_quantity,
                    main_size=main_size,
                )
                ProductInventoryDB.objects.create(product_id=size_var_obj, stock=stock)
                response_obj["size_id"] = size_var_obj.id
                response_obj["selling_price"] = selling_price
                return Response(
                    {"status": True, "msg": "Success", "data": response_obj}
                )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def seller_add_update_ProductBulkPurchaseDetails(request):
    user_id = request.headers["User-id"]
    if userDB.objects.filter(user_id=user_id).exists():
        try:
            size_id = int(request.data["size_id"])
            bulk_details = request.data.get("bulk_details", None)  # get a list of dicts
            if bulk_details:
                bulk_details = list(eval(bulk_details))
                size_id = ProductSizeVariations.objects.get(id=size_id)
                bulk_dets = ProductBulkPurchaseDetails.objects.filter(
                    product_id=size_id
                )
                if bulk_dets:
                    for bulk_det in bulk_dets:
                        bulk_det.delete()

                for i in bulk_details:
                    ProductBulkPurchaseDetails.objects.create(
                        product_id=size_id,
                        min_quantity=int(i["min"]),
                        max_quantity=int(i["max"]),
                        price=int(i["price"]),
                        manufacturing_time=int(i["manufacturing_time"]),
                    )
                return Response({"status": True, "msg": "Success", "data": {}})
            else:
                return Response({"status": True, "msg": "Success", "data": {}})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def show_category_dropdown_addnewproduct(request):
    try:
        categories = []
        for i in productCategory.objects.all():
            categories.append(i.name)

        return Response({"status": True, "msg": "Success", "data": categories})
    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["POST"])
def show_subcategory_dropdown_addnewproduct(request):
    try:
        category = request.data.get("category", None)

        if productCategory.objects.filter(name=category):
            file_path = os.path.join(
                settings.BASE_DIR, "static_files", "categories", f"{category}.txt"
            )
            with open(file_path, "r") as f:
                file_contents = f.read()

            lines = file_contents.split("\n")
            #     filtered_lines = []
            #     for line in lines:
            #         line = line.split(">")
            #         if len(line)>1 and line[1].replace(" ", "") not in filtered_lines:
            #             filtered_lines.append(line[1].replace(" ", ""))

            #     catpage[category.name]=filtered_lines
            # categories = []
            # for i in productCategory.objects.all():
            #     categories.append(i.name)

            return Response({"status": True, "msg": "Success", "data": lines})
        else:
            return Response(
                {"status": False, "msg": "Invalid Category Name", "data": lines}
            )
    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["POST"])
def show_ports_dropdown(request):
    try:
        country = request.data.get("country", None)
        ports = []
        if country:
            for i in port.objects.filter(country=country):
                ports.append(i.name)
        else:
            for i in port.objects.all():
                ports.append(i.name)

        return Response({"status": True, "msg": "Success", "data": ports})
    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["POST"])
def seller_submit_draft_product_for_verification(request):
    user_id = request.headers["User-id"]
    if userDB.objects.filter(user_id=user_id):
        try:
            product_id = int(request.data["product_id"])
            product = ProductDB.objects.get(id=product_id)
            product.verification_status = "under verification"
            product.main_sku_id = (
                "MAIN" + "-" + product.product_title[:3] + "-" + str(product.id)
            ).upper()
            product.save()

            for col_var in ProductColorVariations.objects.filter(
                main_product_id=product
            ):
                for size_var in ProductSizeVariations.objects.filter(
                    variant_id=col_var
                ):
                    SubSkuIdDB.objects.create(
                        product_id=size_var,
                        sub_sku_id=(
                            product.product_title[:3] + "-" + str(size_var.id)
                        ).upper(),
                    )

            return Response({"status": True, "msg": "Success", "data": {}})

        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


# @api_view(['POST'])
# def seller_add_new_product(request):
#     user_id1 = request.headers['User-id']
#     user_id = userDB.objects.get(user_id=user_id1).user_id
#     if user_id:
#         # try:
#             ### ProductDB ###
#             product_title = request.data.get('title', None)
#             category_name = request.data.get('category_name', None)
#             subcategory = request.data.get('subcategory', None)
#             keywords = request.data.get('keywords', None)
#             packing_address = request.data.get('address', None)
#             brand = request.data.get('brand', None)
#             video = request.FILES.get('video', None)
#             detailed_description = request.data.get('detailed_description', None)
#             transportation_port = request.data.get('transportation_port', None)
#             price_choice = request.data.get('price_choice', None)
#             sale_startdate = request.data.get('sale_startdate', None)
#             sale_enddate = request.data.get('sale_enddate', None)

#             seller_id = (userDB.objects.filter(user_id=user_id1))[0]
#             category_id = (productCategory.objects.filter(name=category_name))[0]
#             productdb_obj=ProductDB.objects.create(seller_id=seller_id, product_title=product_title, category_id=category_id, subcategory=subcategory, keywords=keywords, packing_address=packing_address, brand=brand, video=video, detailed_description=detailed_description, transportation_port=transportation_port, price_choice=price_choice, sale_startdate=sale_startdate, sale_enddate=sale_enddate)

#             ### ProductDetail ###
#             product_details = request.data.get('product_details', None)  ## get a list of dicts
#             if product_details != None:
#                 res = list(eval(product_details))
#                 for i in res:
#                     ProductDetails.objects.create(main_product_id=productdb_obj, heading=i['heading'], description=i['description'])

#             ### ProductImageDescription ###
#             img_head_desc = request.data.get('img_head_desc', None)   ## get a list of dicts of head and desc only
#             if img_head_desc != None:
#                 img_head_desc = list(eval(img_head_desc))
#                 images_img_head_desc = request.FILES.getlist('images_img_head_desc', None) ## get all images at a time
#                 if images_img_head_desc != None:
#                     for j in range(len(img_head_desc)):
#                         ProductImageDescription.objects.create(main_product_id=productdb_obj, image=images_img_head_desc[j], heading=img_head_desc[j]['heading'], description=img_head_desc[j]['description'])

#             ### ProductColorVariations ###
#             color_variants = request.data.get('color_variants', [{'color':None, 'main_variant':True}])   ## get a list of dict {color name, T/f for Main variant}
#             color_variants = list(eval(color_variants))
#             main_images = request.FILES.getlist('main_images', None)  ## sequence must be same as of color variants
#             extra_images = {}
#             for k in range(len(color_variants)):
#                 prod_color_obj=ProductColorVariations.objects.create(main_product_id=productdb_obj, color=color_variants[k]['color'], main_image=main_images[k], main_variant=bool(color_variants[k]['main_variant']))
#                 extra_images[prod_color_obj] = request.FILES.getlist(f'extra_images{k}', None)  ## get all images at a time of one color variant

#             ### ProductExtraImages ###
#             for key, value in extra_images.items():
#                     for v in value:
#                         ProductExtraImages.objects.create(variant_id=key, image=v)

#             ### ProductSizeVariations ###
#             size_variant_data = {}
#             for idx, color in enumerate(extra_images.keys()):
#                 size_variant_data[color] = request.data.get(f'size_variants_data{idx}', None)  ## list of dicts of all size variants of a color variant including stock

#             for idx1, (key, value) in enumerate(size_variant_data.items()):
#                 value=list(eval(value))
#                 for idx, i in enumerate(value):
#                     size_var_obj=ProductSizeVariations.objects.create(variant_id=key, size=i['size'], mrp=i['mrp'], selling_price=i['selling_price'], sale_price=i['sale_price'], weight=i['weight'], packing_details=i['packing_details'], dim_length=i['dim_length'], dim_width=i['dim_width'], dim_height=i['dim_height'], manufacturing_time=i['manufacturing_time'], max_order_quantity=i['max_order_quantity'])

#                     ### ProductInventoryDB ###
#                     ProductInventoryDB.objects.create(product_id=size_var_obj, stock=i['stock'])

#                     ### ProductBulkPurchaseDetails ###
#                     if productdb_obj.price_choice == 'Price according to quantity':
#                         bulk_purchase_data = {}
#                         bulk_purchase_data[f'bulk_purchase{idx1}{idx}'] = request.data[f'bulk_purchase{idx1}{idx}']  ## list of dicts of all possible prices for one size variant
#                         for i in list(eval(bulk_purchase_data[f'bulk_purchase{idx1}{idx}'])):
#                             ProductBulkPurchaseDetails.objects.create(product_id=size_var_obj, quantity=i['quantity'], price=i['price'], manufacturing_time=i['manufacturing_time'])
#                 size_var_obj.main_size=True
#                 size_var_obj.save()
#             return Response({'msg':True})

#         # except:
#         #     return Response({'msg':'some error occured'})

#     else:
#         return Response({'msg':'unauthorised_user'})


@api_view(["POST"])
def seller_get_products(request):
    user_id1 = request.headers["User-id"]
    user_obj = userDB.objects.get(user_id=user_id1)
    user_id = user_obj.user_id
    if user_id1 == user_id:
        try:
            products = ProductDB.objects.filter(seller_id=user_obj)
            product_list = []
            for i in products:
                title = i.product_title
                date = i.created_at
                product_category = i.category_id.name
                date = date.strftime("%d %B %Y")
                status = i.verification_status
                color_variants = ProductColorVariations.objects.filter(
                    main_product_id=i
                )
                print("qqqqqqqqqqqqqqqqqqqqq")
                variants = len(color_variants)
                main_image = ProductColorVariationsSerializer(color_variants[0]).data[
                    "main_image"
                ]
                sku_id = (ProductDB.objects.filter(id=i.id))[0].main_sku_id
                price = (
                    ProductSizeVariations.objects.filter(variant_id=color_variants[0])
                )[0].selling_price

                prod = {
                    "product_id": i.id,
                    "image": "https://negbuy.com:8080" + main_image,
                    "title": title,
                    "price": price,
                    "product_category": product_category,
                    "sku_id": sku_id,
                    "variants": variants,
                    "created_on": date,
                    "status": status,
                }
                product_list.append(prod)

            return Response(product_list)

        except Exception as e:
            print(product_list)
            return Response({"msg": "some error occured", "error": str(e)})

    else:
        return Response({"msg": "unauthorised user"})


@api_view(["POST"])
def seller_get_product_alldetails(request):
    user_id1 = request.headers["User-id"]
    user_obj = userDB.objects.get(user_id=user_id1)
    user_id = user_obj.user_id
    
    if user_id1 == user_id:
        try:
            product_id = request.data["product_id"]
            product = ProductDB.objects.get(id=int(product_id))
            pro_serialised = ProductDBSerializer(product)
            pricing = pro_serialised.data["price_choice"]
            product_image_description = ProductImageDescription.objects.filter(
                main_product_id=product
            )
            # print("yeahhh")
            img_desc = []
            for i in product_image_description:
                i = ProductImageDescriptionSerializer(i)
                img_desc.append(
                    {
                        "image": "https://negbuy.com:8080" + i.data["image"] if i.data["image"] else None,
                        "heading": i.data["heading"],
                        "description": i.data["description"],
                    }
                )
            # print("yeahhh")
            product_details = ProductDetails.objects.filter(
                main_product_id=product_id
            ).values("heading", "description")

            color_variants = ProductColorVariations.objects.filter(
                main_product_id=product
            )
            # print("yeahhh")
            col_vars = []
            for j in color_variants:
                extra_images = ProductExtraImages.objects.filter(variant_id=j)
                all_images = []
                for k in extra_images:
                    k = ProductExtraImagesSerializer(k)
                    all_images.append("https://negbuy.com:8080" + k.data["image"] if k.data["image"] else None)

                size_variants = ProductSizeVariations.objects.filter(variant_id=j)
                size_data = []
                for p in size_variants:
                    stock = (ProductInventoryDB.objects.get(product_id=p)).stock

                    if pricing == "Price according to quantity":
                        bulk_purchase = ProductBulkPurchaseDetails.objects.filter(
                            product_id=p
                        )
                        bulk_data = []
                        for q in bulk_purchase:
                            bulk = {
                                "quantity": q.quantity,
                                "price": q.price,
                                "manufacturing_time": q.manufacturing_time,
                            }
                            bulk_data.append(bulk)

                        size_variant = {
                            "size": p.size,
                            "mrp": p.mrp,
                            "selling_price": p.selling_price,
                            "sale_price": p.sale_price,
                            "weight": p.weight,
                            "packing_details": p.packing_details,
                            "dim_length": p.dim_length,
                            "dim_width": p.dim_width,
                            "dim_height": p.dim_height,
                            "manufacturing_time": p.manufacturing_time,
                            "max_order_quantity": p.max_order_quantity,
                            "stock": stock,
                            "bulk_purchase_details": bulk_data,
                        }

                    else:
                        size_variant = {
                            "size": p.size,
                            "mrp": p.mrp,
                            "selling_price": p.selling_price,
                            "sale_price": p.sale_price,
                            "weight": p.weight,
                            "packing_details": p.packing_details,
                            "dim_length": p.dim_length,
                            "dim_width": p.dim_width,
                            "dim_height": p.dim_height,
                            "manufacturing_time": p.manufacturing_time,
                            "max_order_quantity": p.max_order_quantity,
                            "stock": stock,
                        }

                    size_data.append(size_variant)

                j = ProductColorVariationsSerializer(j)
                 
                images = {
                    "main_image": "https://negbuy.com:8080" + j.data["main_image"] if j.data["main_image"] else None,
                    "color": j.data.get("color", "Default Color") ,
                    "extra_images": all_images,
                }

                color_variant_data = {
                    "color_images": images,
                    "size_variants": size_data,
                }

                col_vars.append(color_variant_data)

            return Response(
                {
                    "title": pro_serialised.data["product_title"],
                    "keywords": pro_serialised.data["keywords"],
                    "location": pro_serialised.data["packing_address"],
                    "price_choice": pro_serialised.data["price_choice"],
                    "sale_startdate": pro_serialised.data["sale_startdate"],
                    "sale_enddate": pro_serialised.data["sale_enddate"],
                    "video": "https://negbuy.com:8080" + pro_serialised.data["video"] if pro_serialised.data["video"] else None,
                    "transportation_port": pro_serialised.data["transportation_port"],
                    "detailed_description": pro_serialised.data["detailed_description"],
                    "product_image_description": img_desc,
                    "product_details": product_details,
                    "variants": col_vars,
                }
            )

        except Exception as e:
            return Response({"msg": "some error occured", "error": str(e)})

    else:
        return Response({"msg": "unauthorised user"})


@api_view(["POST"])
def seller_current_issues(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.get(user_id=user_id)
    seller = seller_id.user_id
    if user_id == seller:
        try:
            issues = ProductEditRemarks.objects.filter(seller_id=seller_id).values(
                "main_product_id"
            )
            pro_ids = set()
            for i in issues:
                pro_ids.add(i["main_product_id"])

            pro_issues = []
            for idx1, j in enumerate(pro_ids):
                col_var = ProductColorVariations.objects.filter(main_product_id=j)
                image = ProductColorVariationsSerializer(col_var[0])
                image = "https://negbuy.com:8080" + image.data["main_image"]
                product_title = col_var[0].main_product_id.product_title
                product_id = col_var[0].main_product_id.id
                sku_id = col_var[0].main_product_id.main_sku_id
                created_on = col_var[0].main_product_id.created_at
                created_on = created_on.strftime("%d %B %Y")

                issues = ProductEditRemarks.objects.filter(
                    main_product_id=j, status="New"
                ).values("field_name", "remarks", "status", "id")
                one_product_issues = []
                for idx, j in enumerate(issues):
                    one_product_issues.append({f"issue_{idx}": j})
                pro_issues.append(
                    {
                        "product_id": product_id,
                        "image": image,
                        "product_name": product_title,
                        "sku_id": sku_id,
                        "created_on": created_on,
                        f"product_{idx1}_issues": one_product_issues,
                    }
                )

            return Response({"issues": pro_issues})

        except:
            return Response({"msg": "Some error Occured"})

    else:
        return {"msg": "Unauthorised User"}


@api_view(["POST"])
def seller_issues_history(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.get(user_id=user_id)
    seller = seller_id.user_id
    if user_id == seller:
        try:
            issues = ProductEditRemarks.objects.filter(seller_id=seller_id).values(
                "main_product_id"
            )
            pro_ids = set()
            for i in issues:
                pro_ids.add(i["main_product_id"])

            pro_issues = []
            for idx1, j in enumerate(pro_ids):
                col_var = ProductColorVariations.objects.filter(main_product_id=j)
                product_title = col_var[0].main_product_id.product_title
                sku_id = col_var[0].main_product_id.main_sku_id
                resolved_on = None

                issues = (
                    ProductEditRemarks.objects.filter(
                        Q(main_product_id=j) & ~Q(status="New")
                    )
                    .order_by("-modified_at")
                    .values(
                        "field_name", "remarks", "created_at", "modified_at", "status"
                    )
                )
                created_on = issues[0]["created_at"]
                created_on = created_on.strftime("%d %B %Y")
                statuses = []
                one_product_issues = []
                for idx, j in enumerate(issues):
                    one_product_issues.append({f"remark_{idx}": j["remarks"]})
                    statuses.append(j["status"])

                if "Rejected" in statuses:
                    statuses = "Product Rejected"
                elif "Updated" in statuses:
                    statuses = "Under Review"
                elif "Updated" and "Unresolved" not in statuses:
                    statuses = "Issue Resolved"
                    resolved_on = issues[0]["modified_at"]
                    resolved_on = resolved_on.strftime("%d %B %Y")
                elif "Updated" and "Resolved" not in statuses:
                    statuses = "Issue Closed"
                pro_issues.append(
                    {
                        "product_name": product_title,
                        "sku_id": sku_id,
                        "requested_on": created_on,
                        "resolved_on": resolved_on,
                        "status": statuses,
                        f"product_{idx1}_issues": one_product_issues,
                    }
                )

            return Response({"issues": pro_issues})

        except:
            return Response({"msg": "Some error Occured"})

    else:
        return {"msg": "Unauthorised User"}


@api_view(["POST"])
def seller_update_remarked_field(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.get(user_id=user_id)
    seller = seller_id.user_id
    if user_id == seller:
        try:
            issue_id = int(request.data["issue_id"])
            field_name = request.data["field_name"]
            changed_value = request.data.get("changed_value", None)

            product = ProductEditRemarks.objects.get(id=issue_id)
            product_id = product.main_product_id.id
            main_product = ProductDB.objects.get(id=product_id)

            if field_name == "product_title":
                main_product.update(product_title=changed_value)
            elif field_name == "subcategory":
                main_product.update(subcategory=changed_value)
            elif field_name == "keywords":
                main_product.update(keywords=changed_value)
            elif field_name == "packing_address":
                main_product.update(packing_address=changed_value)
            elif field_name == "brand":
                main_product.update(brand=changed_value)
            elif field_name == "detailed_description":
                main_product.update(detailed_description=changed_value)
            elif field_name == "transportation_port":
                main_product.update(transportation_port=changed_value)
            elif field_name == "sale_startdate":
                main_product.update(sale_startdate=changed_value)
            elif field_name == "sale_enddate":
                main_product.update(sale_enddate=changed_value)
            elif field_name == "video":
                video = request.FILES["video"]
                main_product.update(video=video)
            elif field_name == "category_name":
                category_id = (productCategory.objects.filter(name=changed_value))[0]
                main_product.update(category_id=category_id)
            elif field_name == "product_details":
                # get a list of dicts
                product_details = request.data["product_details"]
                res = list(eval(product_details))
                ProductDetails.objects.filter(main_product_id=main_product).delete()
                for i in res:
                    ProductDetails.objects.create(
                        main_product_id=main_product,
                        heading=i["heading"],
                        description=i["description"],
                    )
            elif field_name == "img_head_desc":
                ProductImageDescription.objects.filter(
                    main_product_id=main_product
                ).delete()
                # get a list of dicts of head and desc only
                img_head_desc = request.data.get("img_head_desc")
                img_head_desc = list(eval(img_head_desc))
                images_img_head_desc = request.FILES.getlist(
                    "images_img_head_desc"
                )  # get all images at a time
                for j in range(len(img_head_desc)):
                    ProductImageDescription.objects.create(
                        main_product_id=main_product,
                        image=images_img_head_desc[j],
                        heading=img_head_desc[j]["heading"],
                        description=img_head_desc[j]["description"],
                    )

            return Response({"msg": "Data Updated Successfully"})

        except Exception as e:
            return Response({"msg": str(e)})

    else:
        return {"msg": "Unauthorised User"}


@api_view(["POST"])
def seller_inventory_details(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.get(user_id=user_id)
    seller = seller_id.user_id
    if user_id == seller:
        # try:
        products = ProductDB.objects.filter(
            seller_id=seller_id, verification_status="verified"
        )
        prdcts = []
        for idx, product in enumerate(products):
            product_name = ProductDB.objects.get(id=product.id).product_title
            col_variants = ProductColorVariations.objects.filter(
                main_product_id=product
            )
            col_vars = []

            for col_variant in col_variants:
                color = col_variant.color
                main_image = ProductColorVariationsSerializer(col_variant)
                main_image = "https://negbuy.com:8080" + main_image.data["main_image"]

                size_variants = ProductSizeVariations.objects.filter(
                    variant_id=col_variant
                )
                size_vars = []
                for size_variant in size_variants:
                    inventory_obj = ProductInventoryDB.objects.get(
                        product_id=size_variant
                    )
                    stock = inventory_obj.stock
                    price = size_variant.selling_price
                    size = size_variant.size
                    id = size_variant.id
                    max_order_quantity = size_variant.max_order_quantity

                    # ordrs = orders.objects.filter(product_id=size_variant)
                    # sold = 0
                    # for order in ordrs:
                    #    sold += int(order.order_quantity)

                    size_vars.append(
                        {
                            "id": id,
                            "stock": stock,
                            "price": price,
                            "sku_id": product.main_sku_id,
                            "size": size,
                            "sold": 0,
                            "max_order_quantity": max_order_quantity,
                        }
                    )

                col_vars.append(
                    {"color": color, "image": main_image, "size_variants": size_vars}
                )

            prdcts.append(
                {
                    # f"Product{idx}": {
                    f"Product": {
                        "product_name": product_name,
                        "color_variants": col_vars,
                    }
                }
            )

        return Response(prdcts)


@api_view(["POST"])
def seller_inventory_update(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.get(user_id=user_id)
    seller = seller_id.user_id
    if user_id == seller:
        # try:
        product_id = int(request.data["product_id"])
        price = request.data.get("price", None)
        stock = request.data.get("stock", None)
        product = ProductSizeVariations.objects.filter(id=product_id)[0]
        if price != None:
            product.selling_price = price
            product.save()
        if stock != None:
            product_inventory = ProductInventoryDB.objects.get(product_id=product)
            product_inventory.stock = stock
            product_inventory.save()

        return Response({"msg": "Updated Successfully"})


@api_view(["POST"])
def seller_allproducts_review(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.get(user_id=user_id)
    seller = seller_id.user_id
    if user_id == seller:
        # try:
        products = ProductDB.objects.filter(
            seller_id=seller_id, verification_status="verified"
        )
        all_products = []
        for product in products:
            title = product.product_title
            col_var = ProductColorVariations.objects.get(
                main_product_id=product, main_variant=True
            )
            main_image = ProductColorVariationsSerializer(col_var)
            main_image = "https://negbuy.com:8080" + main_image.data["main_image"]
            category = product.category_id.id
            category = productCategory.objects.get(id=category)
            category = category.name
            price = ProductSizeVariations.objects.get(
                variant_id=col_var, main_size=True
            )
            price = price.selling_price

            reviews = review_db.objects.filter(main_product_id=product)
            if reviews:
                num_reviews = len(reviews)

                ratings = []
                total_reviews = 0
                for review in reviews:
                    ratings.append(review.rating)
                    if review.review_title != None or review.review_description != None:
                        total_reviews += 1

                star1 = ratings.count(1)
                star2 = ratings.count(2)
                star3 = ratings.count(3)
                star4 = ratings.count(4)
                star5 = ratings.count(5)
                star1p = int((star1 / num_reviews) * 100)
                star2p = int((star2 / num_reviews) * 100)
                star3p = int((star3 / num_reviews) * 100)
                star4p = int((star4 / num_reviews) * 100)
                star5p = int((star5 / num_reviews) * 100)
                average_stars = round(
                    (star1 + 2 * star2 + 3 * star3 + 4 * star4 + 5 * star5) / 5, 1
                )
                percent_average_stars = int((average_stars / 5) * 100)
                total_ratings = num_reviews
            else:
                star1 = 0
                star2 = 0
                star3 = 0
                star4 = 0
                star5 = 0

                star1p = 0
                star2p = 0
                star3p = 0
                star4p = 0
                star5p = 0

                average_stars = 0
                total_reviews = 0
                total_ratings = 0
                percent_average_stars = 0

            all_products.append(
                {
                    "product_id": product.id,
                    "product_name": title,
                    "image": main_image,
                    "category": category,
                    "price": price,
                    "stars_out_of_5": average_stars,
                    "percent_average_stars": percent_average_stars,
                    "total_reviews": total_reviews,
                    "total_ratings": total_ratings,
                    "star1": {"percent": star1p, "number": star1},
                    "star2": {"percent": star2p, "number": star2},
                    "star3": {"percent": star3p, "number": star3},
                    "star4": {"percent": star4p, "number": star4},
                    "star5": {"percent": star5p, "number": star5},
                }
            )

        return Response(all_products)


@api_view(["POST"])
def seller_product_detailed_review(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.get(user_id=user_id)
    seller = seller_id.user_id
    if user_id == seller:
        # try:
        product_id = int(request.data["product_id"])
        product = ProductDB.objects.get(id=product_id)
        reviews = review_db.objects.filter(
            Q(main_product_id=product)
            & ~Q(review_title=None)
            & ~Q(review_description=None)
        )

        all_reviews = []
        for review in reviews:
            images = Review_Images_DB.objects.filter(review_id=review)
            files = []
            for image in images:
                file = (
                    "https://negbuy.com:8080"
                    + Review_Images_DBSerializer(image).data["file"]
                )
                files.append(file)
            all_reviews.append(
                {
                    "review_title": review.review_title,
                    "review_description": review.review_description,
                    "rating": review.rating,
                    "seller_name": review.main_product_id.seller_id.seller_name,
                    # "last_name": review.main_product_id..last_name,
                    "city": review.main_product_id.seller_id.city,
                    "date": (review.created_at).strftime("%d %B %Y"),
                    "images": files,
                }
            )

        return Response(all_reviews)


def calculate_time(diff):
    diff = diff.total_seconds()
    if diff / 60 >= 1:
        diff = diff / 60
        time = str(int(round(diff, 0))) + " " + "Minutes Ago"
        if diff / 60 >= 1:
            diff = diff / 60
            time = str(int(round(diff, 0))) + " " + "Hours Ago"
            if diff / 24 >= 1:
                diff = diff / 24
                time = str(int(round(diff, 0))) + " " + "Days Ago"
                if diff / 7 >= 1:
                    diff = diff / 7
                    time = str(int(round(diff, 0))) + " " + "Weeks Ago"
                    if diff / (52 / 12) >= 1:
                        diff = diff / (52 / 12)
                        time = str(int(round(diff, 0))) + " " + "Months Ago"
                        if diff / 12 >= 1:
                            diff = diff / 12
                            time = str(int(round(diff, 0))) + " " + "Years Ago"
    else:
        time = str(int(round(diff, 0))) + " " + "Seconds Ago"
    return time


@api_view(["POST"])
def seller_notice_board(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.get(user_id=user_id)
    seller = seller_id.user_id
    if user_id == seller:
        notices = SellerNoticeBoard.objects.filter(seller_id=seller_id)

        all_notices = []
        for notice in notices:
            send_time = notice.created_at
            current_time = timezone.now()
            diff = current_time - send_time
            time = calculate_time(diff)

            all_notices.append(
                {
                    "notice_id": notice.id,
                    "subject": notice.subject,
                    "sender": notice.sender,
                    "time_ago": time,
                    "read_status": notice.read,
                }
            )

        return Response(all_notices)


@api_view(["POST"])
def seller_read_notice(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.get(user_id=user_id)
    seller = seller_id.user_id
    if user_id == seller:
        notice_id = int(request.data["notice_id"])
        notice = SellerNoticeBoard.objects.get(id=notice_id)

        if str(seller_id.last_name) != "null":
            name = "Hi " + seller_id.first_name + " " + seller_id.last_name
        else:
            name = "Hi " + seller_id.first_name

        sender = notice.data["sender"]
        if sender == "admin":
            sender = "Administrator"
        elif sender == "negbuy":
            sender = "Team Negbuy"
        else:
            sender = "Others"

        notice.read = True
        notice.save()

        return Response(
            {
                "subject": notice.subject,
                "seller_name": name,
                "message": notice.message,
                "help": "At your help",
                "sender": sender,
            }
        )


#########        Views Tracking DB Automatic Updates via CRON JOB      ##############


# This API is called at :00 minutes of each hour
@api_view(["GET"])
def hourly_views_update(request):
    views = ViewsTrackingDB.objects.all()
    for view in views:
        hourly_views = list(eval(view.hourly_views))
        hour = (timezone.now()).strftime("%I %p")
        if len(hourly_views) <= 25:
            new_hour_data = {
                "hrs": hour,
                "browser": view.browser_views,
                "app": view.app_views,
                "screen_time": view.total_screen_time,
                "searched_clicks": view.total_searched_clicks,
                "other_clicks": view.total_other_clicks,
            }
            view.hourly_views = str(hourly_views.append(new_hour_data))
        else:
            new_hour_data = {
                "hrs": hour,
                "browser": view.browser_views,
                "app": view.app_views,
                "screen_time": view.total_screen_time,
                "searched_clicks": view.total_searched_clicks,
                "other_clicks": view.total_other_clicks,
            }
            hourly_views.pop(0)
            view.hourly_views = str(hourly_views.append(new_hour_data))
        view.save()
    return Response(True)


# this api is called in midnight every day at 12:00 AM
@api_view(["GET"])
def daily_views_update(request):
    views = ViewsTrackingDB.objects.all()

    for view in views:
        daily_views = list(eval(view.daily_views))
        weekday = (timezone.now()).strftime("%A")

        if len(daily_views) <= 8:
            new_day_data = {
                "day": weekday,
                "browser": view.browser_views,
                "app": view.app_views,
                "screen_time": view.total_screen_time,
                "searched_clicks": view.total_searched_clicks,
                "other_clicks": view.total_other_clicks,
            }
            view.daily_views = str(daily_views.append(new_day_data))
        else:
            new_day_data = {
                "day": weekday,
                "browser": view.browser_views,
                "app": view.app_views,
                "screen_time": view.total_screen_time,
                "searched_clicks": view.total_searched_clicks,
                "other_clicks": view.total_other_clicks,
            }
            daily_views.pop(0)
            view.daily_views = str(daily_views.append(new_day_data))
        view.save()


# this api is called in midnight every month at 12:00 AM
@api_view(["GET"])
def monthly_views_update(request):
    views = ViewsTrackingDB.objects.all()

    for view in views:
        monthly_views = list(eval(view.monthly_views))
        month = (timezone.now()).strftime("%B")

        if len(monthly_views) <= 13:
            new_day_data = {
                "month": month,
                "browser": view.browser_views,
                "app": view.app_views,
                "screen_time": view.total_screen_time,
                "searched_clicks": view.total_searched_clicks,
                "other_clicks": view.total_other_clicks,
            }
            view.monthly_views = str(monthly_views.append(new_day_data))
        else:
            new_day_data = {
                "month": month,
                "browser": view.browser_views,
                "app": view.app_views,
                "screen_time": view.total_screen_time,
                "searched_clicks": view.total_searched_clicks,
                "other_clicks": view.total_other_clicks,
            }
            monthly_views.pop(0)
            view.monthly_views = str(monthly_views.append(new_day_data))
        view.save()
    return Response(True)


# this api is called on Sunday at 2:00 AM night
@api_view(["GET"])
def last_week_data_collection(request):
    views = ViewsTrackingDB.objects.all()

    for view in views:
        weekly_views = list(eval(view.daily_views))
        total = 0
        for i in range(len(weekly_views) - 1):
            total += weekly_views[i + 1]["screen_time"] - weekly_views[i]["screen_time"]
        view.last_week_screen_time = total
        view.save()
    return Response(True)


#######################  Seller Dashboard  #############


@api_view(["GET"])
def seller_dashboard_navbar(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            products = ProductDB.objects.filter(
                seller_id=seller_id[0], verification_status="verified"
            )
            total_sales = 0
            total_views = 0
            new_reviews = 0

            for product in products:
                browser_views = product.viewstrackingdb.browser_views
                app_views = product.viewstrackingdb.app_views
                views = browser_views + app_views
                total_views += views

                reviews = len(
                    review_db.objects.filter(
                        main_product_id=product,
                        created_at__gte=(timezone.now()).replace(hour=0, minute=0),
                    )
                )
                new_reviews += reviews

            total_rfqs = len(
                rfq.objects.filter(rfq_status="accepted", status="ongoing")
            )

            ddp_orders = DDP_Orders_DB.objects.filter(
                ddp_seller_id=seller_id[0], buyer_status="Delivered"
            )
            exwork_orders = Exwork_Orders_DB.objects.filter(
                exwork_seller_id=seller_id[0], buyer_status="Picked Up"
            )

            for ddp_order in ddp_orders:
                ddp_price = ddp_order.product_price
                ddp_quantity = ddp_order.quantity
                ddp_total = float(ddp_price) * ddp_quantity
                total_sales += ddp_total

            for exwork_order in exwork_orders:
                ex_price = exwork_order.product_price
                ex_quantity = exwork_order.quantity
                ex_total = float(ex_price) * ex_quantity
                total_sales += ex_total

            api_logger("seller dashboard api", user_id)

            return Response(
                {
                    "status": True,
                    "msg": "Success",
                    "data": {
                        "total_products": len(products),
                        "total_sales": total_sales,
                        "total_views": total_views,
                        "new_reviews": new_reviews,
                        "total_rfqs": total_rfqs,
                    },
                }
            )

        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_dashboard_show_rfqs(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            rfqs = rfq.objects.filter(Q(rfq_status="accepted") & Q(status="ongoing"))
            rfqs_data = []
            for rf in rfqs:
                seller_rfq_reply = SellerRfqReply.objects.filter(rfq_id=rf)
                if seller_rfq_reply:
                    reply_status = True
                else:
                    reply_status = False
                rfqs_data.append(
                    {
                        "rfq_id": rf.id,
                        "product": rf.requirement,
                        "target_price": rf.target_price,
                        "quantity": rf.quantity,
                        "seller_reply": reply_status,
                    }
                )

            return Response({"status": True, "msg": "Success", "data": rfqs_data})

        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def seller_dashboard_upload_quotation(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            rfq_id = int(request.data["rfq_id"])
            quotation = request.FILES["quotation"]
            rfq_id = rfq.objects.get(id=rfq_id)
            SellerRfqReply.objects.create(
                seller_id=seller_id[0], rfq_id=rfq_id, file=quotation
            )

            return Response({"status": True, "msg": "Success", "data": {}})

        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_dashboard_active_buyers_daily(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            products = ProductDB.objects.filter(
                seller_id=seller_id[0], verification_status="verified"
            )
            last_24_hours_data = []
            hour_names = []
            for i in range(24):
                hour_names.append(
                    (timezone.localtime() - timezone.timedelta(hours=i)).strftime(
                        "%I %p - %d %B"
                    )
                )

            for i in hour_names:
                last_24_hours_data.append(
                    {"hour": i, "browser": 0, "app": 0, "total": 0}
                )

            for product in products:
                view_data = ViewsTrackingDB.objects.get(main_product_id=product)
                hourly = list(eval(view_data.hourly_views))[::-1]
                for j in range(len(hourly) - 1):
                    last_24_hours_data[j]["browser"] += (
                        hourly[j]["browser"] - hourly[j + 1]["browser"]
                    )
                    last_24_hours_data[j]["app"] += (
                        hourly[j]["app"] - hourly[j + 1]["app"]
                    )
                    last_24_hours_data[j]["total"] = (
                        last_24_hours_data[j]["browser"] + last_24_hours_data[j]["app"]
                    )

            last_24_hours_data.reverse()

            return Response(
                {"status": True, "msg": "Success", "data": last_24_hours_data}
            )

        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_dashboard_visit_ratio_weekly(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            products = ProductDB.objects.filter(
                seller_id=seller_id[0], verification_status="verified"
            )
            last_7_days_data = []
            day_names = []
            for i in range(1, 8):
                day_names.append(
                    (timezone.localtime() - timezone.timedelta(days=i)).strftime(
                        "%A - %d %B"
                    )
                )

            for i in day_names:
                last_7_days_data.append(
                    {
                        "day": i,
                        "browser": 0,
                        "app": 0,
                    }
                )

            for product in products:
                view_data = ViewsTrackingDB.objects.get(main_product_id=product)
                daily = list(eval(view_data.daily_views))[::-1]
                for j in range(len(daily) - 1):
                    last_7_days_data[j]["browser"] += (
                        daily[j]["browser"] - daily[j + 1]["browser"]
                    )
                    last_7_days_data[j]["app"] += daily[j]["app"] - daily[j + 1]["app"]

            last_7_days_data.reverse()

            return Response(
                {"status": True, "message": "Success", "data": last_7_days_data}
            )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_dashboard_user_traffic_monthly(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            products = ProductDB.objects.filter(
                seller_id=seller_id[0], verification_status="verified"
            )
            last_12_months_data = []
            today = timezone.localdate()
            month_names = []
            for i in range(12):
                date = today - timezone.timedelta(days=365 - i * 30)
                month_name = date.strftime("%B %Y")
                month_names.append(month_name)
            month_names.reverse()

            for i in month_names:
                last_12_months_data.append(
                    {"month": i, "browser": 0, "app": 0, "total": 0}
                )

            for product in products:
                view_data = ViewsTrackingDB.objects.get(main_product_id=product)
                daily = list(eval(view_data.monthly_views))[::-1]
                for j in range(len(daily) - 1):
                    last_12_months_data[j]["browser"] += (
                        daily[j]["browser"] - daily[j + 1]["browser"]
                    )
                    last_12_months_data[j]["app"] += (
                        daily[j]["app"] - daily[j + 1]["app"]
                    )
                    last_12_months_data[j]["total"] = (
                        last_12_months_data[j]["browser"]
                        + last_12_months_data[j]["app"]
                    )

            last_12_months_data.reverse()

            return Response(
                {"status": True, "message": "Success", "data": last_12_months_data}
            )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_dashboard_total_screen_time(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            products = ProductDB.objects.filter(
                seller_id=seller_id[0], verification_status="verified"
            )
            screen_time = 0
            for product in products:
                pro_views = ViewsTrackingDB.objects.get(main_product_id=product)
                screen_time += pro_views.total_screen_time

            return Response({"status": True, "message": "Success", "data": screen_time})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_dashboard_messages_today(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            products = ProductDB.objects.filter(seller_id=seller_id[0])
            messages = []
            for product in products:
                reviews = review_db.objects.filter(
                    main_product_id=product,
                    created_at__gte=(
                        timezone.localtime().replace(hour=0, minute=0)
                        - timezone.timedelta(days=1)
                    ),
                )
                for review in reviews:
                    messages.append(
                        {
                            "review_id": review.id,
                            "time": timezone.localtime(review.created_at).strftime(
                                "%I %p"
                            ),
                        }
                    )

            notices = SellerNoticeBoard.objects.filter(
                seller_id=seller_id[0],
                read=False,
                created_at__gte=(
                    timezone.localtime().replace(hour=0, minute=0)
                    - timezone.timedelta(days=1)
                ),
            )
            for notice in notices:
                messages.append(
                    {
                        "notice_id": notice.id,
                        "heading": notice.subject,
                        "message": notice.message,
                        "icon": notice.sender,
                        "time": timezone.localtime(notice.created_at).strftime("%I %p"),
                    }
                )

            final_msgs_data = sorted(messages, key=lambda x: x["time"])

            return Response(
                {"status": True, "message": "Success", "data": final_msgs_data}
            )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_dashboard_messages_week(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            products = ProductDB.objects.filter(seller_id=seller_id[0])
            messages = []
            for product in products:
                reviews = review_db.objects.filter(
                    main_product_id=product,
                    created_at__gte=(
                        timezone.localtime().replace(hour=0, minute=0)
                        - timezone.timedelta(days=7)
                    ),
                )
                for review in reviews:
                    messages.append(
                        {
                            "review_id": review.id,
                            "time": timezone.localtime(review.created_at).strftime(
                                "%I %p - %d %B"
                            ),
                        }
                    )

            notices = SellerNoticeBoard.objects.filter(
                seller_id=seller_id[0],
                read=False,
                created_at__gte=(
                    timezone.localtime().replace(hour=0, minute=0)
                    - timezone.timedelta(days=7)
                ),
            )
            for notice in notices:
                messages.append(
                    {
                        "notice_id": notice.id,
                        "heading": notice.subject,
                        "message": notice.message,
                        "icon": notice.sender,
                        "time": timezone.localtime(notice.created_at).strftime("%I %p"),
                    }
                )

            final_msgs_data = sorted(messages, key=lambda x: x["time"])

            return Response(
                {"status": True, "message": "Success", "data": final_msgs_data}
            )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_dashboard_messages_month(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            products = ProductDB.objects.filter(seller_id=seller_id[0])
            messages = []
            for product in products:
                reviews = review_db.objects.filter(
                    main_product_id=product,
                    created_at__gte=(
                        timezone.localtime() - timezone.timedelta(days=30)
                    ),
                )
                for review in reviews:
                    messages.append(
                        {
                            "review_id": review.id,
                            "time": timezone.localtime(review.created_at).strftime(
                                "%I %p - %d %B"
                            ),
                        }
                    )

            notices = SellerNoticeBoard.objects.filter(
                seller_id=seller_id[0],
                read=False,
                created_at__gte=(timezone.localtime() - timezone.timedelta(days=30)),
            )
            for notice in notices:
                messages.append(
                    {
                        "notice_id": notice.id,
                        "heading": notice.subject,
                        "message": notice.message,
                        "icon": notice.sender,
                        "time": timezone.localtime(notice.created_at).strftime("%I %p"),
                    }
                )

            final_msgs_data = sorted(messages, key=lambda x: x["time"])

            return Response(
                {"status": True, "message": "Success", "data": final_msgs_data}
            )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_dashboard_sales_acc_to_category(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            ddp_orders = DDP_Orders_DB.objects.filter(
                ddp_seller_id=seller_id[0], buyer_status="Delivered"
            )
            exwork_orders = Exwork_Orders_DB.objects.filter(
                exwork_seller_id=seller_id[0], buyer_status="Picked Up"
            )
            category = []
            for ddp_order in ddp_orders:
                category.append(
                    ddp_order.product_id.variant_id.main_product_id.category_id.id
                )
            for exwork_order in exwork_orders:
                category.append(
                    exwork_order.product_id.variant_id.main_product_id.category_id.id
                )

            cats = list(set(category))
            total_sold = 0
            pie_data = []
            for cat in cats:
                sold = category.count(cat)
                total_sold += sold
                cat_name = productCategory.objects.get(id=cat).name
                pie_data.append({"name": cat_name, "sold": sold})

            for pie in pie_data:
                pie["sold"] = round((pie["sold"] / total_sold) * 100, 2)

            return Response({"status": True, "message": "Success", "data": pie_data})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_dashboard_customer_satisfaction(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            products = ProductDB.objects.filter(
                seller_id=seller_id[0], verification_status="verified"
            )
            pro_ratings = []
            for product in products:
                reviews = review_db.objects.filter(main_product_id=product)
                for review in reviews:
                    pro_ratings.append(review.rating)

            total_ratings = len(pro_ratings)
            negative_ratings = pro_ratings.count(1)
            positive_ratings = total_ratings - negative_ratings
            try:
                positive_percent = round(((positive_ratings / total_ratings) * 100), 1)
                negative_percent = round(((negative_ratings / total_ratings) * 100), 1)
            except:
                positive_percent = 0
                negative_percent = 0

            return Response(
                {
                    "status": True,
                    "message": "Success",
                    "data": {
                        "positive_rating_percentage": positive_percent,
                        "negative_rating_percentage": negative_percent,
                        "total_ratings": total_ratings,
                        "positive_ratings": positive_ratings,
                        "negative_ratings": negative_ratings,
                    },
                }
            )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def seller_dashboard_supply_chain_overview(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            yr = int(request.data["year"])
            year_data = []
            for i in range(1, 13):
                total_revenue = 0
                total_ddp = 0
                total_exwork = 0
                total_sold = 0
                num_days = calendar.monthrange(yr, i)[1]
                start_time = (timezone.localtime()).replace(
                    year=yr, month=i, day=1, hour=0, minute=0, second=0, microsecond=0
                )
                end_time = (timezone.localtime()).replace(
                    year=yr,
                    month=i,
                    day=num_days,
                    hour=23,
                    minute=59,
                    second=59,
                    microsecond=999999,
                )
                ddp_orders = DDP_Orders_DB.objects.filter(
                    Q(ddp_seller_id=seller_id[0])
                    & Q(buyer_status="Delivered")
                    & Q(created_at__gte=start_time)
                    & Q(created_at__lte=end_time)
                )
                exwork_orders = Exwork_Orders_DB.objects.filter(
                    Q(exwork_seller_id=seller_id[0])
                    & Q(buyer_status="Picked Up")
                    & Q(created_at__gte=start_time)
                    & Q(created_at__lte=end_time)
                )

                for ddp_order in ddp_orders:
                    price = ddp_order.product_price
                    quantity = ddp_order.quantity
                    total = price * quantity
                    total_revenue += total
                    total_sold += quantity
                    total_ddp += quantity

                for exwork_order in exwork_orders:
                    price = exwork_order.product_price
                    quantity = exwork_order.quantity
                    total = price * quantity
                    total_revenue += total
                    total_sold += quantity
                    total_exwork += quantity

                month_name = {
                    1: "January",
                    2: "February",
                    3: "March",
                    4: "April",
                    5: "May",
                    6: "June",
                    7: "July",
                    8: "August",
                    9: "September",
                    10: "October",
                    11: "November",
                    12: "December",
                }[i]

                av_revenues = []

                sellers = userDB.objects.filter(role="Seller")
                for seller in sellers:
                    dp_orders = DDP_Orders_DB.objects.filter(
                        Q(ddp_seller_id=seller)
                        & Q(buyer_status="Delivered")
                        & Q(created_at__gte=start_time)
                        & Q(created_at__lte=end_time)
                    )
                    xwork_orders = Exwork_Orders_DB.objects.filter(
                        Q(exwork_seller_id=seller)
                        & Q(buyer_status="Picked Up")
                        & Q(created_at__gte=start_time)
                        & Q(created_at__lte=end_time)
                    )
                    tot_rev = 0

                    for dp_order in dp_orders:
                        price = dp_order.product_price
                        quantity = dp_order.quantity
                        total = price * quantity
                        tot_rev += total

                    for xwork_order in xwork_orders:
                        price = xwork_order.product_price
                        quantity = xwork_order.quantity
                        total = price * quantity
                        tot_rev += total

                    av_revenues.append(tot_rev)

                average_revenue = sum(av_revenues) / len(av_revenues)

                year_data.append(
                    {
                        "month": month_name,
                        "units_sold": total_sold,
                        "ddp_units": total_ddp,
                        "exwork_units": total_exwork,
                        "revenue": total_revenue,
                        "average_revenue": average_revenue,
                    }
                )

            return Response({"status": True, "message": "Success", "data": year_data})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def seller_dashboard_summary(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            month = request.data["month"]
            month_num = {
                "January": 1,
                "February": 2,
                "March": 3,
                "April": 4,
                "May": 5,
                "June": 6,
                "July": 7,
                "August": 8,
                "September": 9,
                "October": 10,
                "November": 11,
                "December": 12,
            }[month]

            total_sold = 0
            diff = int(
                ((((timezone.localtime()).month) - month_num) * 30.5)
                + (timezone.localtime()).day
            )
            ddp_orders = DDP_Orders_DB.objects.filter(
                Q(ddp_seller_id=seller_id[0])
                & Q(buyer_status="Delivered")
                & Q(
                    created_at__gte=(
                        (timezone.localtime()).replace(
                            hour=0, minute=0, second=0, microsecond=0
                        )
                        - timezone.timedelta(days=diff)
                    )
                )
                & Q(
                    created_at__lte=(
                        (timezone.localtime()).replace(
                            hour=0, minute=0, second=0, microsecond=0
                        )
                        - timezone.timedelta(days=(diff - 30))
                    )
                )
            )
            for ddp_order in ddp_orders:
                total_sold += ddp_order.quantity
            exwork_orders = Exwork_Orders_DB.objects.filter(
                Q(exwork_seller_id=seller_id[0])
                & Q(buyer_status="Picked Up")
                & Q(
                    created_at__gte=(
                        (timezone.localtime()).replace(
                            hour=0, minute=0, second=0, microsecond=0
                        )
                        - timezone.timedelta(days=diff)
                    )
                )
                & Q(
                    created_at__lte=(
                        (timezone.localtime()).replace(
                            hour=0, minute=0, second=0, microsecond=0
                        )
                        - timezone.timedelta(days=(diff - 30))
                    )
                )
            )
            for exwork_order in exwork_orders:
                total_sold += exwork_order.quantity
            total_products = len(
                ProductDB.objects.filter(
                    seller_id=seller_id[0], verification_status="verified"
                )
            )

            return Response(
                {
                    "status": True,
                    "message": "Success",
                    "data": {
                        "total_products": total_products,
                        "total_sold": total_sold,
                        "returned": 0,
                    },
                }
            )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_profile_view(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            bank_details = bankDetail.objects.filter(user=seller_id[0])
            if bank_details:
                bank_details = bankDetail.objects.get(user=seller_id[0])
            else:
                bank_details = bankDetail.objects.create(user=seller_id[0])

            if seller_id[0].profile_picture:
                profile_pic = (
                    "https://negbuy.com:8080"
                    + UserSerializer(seller_id[0]).data["profile_picture"]
                )
            else:
                profile_pic = False

            if seller_id[0].document_verification:
                govt_id = (
                    "https://negbuy.com:8080"
                    + UserSerializer(seller_id[0]).data["document_verification"]
                )
            else:
                govt_id = False

            profile_data = {
                "profile_pic": profile_pic,
                "name": str(seller_id[0].seller_name),
                "email": seller_id[0].email,
                "phone": seller_id[0].phone,
                "PINCODE": seller_id[0].postal_code,
                "city": seller_id[0].city,
                "state": seller_id[0].state,
                "country": seller_id[0].country,
                "address_line_1": seller_id[0].address_line1,
                "address_line_2": seller_id[0].address_line2,
                "bank_account_holder_name": bank_details.accountName,
                "bank_account_number": bank_details.accountNumber,
                "bank_ifsc_code": bank_details.accountIfsc,
                "gst_number": seller_id[0].gst_number,
                "identification_id": govt_id,
            }
            return Response(
                {"status": True, "message": "Success", "data": profile_data}
            )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def seller_profile_update(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            name = request.data.get("name", None)
            profile_pic = request.FILES.get("profile_pic", None)
            email = request.data.get("email", None)
            phone = request.data.get("phone", None)
            pincode = request.data.get("pincode", None)
            city = request.data.get("city", None)
            state = request.data.get("state", None)
            country = request.data.get("country", None)
            address_line_1 = request.data.get("address_line_1", None)
            address_line_2 = request.data.get("address_line_2", None)
            bank_account_holder_name = request.data.get(
                "bank_account_holder_name", None
            )
            bank_account_number = request.data.get("bank_account_number", None)
            bank_ifsc_code = request.data.get("bank_ifsc_code", None)
            gst_number = request.data.get("gst_number", None)
            identification_id = request.FILES.get("identification_id", None)

            bank_details = bankDetail.objects.get(user=seller_id[0])

            if name:
                seller_id[0].seller_name = name
            if profile_pic:
                seller_id[0].profile_picture = profile_pic
            if email:
                seller_id[0].email = email
            if phone:
                seller_id[0].phone = phone
            if pincode:
                seller_id[0].postal_code = pincode
            if city:
                seller_id[0].city = city
            if state:
                seller_id[0].state = state
            if country:
                seller_id[0].country = country
            if address_line_1:
                seller_id[0].address_line1 = address_line_1
            if address_line_2:
                seller_id[0].address_line2 = address_line_2
            if gst_number:
                seller_id[0].gst_number = gst_number
            if identification_id:
                seller_id[0].document_verification = identification_id
            seller_id[0].save()

            if bank_account_holder_name:
                bank_details.accountName = bank_account_holder_name
            if bank_account_number:
                bank_details.accountNumber = bank_account_number
            if bank_ifsc_code:
                bank_details.accountIfsc = bank_ifsc_code
            bank_details.save()

            return Response({"status": True, "message": "Success", "data": {}})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_all_orders(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            ddp_orders = DDP_Orders_DB.objects.filter(ddp_seller_id=seller_id[0])
            orders = []
            if ddp_orders:
                for ddp_order in ddp_orders:
                    orders.append(
                        {
                            "product_image": "https://negbuy.com:8080"
                            + ProductColorVariationsSerializer(
                                ddp_order.product_id.variant_id
                            ).data["main_image"],
                            "product_title": ddp_order.product_id.variant_id.main_product_id.product_title,
                            "order_id": ddp_order.order_id,
                            "quantity": ddp_order.quantity,
                            "order_date": ddp_order.created_at.strftime("%d %B %Y"),
                            "delivery_option": "DDP",
                            "status": ddp_order.seller_status,
                        }
                    )

            exwork_orders = Exwork_Orders_DB.objects.filter(
                Q(exwork_seller_id=seller_id[0]) & ~Q(seller_status="Pending")
            )
            for exwork_order in exwork_orders:
                orders.append(
                    {
                        "product_image": "https://negbuy.com:8080"
                        + ProductColorVariationsSerializer(
                            exwork_order.product_id.variant_id
                        ).data["main_image"],
                        "product_title": exwork_order.product_id.variant_id.main_product_id.product_title,
                        "order_id": exwork_order.order_id,
                        "quantity": exwork_order.quantity,
                        "order_date": exwork_order.created_at.strftime("%d %B %Y"),
                        "delivery_option": "ExWork",
                        "status": exwork_order.seller_status,
                    }
                )

            orders = sorted(orders, key=lambda x: x["order_date"], reverse=True)
            return Response({"status": True, "message": "Success", "data": orders})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_exwork_pending_orders(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            exwork_orders = Exwork_Orders_DB.objects.filter(
                exwork_seller_id=seller_id[0], seller_status="Pending"
            ).order_by("-created_at")
            orders = []
            for exwork_order in exwork_orders:
                orders.append(
                    {
                        "product_image": "https://negbuy.com:8080"
                        + ProductColorVariationsSerializer(
                            exwork_order.product_id.variant_id
                        ).data["main_image"],
                        "product_title": exwork_order.product_id.variant_id.main_product_id.product_title,
                        "order_id": exwork_order.order_id,
                        "quantity": exwork_order.quantity,
                        "order_date": exwork_order.created_at.strftime("%d %B %Y"),
                        "delivery_option": "ExWork",
                        "status": exwork_order.seller_status,
                    }
                )

            return Response({"status": True, "message": "Success", "data": orders})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_unshipped_orders(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            exwork_orders = Exwork_Orders_DB.objects.filter(
                Q(exwork_seller_id=seller_id[0]) & Q(seller_status="Schedule Pick Up")
                | Q(seller_status="Unshipped")
            )
            ddp_orders = DDP_Orders_DB.objects.filter(
                Q(ddp_seller_id=seller_id[0]) & Q(seller_status="Schedule Pick Up")
                | Q(seller_status="Unshipped")
            )
            orders = []
            for ddp_order in ddp_orders:
                orders.append(
                    {
                        "product_image": "https://negbuy.com:8080"
                        + ProductColorVariationsSerializer(
                            ddp_order.product_id.variant_id
                        ).data["main_image"],
                        "product_title": ddp_order.product_id.variant_id.main_product_id.product_title,
                        "order_id": ddp_order.order_id,
                        "quantity": ddp_order.quantity,
                        "order_date": ddp_order.created_at.strftime("%d %B %Y"),
                        "delivery_option": "DDP",
                        "status": ddp_order.seller_status,
                    }
                )

            for exwork_order in exwork_orders:
                orders.append(
                    {
                        "product_image": "https://negbuy.com:8080"
                        + ProductColorVariationsSerializer(
                            exwork_order.product_id.variant_id
                        ).data["main_image"],
                        "product_title": exwork_order.product_id.variant_id.main_product_id.product_title,
                        "order_id": exwork_order.order_id,
                        "quantity": exwork_order.quantity,
                        "order_date": exwork_order.created_at.strftime("%d %B %Y"),
                        "delivery_option": "Exwork",
                        "status": exwork_order.seller_status,
                    }
                )

            orders = sorted(orders, key=lambda x: x["order_date"], reverse=True)
            return Response({"status": True, "message": "Success", "data": orders})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_shipped_orders(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            exwork_orders = Exwork_Orders_DB.objects.filter(
                Q(exwork_seller_id=seller_id[0]) & Q(seller_status="Shipped")
            )
            ddp_orders = DDP_Orders_DB.objects.filter(
                Q(ddp_seller_id=seller_id[0])
                & (
                    Q(seller_status="Shipped")
                    | Q(seller_status="Shipped - Arrived at Negbuy")
                    | Q(seller_status="Shipped - Shipped from Negbuy")
                )
            )
            orders = []
            for ddp_order in ddp_orders:
                orders.append(
                    {
                        "product_image": "https://negbuy.com:8080"
                        + ProductColorVariationsSerializer(
                            ddp_order.product_id.variant_id
                        ).data["main_image"],
                        "product_title": ddp_order.product_id.variant_id.main_product_id.product_title,
                        "order_id": ddp_order.order_id,
                        "quantity": ddp_order.quantity,
                        "order_date": ddp_order.created_at.strftime("%d %B %Y"),
                        "delivery_option": "DDP",
                        "shipped_on": ddp_order.seller_shipped_on.strftime(
                            "%I %p  %d %B %Y"
                        ),
                        "status": ddp_order.seller_status,
                    }
                )

            for exwork_order in exwork_orders:
                orders.append(
                    {
                        "product_image": "https://negbuy.com:8080"
                        + ProductColorVariationsSerializer(
                            exwork_order.product_id.variant_id
                        ).data["main_image"],
                        "product_title": exwork_order.product_id.variant_id.main_product_id.product_title,
                        "order_id": exwork_order.order_id,
                        "quantity": exwork_order.quantity,
                        "order_date": exwork_order.created_at.strftime("%d %B %Y"),
                        "delivery_option": "Exwork",
                        "shipped_on": exwork_order.seller_shipped_on.strftime(
                            "%I %p  %d %B %Y"
                        ),
                        "status": exwork_order.seller_status,
                    }
                )

            orders = sorted(orders, key=lambda x: x["order_date"], reverse=True)

            return Response({"status": True, "message": "Success", "data": orders})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_delivered_orders(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            exwork_orders = Exwork_Orders_DB.objects.filter(
                Q(exwork_seller_id=seller_id[0]) & Q(seller_status="Delivered")
            )
            ddp_orders = DDP_Orders_DB.objects.filter(
                Q(ddp_seller_id=seller_id[0]) & Q(seller_status="Delivered")
            )
            orders = []
            for ddp_order in ddp_orders:
                orders.append(
                    {
                        "product_image": "https://negbuy.com:8080"
                        + ProductColorVariationsSerializer(
                            ddp_order.product_id.variant_id
                        ).data["main_image"],
                        "product_title": ddp_order.product_id.variant_id.main_product_id.product_title,
                        "order_id": ddp_order.order_id,
                        "quantity": ddp_order.quantity,
                        "order_date": ddp_order.created_at.strftime("%d %B %Y"),
                        "delivery_option": "DDP",
                        "delivered_on": ddp_order.modified_at.strftime(
                            "%I %p  %d %B %Y"
                        ),
                        "status": ddp_order.seller_status,
                    }
                )

            for exwork_order in exwork_orders:
                orders.append(
                    {
                        "product_image": "https://negbuy.com:8080"
                        + ProductColorVariationsSerializer(
                            exwork_order.product_id.variant_id
                        ).data["main_image"],
                        "product_title": exwork_order.product_id.variant_id.main_product_id.product_title,
                        "order_id": exwork_order.order_id,
                        "quantity": exwork_order.quantity,
                        "order_date": exwork_order.created_at.strftime("%d %B %Y"),
                        "delivery_option": "Exwork",
                        "delivered_on": exwork_order.modified_at.strftime(
                            "%I %p  %d %B %Y"
                        ),
                        "status": exwork_order.seller_status,
                    }
                )

            orders = sorted(orders, key=lambda x: x["order_date"], reverse=True)

            return Response({"status": True, "message": "Success", "data": orders})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_cancelled_orders(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            exwork_orders = Exwork_Orders_DB.objects.filter(
                Q(exwork_seller_id=seller_id[0]) & Q(seller_status="Cancelled")
            )
            ddp_orders = DDP_Orders_DB.objects.filter(
                Q(ddp_seller_id=seller_id[0]) & Q(seller_status="Cancelled")
            )
            orders = []
            for ddp_order in ddp_orders:
                orders.append(
                    {
                        "product_image": "https://negbuy.com:8080"
                        + ProductColorVariationsSerializer(
                            ddp_order.product_id.variant_id
                        ).data["main_image"],
                        "product_title": ddp_order.product_id.variant_id.main_product_id.product_title,
                        "order_id": ddp_order.order_id,
                        "quantity": ddp_order.quantity,
                        "order_date": ddp_order.created_at.strftime("%d %B %Y"),
                        "delivery_option": "DDP",
                        "cancelled_on": ddp_order.modified_at.strftime(
                            "%I %p  %d %B %Y"
                        ),
                        "status": ddp_order.seller_status,
                    }
                )

            for exwork_order in exwork_orders:
                orders.append(
                    {
                        "product_image": "https://negbuy.com:8080"
                        + ProductColorVariationsSerializer(
                            exwork_order.product_id.variant_id
                        ).data["main_image"],
                        "product_title": exwork_order.product_id.variant_id.main_product_id.product_title,
                        "order_id": exwork_order.order_id,
                        "quantity": exwork_order.quantity,
                        "order_date": exwork_order.created_at.strftime("%d %B %Y"),
                        "delivery_option": "Exwork",
                        "cancelled_on": exwork_order.modified_at.strftime(
                            "%I %p  %d %B %Y"
                        ),
                        "status": exwork_order.seller_status,
                    }
                )

            orders = sorted(orders, key=lambda x: x["order_date"], reverse=True)

            return Response({"status": True, "message": "Success", "data": orders})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def seller_schedule_pick_up_page(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            order_id = request.data["order_id"]
            if order_id.split("-")[0] == "DDP":
                order = DDP_Orders_DB.objects.get(id=int(order_id.split("-")[1]))
                delivery_type = "DDP"
            else:
                order = Exwork_Orders_DB.objects.get(id=int(order_id.split("-")[1]))
                delivery_type = "Ex-Work"

            return Response(
                {
                    "status": True,
                    "message": "Success",
                    "data": {
                        "order_id": order.order_id,
                        "delivery_type": delivery_type,
                        "shipping_warehouse": order.product_id.variant_id.main_product_id.packing_address,
                        "Negbuy_Warehouse": Negbuy_Warehouse.objects.first().address,
                        "product_image": "https://negbuy.com:8080"
                        + ProductColorVariationsSerializer(
                            order.product_id.variant_id
                        ).data["main_image"],
                        "product_name": order.product_id.variant_id.main_product_id.product_title,
                        "sku_id": order.product_id.subskuiddb.sub_sku_id,
                        "size_variant": order.product_id.size,
                        "color_variant": order.product_id.variant_id.color,
                        "price": order.product_price,
                        "quantity": order.quantity,
                    },
                }
            )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def seller_schedule_pick_up(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            order_id = request.data["order_id"]
            package_weight = request.data.get("package_weight", None)
            dim_lenght = request.data.get("dim_lenght", None)
            dim_width = request.data.get("dim_width", None)
            dim_height = request.data.get("dim_height", None)
            pick_up_date_time = request.data.get("pick_up_date_time", None)

            if order_id.split("-")[0] == "DDP":
                order = DDP_Orders_DB.objects.get(id=int(order_id.split("-")[1]))
                schedule = Pick_Up_Schedules.objects.filter(ddp_order_id=order)
            else:
                order = Exwork_Orders_DB.objects.get(id=int(order_id.split("-")[1]))
                schedule = Pick_Up_Schedules.objects.filter(exwork_order_id=order)

            if schedule:
                if package_weight:
                    schedule[0].weight = package_weight
                if dim_lenght:
                    schedule[0].dim_lenght = dim_lenght
                if dim_width:
                    schedule[0].dim_width = dim_width
                if dim_height:
                    schedule[0].dim_height = dim_height
                if pick_up_date_time:
                    date1 = pick_up_date_time.split("-")
                    time1 = (timezone.now()).replace(
                        hour=int(date1[3]),
                        minute=int(date1[4]),
                        day=int(date1[0]),
                        month=int(date1[1]),
                        year=int(date1[2]),
                    )
                    time1 = time1 - datetime.timedelta(hours=5, minutes=30)
                    time2 = (timezone.now()).replace(
                        hour=int(date1[5]),
                        minute=int(date1[6]),
                        day=int(date1[0]),
                        month=int(date1[1]),
                        year=int(date1[2]),
                    )
                    time2 = time2 - datetime.timedelta(hours=5, minutes=30)
                    schedule[0].pick_up_time1 = time1
                    schedule[0].pick_up_time2 = time2
            else:
                date1 = pick_up_date_time.split("-")
                time1 = (timezone.now()).replace(
                    hour=int(date1[3]),
                    minute=int(date1[4]),
                    day=int(date1[0]),
                    month=int(date1[1]),
                    year=int(date1[2]),
                )
                time1 = time1 - datetime.timedelta(hours=5, minutes=30)
                time2 = (timezone.now()).replace(
                    hour=int(date1[5]),
                    minute=int(date1[6]),
                    day=int(date1[0]),
                    month=int(date1[1]),
                    year=int(date1[2]),
                )
                time2 = time2 - datetime.timedelta(hours=5, minutes=30)
                if order_id.split("-")[0] == "DDP":
                    Pick_Up_Schedules.objects.create(
                        ddp_order_id=order,
                        weight=package_weight,
                        dim_lenght=dim_lenght,
                        dim_width=dim_width,
                        dim_height=dim_height,
                        pick_up_time1=time1,
                        pick_up_time2=time2,
                    )
                else:
                    Pick_Up_Schedules.objects.create(
                        exwork_order_id=order,
                        weight=package_weight,
                        dim_lenght=dim_lenght,
                        dim_width=dim_width,
                        dim_height=dim_height,
                        pick_up_time1=time1,
                        pick_up_time2=time2,
                    )

            return Response({"status": True, "message": "Success", "data": {}})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_analytics_margins_screen_time(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            products = ProductDB.objects.filter(
                seller_id=seller_id[0], verification_status="verified"
            )
            for product in products:
                view = ViewsTrackingDB.objects.get(main_product_id=product)
                last_week_screen_time = view.last_week_screen_time

                hourly_views = list(eval(view.hourly_views))[::-1]
                today_screen_time = 0
                for i in range(len(hourly_views) - 1):
                    if hourly_views[i]["hrs"] == "12 AM":
                        break
                    else:
                        today_screen_time += (
                            hourly_views[i]["screen_time"]
                            - hourly_views[i + 1]["screen_time"]
                        )

                daily_views = list(eval(view.daily_views))[::-1]
                current_week_screen_time = 0
                x = 0
                for i in range(len(daily_views) - 1):
                    if daily_views[i]["day"] == "Sunday":
                        break
                    else:
                        current_week_screen_time += (
                            daily_views[i]["screen_time"]
                            - daily_views[i + 1]["screen_time"]
                        )
                    x = i + 1
                week_av = last_week_screen_time * (x / 7)
                if week_av:
                    change = round(
                        ((current_week_screen_time - week_av) / week_av) * 100, 2
                    )
                else:
                    change = current_week_screen_time * 100

            return Response(
                {
                    "status": True,
                    "message": "Success",
                    "data": {
                        "today": today_screen_time,
                        "current_week": current_week_screen_time,
                        "last_week": last_week_screen_time,
                        "change": f"{change}%",
                    },
                }
            )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_analytics_margins_sold_products(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            ddp_orders = DDP_Orders_DB.objects.filter(
                Q(ddp_seller_id=seller_id[0]) & ~Q(buyer_status="Cancelled")
            )
            exwork_orders = Exwork_Orders_DB.objects.filter(
                Q(exwork_seller_id=seller_id[0])
                & ~Q(buyer_status="Cancelled")
                & ~Q(seller_status="Pending")
            )
            total_sold = 0
            for ddp_order in ddp_orders:
                total_sold += ddp_order.quantity
            for exwork_order in exwork_orders:
                total_sold += exwork_order.quantity

            t = (timezone.now() - timedelta(days=timezone.now().weekday())).replace(
                hour=0, minute=0
            )
            ddp_orders = DDP_Orders_DB.objects.filter(
                Q(ddp_seller_id=seller_id[0])
                & ~Q(buyer_status="Cancelled")
                & Q(created_at__lte=timezone.now())
                & Q(created_at__gte=t)
            )
            exwork_orders = Exwork_Orders_DB.objects.filter(
                Q(exwork_seller_id=seller_id[0])
                & ~Q(buyer_status="Cancelled")
                & ~Q(seller_status="Pending")
                & Q(created_at__lte=timezone.now())
                & Q(created_at__gte=t)
            )

            sold_in_last_7days = 0
            for ddp_order in ddp_orders:
                sold_in_last_7days += ddp_order.quantity
            for exwork_order in exwork_orders:
                sold_in_last_7days += exwork_order.quantity

            return Response(
                {
                    "status": True,
                    "message": "Success",
                    "data": {
                        "total_sold": total_sold,
                        "sold_in_current_week": sold_in_last_7days,
                    },
                }
            )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_analytics_margins_searched_clicks(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        total_searched_clicks = 0
        current_week_searched_clicks = 0
        try:
            for product in ProductDB.objects.filter(
                seller_id=seller_id[0], verification_status="verified"
            ):
                view = ViewsTrackingDB.objects.get(main_product_id=product)
                total_searched_clicks += view.total_searched_clicks

                daily_views = list(eval(view.daily_views))[::-1]
                for i in range(len(daily_views) - 1):
                    if daily_views[i]["day"] == "Sunday":
                        break
                    else:
                        current_week_searched_clicks += (
                            daily_views[i]["searched_clicks"]
                            - daily_views[i + 1]["searched_clicks"]
                        )

            return Response(
                {
                    "status": True,
                    "message": "Success",
                    "data": {
                        "total_searched_clicks": total_searched_clicks,
                        "current_week_searched_clicks": current_week_searched_clicks,
                    },
                }
            )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_analytics_margins_other_clicks(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        total_other_clicks = 0
        current_week_other_clicks = 0
        try:
            for product in ProductDB.objects.filter(
                seller_id=seller_id[0], verification_status="verified"
            ):
                view = ViewsTrackingDB.objects.get(main_product_id=product)
                total_other_clicks += view.total_other_clicks

                daily_views = list(eval(view.daily_views))[::-1]
                for i in range(len(daily_views) - 1):
                    if daily_views[i]["day"] == "Sunday":
                        break
                    else:
                        current_week_other_clicks += (
                            daily_views[i]["other_clicks"]
                            - daily_views[i + 1]["other_clicks"]
                        )

            return Response(
                {
                    "status": True,
                    "message": "Success",
                    "data": {
                        "total_other_clicks": total_other_clicks,
                        "current_week_other_clicks": current_week_other_clicks,
                    },
                }
            )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_analytics_margins_last_7days_searched_clicks(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            products = ProductDB.objects.filter(
                seller_id=seller_id[0], verification_status="verified"
            )
            last_7_days_data = []
            day_names = []
            for i in range(1, 8):
                day_names.append(
                    (timezone.now() - timezone.timedelta(days=i)).strftime("%A - %d %B")
                )

            for i in day_names:
                last_7_days_data.append({"day": i, "searched_clicks": 0})

            for product in products:
                view_data = ViewsTrackingDB.objects.get(main_product_id=product)
                daily = list(eval(view_data.daily_views))[::-1]
                for j in range(len(daily) - 1):
                    last_7_days_data[j]["searched_clicks"] += (
                        daily[j]["searched_clicks"] - daily[j + 1]["searched_clicks"]
                    )

            last_7_days_data.reverse()

            return Response(
                {"status": True, "message": "Success", "data": last_7_days_data}
            )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_analytics_margins_last_7days_other_clicks(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            products = ProductDB.objects.filter(
                seller_id=seller_id[0], verification_status="verified"
            )
            last_7_days_data = []
            day_names = []
            for i in range(1, 8):
                day_names.append(
                    (timezone.now() - timezone.timedelta(days=i)).strftime("%A - %d %B")
                )

            for i in day_names:
                last_7_days_data.append(
                    {
                        "day": i,
                        "other_clicks": 0,
                    }
                )

            for product in products:
                view_data = ViewsTrackingDB.objects.get(main_product_id=product)
                daily = list(eval(view_data.daily_views))[::-1]
                for j in range(len(daily) - 1):
                    last_7_days_data[j]["other_clicks"] += (
                        daily[j]["other_clicks"] - daily[j + 1]["other_clicks"]
                    )

            last_7_days_data.reverse()

            return Response(
                {"status": True, "message": "Success", "data": last_7_days_data}
            )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def seller_analytics_margins_wishlist(request):
    user_id = request.headers["User-id"]
    seller_id = userDB.objects.filter(user_id=user_id)
    if seller_id:
        try:
            t = (timezone.now() - timedelta(days=timezone.now().weekday())).replace(
                hour=0, minute=0
            )
            total = len(Wishlist_DB.objects.filter(seller_id=seller_id[0]))
            week = len(
                Wishlist_DB.objects.filter(
                    Q(seller_id=seller_id[0]) & Q(created_at__gte=t)
                )
            )

            return Response(
                {
                    "status": True,
                    "message": "Success",
                    "data": {"total_wishes": total, "current_week_wishes": week},
                }
            )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def file_upload(request):
    try:
        users = userDB.objects.filter(role="Seller")
        x = []
        for user in users:
            x.append(user.first_name)
        # fileobj = FileTest.objects.create()
        # up_file = request.FILES['file']
        # fileobj.file = up_file
        # fileobj.save()
        return Response(x)
    except Exception as e:
        return Response(
            {
                "status": "Error",
                "message": e,
            }
        )


######################################    Home Page   #######################################


@api_view(["POST"])
def homepage_single_category_products(request):
    try:
        category = request.data["category"]
        number = int(request.data["number"])
        city_list = [
            "Haryana",
            "Mumbai",
            "Delhi",
            "UP",
            "Pune",
            "Banglore",
            "Maharastra",
        ]

        cat = productCategory.objects.get(name=category)
        # products = ProductDB.objects.filter(
        #    verification_status="verified"
        # )
        products = ProductDB.objects.filter(
            category_id=cat, verification_status="verified"
        )
        if len(products) > number:
            products = products[:number]

        products_data = []
        for product in products:
            col_var = ProductColorVariations.objects.get(
                main_product_id=product, main_variant=True
            )
            size_var = ProductSizeVariations.objects.get(
                variant_id=col_var, main_size=True
            )
            reviews = review_db.objects.filter(main_product_id=product)
            rating = 0
            for review in reviews:
                rating += review.rating

            products_data.append(
                {
                    "product_id": product.id,
                    "product_image": "https://negbuy.com:8080"
                    + ProductColorVariationsSerializer(col_var).data["main_image"],
                    "product_name": product.product_title,
                    "brand": product.brand,
                    "mrp": size_var.mrp,
                    "price": size_var.selling_price,
                    "discount": round(
                        (
                            ((size_var.mrp - size_var.selling_price) / size_var.mrp)
                            * 100
                        ),
                        2,
                    ),
                    # changing the city to random value until seller onboarded
                    "city": random.choice(city_list),
                    "country": product.seller_id.country,
                    "flag": "https://flagcdn.com/16x12/in.png",
                    "rating": round(
                        (rating / len(reviews) if len(reviews) != 0 else 0), 1
                    ),
                }
            )
        shuffled_list = sorted(products_data, key=lambda x: random.random())
        return Response({"status": True, "message": "Success", "data": shuffled_list})
    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["GET"])
def homepage_news(request):
    try:
        news = Blogs_DB.objects.all().order_by("-created_at")

        if len(news) > 8:
            news = news[:8]

        news_data = []
        for new in news:
            news_data.append(
                {
                    "news_id": new.id,
                    "image": "https://negbuy.com:8080"
                    + Blogs_DBSerializer(new).data["image"],
                    "heading": new.heading,
                    "description": new.description,
                }
            )
        return Response({"status": True, "message": "Success", "data": news_data})
    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["GET"])
def detailed_homepage_news(request):
    try:
        news = Blogs_DB.objects.all().order_by("-created_at")

        if len(news) > 8:
            news = news[:8]

        news_data = []

        for new in news:
            detailed_news = Detailed_blogs_DB.objects.filter(blog=new)
            detailed_news_data = []
            for single_news in detailed_news:
                detailed_news_data.append(
                    {
                        "heading": single_news.heading,
                        "description": single_news.description,
                    }
                )

            news_data.append(
                {
                    "news_id": new.id,
                    "image": "https://negbuy.com:8080"
                    + Blogs_DBSerializer(new).data["image"],
                    "heading": new.heading,
                    "description": new.description,
                    "detailed_news": detailed_news_data,
                }
            )
        return Response({"status": True, "message": "Success", "data": news_data})
    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["POST"])
def homepage_add_rfq(request):
    user_id = request.headers["User-id"]
    user = userDB.objects.filter(user_id=user_id)
    if user:
        try:
            requirement = request.data["requirement"]
            delivery_date = request.data["delivery_date"]
            quantity = request.data["quantity"]
            target_price = request.data["target_price"]
            r = rfq.objects.create(
                user=user[0],
                requirement=requirement,
                target_price=target_price,
                quantity=quantity,
                delivery_expected_date=delivery_date,
            )

            subject = f"{user[0].first_name} Added a New RFQ"
            message_html = render_to_string(
                "./new_rfq_email_to_admin.html",
                {
                    "first_name": user[0].first_name,
                    "last_name": user[0].last_name,
                    "phone": user[0].phone,
                    "email": user[0].email,
                    "city": user[0].city,
                    "state": user[0].state,
                    "country": user[0].country,
                    "pincode": user[0].postal_code,
                    "requirements": r.requirement,
                    "quantity": r.quantity,
                    "price": r.target_price,
                    "date": r.delivery_expected_date,
                },
            )
            message_text = strip_tags(message_html)
            from_email = settings.EMAIL_HOST_USER
            recipient_list = [
                "amirmohd233@gmail.com",
                "amir@negbuy.com",
                "contactus@negbuy.com",
                "akhtersameer736@gmail.com",
            ]
            email = EmailMultiAlternatives(
                subject, message_text, from_email, recipient_list
            )
            email.attach_alternative(message_html, "text/html")
            email.send()

            return Response({"status": True, "message": "Success", "data": {}})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def get_all_rfq(request):
    user_id = request.headers["User-id"]
    user = userDB.objects.get(user_id=user_id)
    rfq_list = list()
    if user:
        try:
            all_rfq = rfq.objects.filter(user=user)

            for each_rfq in all_rfq:
                rfq_dict = dict()

                rfq_dict["id"] = each_rfq.id
                rfq_dict["requirement"] = each_rfq.requirement
                rfq_dict["target_price"] = each_rfq.target_price
                rfq_dict["quantity"] = each_rfq.quantity
                rfq_dict["delivery_expected_date"] = each_rfq.delivery_expected_date
                rfq_dict["status"] = each_rfq.status
                rfq_dict["rfq_status"] = each_rfq.rfq_status
                rfq_dict["reason"] = each_rfq.reason
                rfq_dict["created_at"] = each_rfq.created_at

                rfq_list.append(rfq_dict)

            return Response({"status": True, "message": "Success", "data": rfq_list})

        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})

    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def buyer_notification(request):
    user_id = request.headers["User-id"]
    buyer_id = userDB.objects.get(user_id=user_id)

    if buyer_id:
        notices = BuyerNoticeBoard.objects.filter(Buyer_id=buyer_id)
        try:
            all_notices = []
            for notice in notices:
                send_time = notice.created_at
                current_time = timezone.now()
                diff = current_time - send_time
                time = calculate_time(diff)

                all_notices.append(
                    {
                        "notice_id": notice.id,
                        "subject": notice.subject,
                        "message": notice.message,
                        "sender": notice.sender,
                        "time_ago": time,
                        "read_status": notice.read,
                    }
                )

            return Response({"status": True, "message": "Success", "data": all_notices})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


###############################      Buyer Profile      ###################################


@api_view(["POST"])
def buyer_profile_page(request):
    user_id = request.data["user_id"]
    user = userDB.objects.filter(user_id=user_id)
    if user:
        try:
            return Response(
                {
                    "status": True,
                    "message": "Success",
                    "data": {
                        "profile_pic": "https://negbuy.com:8080"
                        + UserSerializer(user[0]).data["profile_picture"]
                        if user[0].profile_picture
                        else user[0].profile_picture,
                        "first_name": user[0].first_name,
                        "last_name": user[0].last_name,
                        "gender": user[0].gender,
                        "email": user[0].email,
                        "phone": user[0].phone,
                        "pincode": user[0].postal_code,
                        "city": user[0].city,
                        "state": user[0].state,
                        "country": user[0].country,
                        "address_line1": user[0].address_line1,
                        "address_line2": user[0].address_line2,
                    },
                }
            )
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def buyer_profile_update(request):
    user_id = request.headers["User-id"]
    user = userDB.objects.filter(user_id=user_id)
    if user:
        try:
            first_name = request.data.get("first_name", None)
            last_name = request.data.get("last_name", None)
            gender = request.data.get("gender", None)
            pincode = request.data.get("pincode", None)
            city = request.data.get("city", None)
            state = request.data.get("state", None)
            country = request.data.get("country", None)
            address_line1 = request.data.get("address_line1", None)
            address_line2 = request.data.get("address_line2", None)
            profile_pic = request.FILES.get("profile_pic", None)

            if first_name:
                user[0].first_name = first_name
            if last_name:
                user[0].last_name = last_name
            if gender:
                user[0].gender = gender
            if pincode:
                user[0].postal_code = pincode
            if city:
                user[0].city = city
            if state:
                user[0].state = state
            if country:
                user[0].country = country
            if address_line1:
                user[0].address_line1 = address_line1
            if address_line2:
                user[0].address_line2 = address_line2
            if profile_pic:
                user[0].profile_picture = profile_pic
            user[0].save()

            return Response({"status": True, "message": "Success", "data": {}})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["POST"])
def buyer_send_email_verification_email(request):
    user_id = request.headers["User-id"]
    user = userDB.objects.filter(user_id=user_id)
    if user:
        try:
            new_email = request.data["new_email"]

            if user[0].email == new_email:
                return Response(
                    {
                        "status": "Error",
                        "message": "The email is already registered",
                        "data": {},
                    }
                )

            current_site = get_current_site(request)
            mydict = {
                "username": user[0].first_name + " " + user[0].last_name
                if user[0].first_name and user[0].last_name
                else user[0].first_name
                if user[0].first_name
                else "User",
                "domain": current_site,
                "uidb64": urlsafe_base64_encode(force_bytes(user_id)),
                "new_email": urlsafe_base64_encode(force_bytes(new_email)),
                "token": generate_token.make_token(user[0]),
                "email": new_email,
            }

            html_template = "./email_verification_link_email.html"
            html_content = render_to_string(html_template, mydict)
            text_content = strip_tags(html_content)

            email = EmailMultiAlternatives(
                subject="Verify your Email",
                body=text_content,
                from_email=settings.EMAIL_HOST_USER,
                to=[new_email],
            )
            email.attach_alternative(html_content, "text/html")
            email.send(fail_silently=False)

            return Response({"status": True, "message": "Success", "data": {}})

        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


@api_view(["GET"])
def buyer_email_verify_update(request, uidb64, token, new_email):
    try:
        email = force_text(urlsafe_base64_decode(new_email))
        uid = force_text(urlsafe_base64_decode(uidb64))
        user = userDB.objects.filter(user_id=uid)
        if user and generate_token.check_token(user[0], token):
            user[0].email = email
            user[0].save()
            return redirect("https://negbuy.com/email-verified")

        return Response(
            {"status": False, "message": "Wrong Verification Token", "data": {}}
        )
    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["POST"])
def buyer_phone_number_update(request):
    user_id = request.headers["User-id"]
    user = userDB.objects.filter(user_id=user_id)
    if user:
        try:
            new_user_id = request.data["new_user_id"]
            new_phone = request.data["new_phone"]

            user[0].user_id = new_user_id
            user[0].phone = new_phone
            user[0].save()

            return Response({"status": True, "message": "Success", "data": new_user_id})
        except Exception as e:
            return Response({"status": "Error", "message": str(e), "data": {}})
    else:
        return Response({"status": False, "message": "Unauthorised User", "data": {}})


############################      Show Categorywise Products       ####################


@api_view(["POST"])
def category_specific_products(request):
    try:
        category = request.data["category"]
        subcategory = request.data.get("subcategory", None)
        id = productCategory.objects.get(name=category)
        if subcategory:
            products = ProductDB.objects.filter(
                subcategory=subcategory, verification_status="verified"
            )
        else:
            products = ProductDB.objects.filter(
                category_id=id, verification_status="verified"
            )
        products_data = []
        for product in products:
            col_var = ProductColorVariations.objects.get(
                main_product_id=product, main_variant=True
            )
            size_var = ProductSizeVariations.objects.get(
                variant_id=col_var, main_size=True
            )
            reviews = review_db.objects.filter(main_product_id=product)
            rating = 0
            for review in reviews:
                rating += review.rating

            products_data.append(
                {
                    "product_id": product.id,
                    "product_image": "https://negbuy.com:8080"
                    + ProductColorVariationsSerializer(col_var).data["main_image"],
                    "product_name": product.product_title,
                    "brand": product.brand,
                    "mrp": size_var.mrp,
                    "price": size_var.selling_price,
                    "discount": round(
                        (
                            ((size_var.mrp - size_var.selling_price) / size_var.mrp)
                            * 100
                        ),
                        2,
                    ),
                    "city": product.seller_id.city,
                    "country": product.seller_id.country,
                    "rating": random.randint(0, 5)
                    # "rating": round((rating / len(reviews)), 1)
                    # if len(reviews) != 0
                    # else 0,
                }
            )

        return Response({"status": True, "message": "Success", "data": products_data})
    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


############################   Product Detailed Page     ########################


@api_view(["GET"])
def product_details_page(request):
    try:
        user_id = request.headers.get("User-id", None)
        product_id = int(request.query_params["product_id"])
        # product_id = int(request.data["product_id"])
        search_click = request.data.get("search_click", False)
        app = request.data.get("app", False)
        product = ProductDB.objects.get(id=product_id, verification_status="verified")
        color_obj = ProductColorVariations.objects.get(
            main_product_id=product, main_variant=True
        )
        size_obj = ProductSizeVariations.objects.get(
            variant_id=color_obj, main_size=True
        )
        view = ViewsTrackingDB.objects.get(main_product_id=product)

        if search_click:
            view.total_searched_clicks += 1
        else:
            view.total_other_clicks += 1
        if app:
            view.app_views += 1
        else:
            view.browser_views += 1
        view.save()

        if user_id:
            recent = userDB.objects.get(user_id=user_id).recently_viewed_products
            if product not in recent.all():
                if len(recent.all()) >= 6:
                    recent.remove(recent.first())
                recent.add(product)

        return Response(
            {
                "status": True,
                "message": "Success",
                "data": {
                    "User-id": product.seller_id.user_id,
                    "product_id": product.id,
                    "product_title": product.product_title,
                    "main_sku_id": product.main_sku_id,
                    "category": product.category_id.name,
                    "subcategory": product.subcategory,
                    "brand": product.brand,
                    "keywords": product.keywords,
                    "address": product.packing_address,
                    "port": product.transportation_port,
                    "sale_start": product.sale_startdate,
                    "sale_end": product.sale_enddate,
                    "video": "https://negbuy.com:8080"
                    + ProductDBSerializer(product).data["video"]
                    if product.video
                    else None,
                    "detailed_desc": product.detailed_description,
                    "variants": [
                        {
                            "color": col.color,
                            "main_variant": col.main_variant,
                            "main_image": "https://negbuy.com:8080"
                            + ProductColorVariationsSerializer(col).data["main_image"],
                            "extra_images": [
                                "https://negbuy.com:8080"
                                + ProductColorVariationsSerializer(col).data[
                                    "main_image"
                                ]
                            ]
                            + [
                                "https://negbuy.com:8080"
                                + ProductExtraImagesSerializer(j).data["image"]
                                for j in ProductExtraImages.objects.filter(
                                    variant_id=col
                                )
                            ],
                            # if product.price_choice == "Add Price":
                            "size_variants": [
                                {
                                    "size_id": k.id,
                                    "size": k.size,
                                    "mrp": k.mrp,
                                    "price": k.selling_price,
                                    "discount": round(
                                        (((k.mrp - k.selling_price) / k.mrp) * 100),
                                        2,
                                    ),
                                    "sale_price": k.sale_price,
                                    "weight": k.weight,
                                    "packing_details": k.packing_details,
                                    "dim_length": k.dim_length,
                                    "dim_width": k.dim_width,
                                    "dim_height": k.dim_height,
                                    "manufacturing_time": k.manufacturing_time,
                                    # "max_order_quantity": k.max_order_quantity,
                                    "main_size": k.main_size,
                                    # "stock": k.productinventorydb.stock
                                    # if hasattr(k, "productinventorydb")
                                    # else None,
                                    # "sub_sku_id": k.subskuiddb.sub_sku_id
                                    # if hasattr(k, "subskuiddb")
                                    # else None,
                                    "bulk_purchase_details": [
                                        {
                                            "min_quantity": d.min_quantity,
                                            "max_quantity": d.max_quantity,
                                            "price": d.price,
                                            "manufacturing_time": d.manufacturing_time,
                                        }
                                        for d in ProductBulkPurchaseDetails.objects.filter(
                                            product_id=k
                                        )
                                    ],
                                }
                                for k in ProductSizeVariations.objects.filter(
                                    variant_id=col
                                )
                            ],
                        }
                        for col in ProductColorVariations.objects.filter(
                            main_product_id=product
                        )
                    ],
                },
            }
        )

    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["POST"])
def product_details_page_head_desc_details(request):
    try:
        product_id = int(request.data["product_id"])
        product = ProductDB.objects.filter(
            id=product_id, verification_status="verified"
        )

        return Response(
            {
                "status": True,
                "message": "success",
                "data": ProductDetails.objects.filter(
                    main_product_id=product[0]
                ).values("heading", "description"),
            }
        )

    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["POST"])
def product_details_page_img_desc(request):
    try:
        # product_id = int(request.data["product_id"])
        product_id = int(request.query_params["product_id"])
        product = ProductDB.objects.filter(
            id=product_id, verification_status="verified"
        )

        return Response(
            {
                "status": True,
                "message": "success",
                "data": [
                    {
                        "image": "https://negbuy.com:8080"
                        + ProductImageDescriptionSerializer(i).data["image"],
                        "heading": i.heading,
                        "description": i.description,
                    }
                    for i in ProductImageDescription.objects.filter(
                        main_product_id=product[0]
                    )
                ],
            }
        )

    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["POST"])
def product_details_page_review_section(request):
    try:
        product_id = int(request.data["product_id"])
        product = ProductDB.objects.filter(
            id=product_id, verification_status="verified"
        )

        return Response(
            {
                "status": True,
                "message": "success",
                "data": [
                    {
                        "review_id": review.id,
                        "title": review.review_title,
                        "description": review.review_description,
                        "rating": review.rating,
                        "user": review.user.first_name + " " + review.user.last_name
                        if review.user.first_name and review.user.last_name
                        else review.user.first_name
                        if review.user.first_name
                        else "Buyer",
                        "city": review.user.city,
                        "time": f"posted {calculate_time(timezone.localtime()-timezone.localtime(review.created_at))}",
                        "files": [
                            "https://negbuy.com:8080"
                            + Review_Images_DBSerializer(img).data["file"]
                            for img in Review_Images_DB.objects.filter(review_id=review)
                        ],
                    }
                    for review in review_db.objects.filter(main_product_id=product[0])
                ],
            }
        )

    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["POST"])
def product_details_page_review_analysis(request):
    try:
        product_id = int(request.data["product_id"])
        product = ProductDB.objects.filter(
            id=product_id, verification_status="verified"
        )
        reviews = review_db.objects.filter(main_product_id=product[0])
        total_ratings = len(reviews)
        ratings = []
        total_reviews = 0
        for review in reviews:
            ratings.append(review.rating)
            if review.review_title != None or review.review_description != None:
                total_reviews += 1

        star1 = ratings.count(1)
        star2 = ratings.count(2)
        star3 = ratings.count(3)
        star4 = ratings.count(4)
        star5 = ratings.count(5)
        star1p = int((star1 / total_ratings) * 100) if total_ratings else 0
        star2p = int((star2 / total_ratings) * 100) if total_ratings else 0
        star3p = int((star3 / total_ratings) * 100) if total_ratings else 0
        star4p = int((star4 / total_ratings) * 100) if total_ratings else 0
        star5p = int((star5 / total_ratings) * 100) if total_ratings else 0
        average_stars = round(
            (star1 + 2 * star2 + 3 * star3 + 4 * star4 + 5 * star5) / 5, 1
        )
        percent_average_stars = int((average_stars / 5) * 100)

        return Response(
            {
                "status": True,
                "message": "success",
                "data": {
                    "stars_out_of_5": average_stars,
                    "percent_average_stars": percent_average_stars,
                    "total_reviews": total_reviews,
                    "total_ratings": total_ratings,
                    "star1": {"percent": star1p, "number": star1},
                    "star2": {"percent": star2p, "number": star2},
                    "star3": {"percent": star3p, "number": star3},
                    "star4": {"percent": star4p, "number": star4},
                    "star5": {"percent": star5p, "number": star5},
                },
            }
        )

    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["GET"])
def product_details_page_countrywise_reviews_piechart(request):
    try:
        product_id = int(request.data["product_id"])
        product = ProductDB.objects.filter(
            id=product_id, verification_status="verified"
        )
        reviews = review_db.objects.filter(main_product_id=product[0])

        countries = dict()
        for review in reviews:
            country = review.user.country
            if country in countries.keys():
                countries[country] += 1
            else:
                countries[country] = 1

        countries = dict(sorted(countries.items(), key=lambda x: x[1], reverse=True))
        count = dict()
        for c in countries.items():
            count[c[0]] = (c[1] / len(reviews)) * 100

        x = dict()

        if len(countries) > 4:
            for i in countries.items():
                if len(x) == 3:
                    break
                else:
                    x[i[0]] = (i[1] / len(reviews)) * 100

            for key in list(x.keys()):
                if key in countries:
                    del countries[key]

            sum = 0
            for y in countries.values():
                sum += y

            x["Others"] = (sum / len(reviews)) * 100

        else:
            for i in countries.items():
                x[i[0]] = (i[1] / len(reviews)) * 100

        return Response(
            {
                "status": True,
                "message": "Success",
                "data": {"piechartx": x, "others": count},
            }
        )

    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["POST"])
def product_page_screen_time_update(request):
    try:
        product_id = int(request.data["product_id"])
        screen_time = int(request.data.get("screen_time", 0))

        product = ProductDB.objects.get(id=product_id)
        view = ViewsTrackingDB.objects.get(main_product_id=product)
        view.total_screen_time += screen_time
        view.save()

        return Response({"status": True, "message": "Success", "data": {}})

    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


import pgeocode


@api_view(["POST"])
def get_city_state_country_with_pincode(request):
    try:
        pincode = request.data["pincode"]
        response = requests.get(f"https://api.postalpincode.in/pincode/{pincode}")
        data = response.json()

        if data[0]["Status"] == "Error" or response.status_code == 404:
            geopy.geocoders.options.default_language = "en"
            geolocator = Nominatim(user_agent="myapp")
            location = geolocator.geocode(pincode)

            if location:
                latitude = location.latitude
                longitude = location.longitude
                location = geolocator.reverse(f"{latitude},{longitude}")
                address = location.raw.get("address", {})
                print(address)
                country = address.get("country")
                state = address.get("state")
                city = address.get("city")

                return Response(
                    {
                        "status": True,
                        "message": "Success",
                        "data": {"country": country, "state": state, "city": city},
                    }
                )
        else:
            dat = response.json()
            dat = dat[0]["PostOffice"][0]

            return Response(
                {
                    "status": True,
                    "message": "Success",
                    "data": {
                        "country": GoogleTranslator(
                            source="auto", target="en"
                        ).translate(dat["Country"]),
                        "state": GoogleTranslator(source="auto", target="en").translate(
                            dat["State"]
                        ),
                        "city": GoogleTranslator(source="auto", target="en").translate(
                            dat["District"]
                        ),
                    },
                }
            )

    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


def port_finder(latitude, longitude, country):
    dist = []
    coords_1 = (latitude, longitude)
    ports = port.objects.filter(country__icontains=country)

    for prt in ports:
        latitude_val = prt.latitude
        parts = latitude_val.split(" ")
        degrees = float(parts[0][:-1])
        minutes = float(parts[1][:-1])
        seconds = float(parts[2][:-2])
        latitude = float(degrees) + float(minutes) / 60 + float(seconds) / 3600
        longitude_val = prt.longitude
        parts = longitude_val.split(" ")
        degrees = float(parts[0][:-1])
        minutes = float(parts[1][:-1])
        seconds = float(parts[2][:-2])
        longitude = float(degrees) + float(minutes) / 60 + float(seconds) / 3600
        coords_2 = (latitude, longitude)
        print(coords_2)
        distance = geopy.distance.geodesic(coords_1, coords_2).km
        dist.append(distance)
    idx = dist.index(min(dist))

    return {
        "name": ports[idx].name,
        "distance": dist[idx],
        "latitude": ports[idx].latitude,
        "longitude": ports[idx].longitude,
    }


@api_view(["GET"])
def get_coordinates_for_globe(request):
    try:
        user_id = request.headers["User-id"]
        user = userDB.objects.filter(user_id=user_id)
        if user:
            buyer_pin = user[0].postal_code
            geolocator = Nominatim(user_agent="myapp")
            buyer_location = geolocator.geocode(buyer_pin)

            if buyer_location:
                buyer_latitude = buyer_location.latitude
                buyer_longitude = buyer_location.longitude
                location = geolocator.reverse(f"{buyer_latitude},{buyer_longitude}")
                buyer_country = location.raw.get("address", {}).get("country")
                buyer_coordinates = {
                    "latitude": buyer_latitude,
                    "longitude": buyer_longitude,
                }

            seller_location = geolocator.geocode(201301)
            if seller_location:
                seller_latitude = seller_location.latitude
                seller_longitude = seller_location.longitude
                location = geolocator.reverse(f"{buyer_latitude},{buyer_longitude}")
                seller_country = location.raw.get("address", {}).get("country")
                seller_coordinates = {
                    "latitude": seller_latitude,
                    "longitude": seller_longitude,
                }

            if buyer_country == "India":
                return Response(
                    {
                        "status": True,
                        "message": "Success",
                        "data": {
                            "seller": seller_coordinates,
                            "buyer": buyer_coordinates,
                        },
                    }
                )
            else:
                return Response(
                    {
                        "status": True,
                        "message": "Success",
                        "data": {
                            "seller": seller_coordinates,
                            "seller_port": port_finder(
                                seller_latitude, seller_longitude, seller_country
                            ),
                            "buyer": buyer_coordinates,
                            "buyer_port": port_finder(
                                buyer_latitude, buyer_longitude, buyer_country
                            ),
                        },
                    }
                )
        else:
            return Response(
                {"status": False, "message": "Unauthorised User", "data": {}}
            )
    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


########################   BUYER MY ORDERS    ##########################################


@api_view(["GET"])
def buyer_all_orders(request):
    try:
        user_id = request.headers["User-id"]
        buyer = userDB.objects.get(user_id=user_id)

        if buyer:
            ddp_orders = DDP_Orders_DB.objects.filter(ddp_buyer_id=buyer)
            orders = []
            if ddp_orders:
                for ddp_order in ddp_orders:
                    orders.append(
                        {
                            "product_image": "https://negbuy.com:8080"
                            + ProductColorVariationsSerializer(
                                ddp_order.product_id.variant_id
                            ).data["main_image"],
                            "product_id": ddp_order.product_id.id,
                            "product_title": ddp_order.product_id.variant_id.main_product_id.product_title,
                            "order_id": ddp_order.id,
                            "quantity": ddp_order.quantity,
                            "order_confirm_date": ddp_order.created_at.strftime(
                                "%d %B %Y"
                            ),
                            "order_date": ddp_order.created_at.strftime("%d %B %Y"),
                            "delivery_option": "DDP",
                            "status": ddp_order.buyer_status,
                            "shipping_address": ddp_order.shipping_address,
                            "expected_delivery_date": ddp_order.expected_delivery_date,
                            "amount": ddp_order.product_price * ddp_order.quantity,
                            "delivered_date": ddp_order.modified_at.strftime("%d %B %Y")
                            if ddp_order.buyer_status == "Delivered"
                            else False,
                            "cancelled_date": ddp_order.modified_at.strftime("%d %B %Y")
                            if ddp_order.buyer_status == "Cancelled"
                            else False,
                        }
                    )

            exwork_orders = Exwork_Orders_DB.objects.filter(Q(exwork_buyer_id=buyer))
            for exwork_order in exwork_orders:
                orders.append(
                    {
                        "product_image": "https://negbuy.com:8080"
                        + ProductColorVariationsSerializer(
                            exwork_order.product_id.variant_id
                        ).data["main_image"],
                        "product_id": exwork_order.product_id.id,
                        "product_title": exwork_order.product_id.variant_id.main_product_id.product_title,
                        "order_id": exwork_order.id,
                        "quantity": exwork_order.quantity,
                        "order_date": exwork_order.created_at.strftime("%d %B %Y"),
                        "delivery_option": "ExWork",
                        "status": exwork_order.buyer_status,
                        "shipping_address": exwork_order.shipping_address,
                        "expected_delivery_date": exwork_order.expected_delivery_date,
                        "amount": exwork_order.product_price * exwork_order.quantity,
                        "delivered_cancelled_date": exwork_order.modified_at.strftime(
                            "%d %B %Y"
                        )
                        if exwork_order.buyer_status
                        else False,
                    }
                )

            orders = sorted(orders, key=lambda x: x["order_date"], reverse=True)
            return Response({"status": True, "message": "Success", "data": orders})

        return Response({"status": False, "message": "Unauthorised User", "data": {}})
    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["POST"])
def buyer_submit_docket_number(request):
    try:
        user_id = request.headers["User-id"]
        buyer = userDB.objects.get(user_id=user_id)

        docket_number = request.data.get("docket_number")
        docket_file = request.FILES.get("docket_file", None)
        exwork_id = request.data.get("order_id")
        if buyer:
            exwork_order_obj = Exwork_Orders_DB.objects.get(id=exwork_id)
            exwork_order_obj.docket_number = docket_number
            exwork_order_obj.docket_file = docket_file
            exwork_order_obj.save()

            return Response({"status": True, "message": "Success", "data": {}})
        else:
            return Response(
                {"status": "Error", "message": "Unauthorised User", "data": {}}
            )

    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["GET"])
def buy(request):
    try:
        user_id = request.headers["User-id"]
        buyer = userDB.objects.get(user_id=user_id)

        if buyer:
            pass

    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


# from django.http import JsonResponse
# import firebase_admin
# from firebase_admin import auth, credentials
# from google.oauth2 import service_account

# # Construct the path to the credentials file
# credentials_path = os.path.join(os.getcwd(), "future-pager-340912-firebase-adminsdk-6gzim-e44aece06c.json")
# print(credentials_path)

# # Load the credentials from the JSON file
# creds = service_account.Credentials.from_service_account_file(credentials_path)
# print(creds)
# # cred = credentials.Certificate('future-pager-340912-firebase-adminsdk-6gzim-9d29282878.json')
# # firebase_admin.initialize_app(creds)
# print(settings.FIREBASE_PRIVATE_KEY)
# from django.conf import settings
# """SETUP FIREBASE CREDENTIALS"""
# # cred = credentials.Certificate({
# #         "type" : settings.FIREBASE_ACCOUNT_TYPE,
# #         "project_id" : settings.FIREBASE_PROJECT_ID,
# #         "private_key_id" : settings.FIREBASE_PRIVATE_KEY_ID,
# #         "private_key" : settings.FIREBASE_PRIVATE_KEY.replace('\\n', '\n'),
# #         "client_email" : settings.FIREBASE_CLIENT_EMAIL,
# #         "client_id" : settings.FIREBASE_CLIENT_ID,
# #         "auth_uri" : settings.FIREBASE_AUTH_URI,
# #         "token_uri" : settings.FIREBASE_TOKEN_URI,
# #         "auth_provider_x509_cert_url" : settings.FIREBASE_AUTH_PROVIDER_X509_CERT_URL,
# #         "client_x509_cert_url" : settings.FIREBASE_CLIENT_X509_CERT_URL
# # })

# cred = credentials.Certificate({
#   "type": "service_account",
#   "project_id": "future-pager-340912",
#   "private_key_id": "e44aece06cc812ed5334581e35c5de52e5db9663",
#   "private_key": "-----BEGIN PRIVATE KEY-----\nMIIEvAIBADANBgkqhkiG9w0BAQEFAASCBKYwggSiAgEAAoIBAQDSjr03F/Dt1Vgu\noOsMKyelXMwhW8zaeZaJuSHjv13Fk9e/FTwcC6JzOV79BwSqj7DNAC+XvKYZ7HVF\nNj7vfDNo9b0hebJ6DCvmdx0onBfL4r359KsJJhm0YfvG09kCWnF6HpQaI6E6+wZ7\ndCx3eX3qbBus+xG0sChVTXqkpCiucoJpId4TeIS/iat42LmIH5Ud0Ve5vA5q9FFf\nH38pN6bF79AjkfT1qwGfyJlmFdcU4QRkn+EHKY8lazIGZwlgR04BhnbZp51re2iF\nI9bdreLMaoT2t6a1MRWCM3in4pnOks0wJQ3e1jird5maFm+sfn6sfIiedNRqQXJD\nLbZysAFxAgMBAAECggEAKOgzaas8kzBiR/jagYw20SE0xXxc0ctSjGJL67/nm3Ws\nMkMRYQ9/maw5QKZm6S4udr2FAZFUoe+3HuG5m3SBVnB7tK+8lYqc0tMsdpHQ67sb\nFp9KxWXCE7H/pPLOk+73Yfbj5ioq7lcLvNtmsjWH2Y30WIiUuYqJ43zAiZEdGDrr\nx5ipf49zauLujpFPKclq3mcYWm5iKDQ0q+/kwcrZQThRhsQkGm+N14NnucJs4mCl\nzq5c0K9JvR48Yd6KK3zLWmRCtdk5yOLATUsPVuFEpqnQKRkBZxjBTz1iSSozeoUO\nhND1Onylh1lI2NknRGz03mZxpx6U/f4nl/EtBKxeqwKBgQDwbPamxeeciLxkfMoz\nz7XASirhtMSrfdfzUnazBd71YwoTdmkLrrPtItko7Q7+64IPgPGRkFUYHIzn8J0l\n+FribHPvF5DTMMDcaAHmENicR2FgPQ6T+cWtza55S0tf+mO0vuwziLuD7LmswP95\nhctSI8flV14gcdeGdpApJh02wwKBgQDgMndgMbXNVeiu1Q3A8NRMuP2I+pY+DMz4\nnykex7ieKeAP6eWRl+DyoMyr1R6OehslMpLTVrLMiZPbW+4ZNYndG3VzLW0khfmL\nuIjryO4tGNNnUXUmO59Z1QRbzudgLTkJgYxz7Wosv31UpIhTuRjx24GJimtG2nuk\nj8oOtFzruwKBgCYM2KxkPdekPEybA1VX1ifslQTLmySY8ghsUKTclQC4/+s8njtf\nFtK85Sl4+xqVIyxY9+JVuWHrgtFiXzyXJasYBvEcBaum158KRuKmX+G72klk4F3C\n9eof1zETHYa/FfCRFsjBJwwl9uzETqpo5ljFmOlO/nKcEcQSF5arlYwJAoGAZHnZ\njC/bFWIp/KiZGF4WFyhU393GFZrxBX93E0dY0vFbbGZbrytM8g/kiiHiJuJMZhhL\nOxHA4e/KZuXFSGD5HzGeKncyUcsFMnwKSPls3KMio0wbX34bBUx1Ppv0j+LUjBEW\nHCJLkjBjf9qZCkSLDGFepYB2YXOrM+4JFRv6BN8CgYBzD1RV8Wi9HIEV9fUmJ2Uw\nqac6gdqht4PGDi+rnULL7PaimKZ3irwPX6q14BtJhBHZuLLW7ywE0J37vOBX2KaA\nSdhcv2fqWxgjlbOL8KmKBiEQ+44OSJG5mgAt32ZuydGZH051AfuQDvp/dS/bRYC2\nqIMSoAxsKaStvFNnnCjA3g==\n-----END PRIVATE KEY-----\n",
#   "client_email": "firebase-adminsdk-6gzim@future-pager-340912.iam.gserviceaccount.com",
#   "client_id": "116292296677429060834",
#   "auth_uri": "https://accounts.google.com/o/oauth2/auth",
#   "token_uri": "https://oauth2.googleapis.com/token",
#   "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
#   "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/firebase-adminsdk-6gzim%40future-pager-340912.iam.gserviceaccount.com"
# })
# default_app = firebase_admin.initialize_app(cred)

# @api_view(["GET"])
# def fire_send_otp(request):
#     phone_number = request.data['phone_number']
#     verification = auth.create_phone_verification(phone_number)
#     return JsonResponse({'status': 'success'})

# @api_view(["GET"])
# def fire_verify_otp(request):
#     phone_number = request.data['phone_number']
#     # phone_number = request.POST.get('phone_number')
#     # otp = request.POST.get('otp')
#     otp = request.data['otp']
#     verification = auth.verify_phone_number(phone_number, otp)
#     user = auth.get_user(verification.uid)
#     # TODO: Log the user in and create a session
#     return JsonResponse({'status': 'success', 'user':user})


###################        WISHLIST         #####################################


@api_view(["POST"])
def add_product_to_wishlist(request):
    try:
        user_id = request.headers["User-id"]
        user = userDB.objects.filter(user_id=user_id)
        if user:
            product_id = int(request.data["product_id"])
            product = ProductDB.objects.get(
                id=product_id, verification_status="verified"
            )
            Wishlist_DB.objects.create(
                buyer_id=user[0], seller_id=product.seller_id, product_id=product
            )

            return Response({"status": True, "message": "Success", "data": {}})

        else:
            return Response(
                {"status": False, "message": "Unauthorised User", "data": {}}
            )

    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["POST"])
def remove_product_from_wishlist(request):
    try:
        user_id = request.headers["User-id"]
        user = userDB.objects.filter(user_id=user_id)
        if user:
            product_id = int(request.data["product_id"])
            Wishlist_DB.objects.filter(
                buyer_id=user[0], product_id=ProductDB.objects.get(id=product_id)
            ).delete()

            return Response({"status": True, "message": "Success", "data": {}})

        else:
            return Response(
                {"status": False, "message": "Unauthorised User", "data": {}}
            )

    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["GET"])
def get_product_from_wishlist(request):
    try:
        user_id = request.headers["User-id"]
        user = userDB.objects.filter(user_id=user_id)
        res_list = list()
        if user:
            wishlist_data = Wishlist_DB.objects.filter(buyer_id=user[0])
            for i in wishlist_data:
                product_obj = ProductDB.objects.get(id=i.product_id.id)
                color_variation_obj = ProductColorVariations.objects.filter(
                    main_product_id=product_obj
                )[0]
                size_variation_obj = ProductSizeVariations.objects.filter(
                    variant_id=color_variation_obj
                )[0]

                res = dict()

                res["product_id"] = product_obj.id
                res["main_image"] = "https://negbuy.com:8080" + str(
                    color_variation_obj.main_image.url
                )
                res["selling_price"] = size_variation_obj.selling_price
                res["mrp"] = size_variation_obj.mrp
                res["title"] = product_obj.product_title

                res_list.append(res)
            return Response({"status": True, "message": "Success", "data": res_list})

        else:
            return Response(
                {"status": False, "message": "Unauthorised User", "data": {}}
            )

    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


#####################        Buying   Process     #################


@api_view(["POST"])
def click_on_buy_now(request):
    try:
        user_id = request.headers["User-id"]
        user = userDB.objects.filter(user_id=user_id)
        if user:
            product_id = int(request.data["product_id"])
            quantity = int(request.data["quantity"])
            product = ProductSizeVariations.objects.get(id=product_id)
            inventory_obj = ProductInventoryDB.objects.get(product_id=product)
            stock = inventory_obj.stock

            if stock < quantity:
                return Response(
                    {
                        "status": False,
                        "message": "Stock is less than entered quantity. Please decrease the quantity and try again.",
                        "data": {},
                    }
                )
            else:
                pricing = product.variant_id.main_product_id.price_choice
                if pricing == "Price according to quantity":
                    bulk_price = ProductBulkPurchaseDetails.objects.filter(
                        product_id=product
                    ).order_by("max_quantity")
                    if bulk_price[0].min_quantity > quantity:
                        return Response(
                            {
                                "status": False,
                                "message": f"You have to order atleast {bulk_price[0].min_quantity} pieces.",
                                "data": {},
                            }
                        )

            return Response({"status": True, "message": "Success", "data": {}})

        else:
            return Response(
                {"status": False, "message": "Unauthorised User", "data": {}}
            )

    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["POST"])
def tracker_page_data(request):
    try:
        user_id = request.headers["User-id"]
        user = userDB.objects.filter(user_id=user_id)
        # product_id, quantity, order_type
        if user:
            product_id = int(request.data["product_id"])
            quantity = int(request.data["quantity"])
            product = ProductSizeVariations.objects.get(id=product_id)
            inventory_obj = ProductInventoryDB.objects.get(product_id=product)
            stock = int(inventory_obj.stock)

            if stock < quantity:
                return Response(
                    {
                        "status": False,
                        "message": "Stock is less than entered quantity. Please decrease the quantity and try again.",
                        "data": {},
                    }
                )

            pricing = product.variant_id.main_product_id.price_choice
            if pricing == "Add Price":
                price = int(product.selling_price)
            else:
                bulk_price = ProductBulkPurchaseDetails.objects.filter(
                    product_id=product
                ).order_by("max_quantity")

                if bulk_price[0].min_quantity > quantity:
                    return Response(
                        {
                            "status": False,
                            "message": f"You have to order atleast {bulk_price[0].min_quantity} pieces.",
                            "data": {},
                        }
                    )
                else:
                    for i in range(len(bulk_price)):
                        if (
                            bulk_price[i].min_quantity
                            <= quantity
                            <= bulk_price[i].max_quantity
                        ):
                            price = int(bulk_price[i].price)
                            break
                        else:
                            price = int(bulk_price[len(bulk_price) - 1].price)

            pp = Purchase_Process_DP.objects.filter(user_id=user[0], status="active")
            if pp:
                pp.delete()

            total_weight = math.ceil(
                int(product.weight) * int(quantity) * 1.1
            )  # 10% of total weight buffer for packing material
            volumetric_weight = math.ceil(
                int(product.dim_length)
                * int(product.dim_width)
                * int(product.dim_height)
                * int(quantity)
                * 1.1
                / 5000
            )

            if total_weight < volumetric_weight:
                total_weight = volumetric_weight

            gst = float(product.variant_id.main_product_id.gst) * price * quantity / 100
            service_charge = price * quantity * 0.02  # 2% of the product price

            order_type = request.data["order_type"]

            if order_type == "DDP":
                transport_mode = request.data["transport_mode"]
                delivery_address = request.data["delivery_address"]
                country = request.data["country"]
                state = request.data["state"]
                city = request.data["city"]
                pincode = int(request.data["pincode"])
                freight_charge = request.data["freight_charge"]

                ppo = Purchase_Process_DP.objects.create(
                    user_id=user[0],
                    product_id=product,
                    quantity=quantity,
                    price=price,
                    total_weight=total_weight,
                    gst=gst,
                    service_charge=service_charge,
                    order_type=order_type,
                    transport_mode=transport_mode,
                    delivery_address=delivery_address,
                    country=country,
                    state=state,
                    city=city,
                    pincode=pincode,
                )
                ppo.company_courier_id = request.data["company_courier_id"]
                ppo.company_courier_name = request.data["company_courier_name"]
                ppo.courier_charges = freight_charge
                ppo.estimated_delivery_days = request.data["estimated_delivery_days"]
                ppo.total_amount = (
                    int(freight_charge)
                    + int(ppo.service_charge)
                    + int(ppo.price * ppo.quantity)
                    + int(ppo.gst)
                )
            else:
                ppo = Purchase_Process_DP.objects.create(
                    user_id=user[0],
                    product_id=product,
                    quantity=quantity,
                    price=price,
                    total_weight=total_weight,
                    gst=gst,
                    service_charge=service_charge,
                    order_type=order_type,
                )
                ppo.total_amount = (
                    int(ppo.service_charge)
                    + int(ppo.price * ppo.quantity)
                    + int(ppo.gst)
                )
            ppo.save()

            return Response(
                {"status": True, "message": "Success", "data": {"instance_id": ppo.id}}
            )

        else:
            return Response(
                {"status": False, "message": "Unauthorised User", "data": {}}
            )

    except Exception as e:
        return Response(
            {
                "status": "Error",
                "message": str(e),
                "data": {},
                "res": request.data,
                "header": request.headers,
            }
        )


@api_view(["POST"])
def payment_page(request):
    try:
        user_id = request.headers["User-id"]
        user = userDB.objects.filter(user_id=user_id)
        if user:
            instance_id = int(request.data["instance_id"])
            ppo = Purchase_Process_DP.objects.filter(id=instance_id)

            if ppo:
                ppo = ppo[0]
                return Response(
                    {
                        "status": True,
                        "message": "Success",
                        "data": {
                            "instance_id": ppo.id,
                            "product_image": "https://negbuy.com:8080"
                            + ProductColorVariationsSerializer(
                                ppo.product_id.variant_id
                            ).data["main_image"],
                            "product_title": ppo.product_id.variant_id.main_product_id.product_title,
                            "courier_name": ppo.company_courier_name,
                            "quantity": ppo.quantity,
                            "price": ppo.price,
                            "gst": ppo.gst,
                            "service_charge": ppo.service_charge,
                            "delivery_charge": ppo.courier_charges,
                            "total_amount": ppo.total_amount,
                        },
                    }
                )

            else:
                return Response(
                    {
                        "status": False,
                        "message": "You replaced your cart with some other item. Please check your cart.",
                        "data": {},
                    }
                )

        else:
            return Response(
                {"status": False, "message": "Unauthorised User", "data": {}}
            )

    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})

    #########  after this RazorPay APIs will be called  #######


@api_view(["POST"])
def exwork_pay_later(request):
    try:
        user_id = request.headers["User-id"]
        user = userDB.objects.filter(user_id=user_id)
        if user:
            instance_id = int(request.data["instance_id"])
            ppo = Purchase_Process_DP.objects.filter(id=instance_id)

            if ppo:
                Exwork_Orders_DB.objects.create(
                    exwork_buyer_id=ppo[0].user_id,
                    exwork_seller_id=ppo[
                        0
                    ].product_id.variant_id.main_product_id.seller_id,
                    product_id=ppo[0].product_id,
                    product_price=ppo[0].price,
                )
                ppo[0].delete()
                return Response({"status": True, "message": "Success", "data": {}})

            else:
                return Response(
                    {
                        "status": False,
                        "message": "You replaced your cart with some other item. Please check your cart.",
                        "data": {},
                    }
                )

        else:
            return Response(
                {"status": False, "message": "Unauthorised User", "data": {}}
            )

    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


##############################       APIs dedicated for APP only       #################

#################   HomePage APIs    #######


@api_view(["GET"])
def category_names_images(request):
    try:
        cats = productCategory.objects.all()
        cats_data = []
        for cat in cats:
            cats_data.append(
                {
                    "name": cat.app_name,
                    "image": "https://negbuy.com:8080"
                    + productCategorySerializer(cat).data["image"]
                    if cat.image
                    else False,
                }
            )

        return Response({"status": True, "message": "Success", "data": cats_data})
    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["GET"])
def subcategories(request):
    try:
        catpage = {}
        for category in productCategory.objects.all():
            file_path = os.path.join(
                settings.BASE_DIR, "static_files", "categories", f"{category.name}.txt"
            )
            with open(file_path, "r") as f:
                file_contents = f.read()

            lines = file_contents.split("\n")
            filtered_lines = []
            for line in lines:
                line = line.split(">")
                if len(line) > 1 and line[1].replace(" ", "") not in filtered_lines:
                    filtered_lines.append(line[1].replace(" ", ""))

            catpage[category.name] = filtered_lines

        return Response({"status": True, "message": "Success", "data": catpage})

    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["GET"])
def homepage_recently_viewed_products(request):
    try:
        user_id = request.headers["User-id"]
        recents = userDB.objects.get(user_id=user_id).recently_viewed_products.all()
        products_data = []
        for recent in recents:
            col_var = ProductColorVariations.objects.get(
                main_product_id=recent, main_variant=True
            )
            size_var = ProductSizeVariations.objects.get(
                variant_id=col_var, main_size=True
            )
            reviews = review_db.objects.filter(main_product_id=recent)
            rating = 0
            for review in reviews:
                rating += review.rating

            products_data.append(
                {
                    "product_id": recent.id,
                    "product_image": "https://negbuy.com:8080"
                    + ProductColorVariationsSerializer(col_var).data["main_image"],
                    "product_name": recent.product_title,
                    "brand": recent.brand,
                    "mrp": size_var.mrp,
                    "price": size_var.selling_price,
                    "discount": round(
                        (
                            ((size_var.mrp - size_var.selling_price) / size_var.mrp)
                            * 100
                        ),
                        2,
                    ),
                    "city": recent.seller_id.city,
                    "country": recent.seller_id.country,
                    "rating": round((rating / len(reviews)), 1)
                    if len(reviews) != 0
                    else 0,
                }
            )

        return Response({"status": True, "message": "Success", "data": products_data})
    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["POST"])
def homepage_posters_brands(request):
    try:
        choice = request.data["choice"]
        return Response(
            {
                "status": True,
                "message": "Success",
                "data": [
                    {
                        "name": poster.name,
                        "image": "https://negbuy.com:8080"
                        + HomePage_Poster_BrandsSerializer(poster).data["image"],
                    }
                    for poster in HomePage_Poster_Brands.objects.filter(
                        choice=choice, status="active"
                    )
                ],
            }
        )
    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


# @api_view(["GET"])
# def subcat_dev(request):
#     try:
#         catpage = {}
#         for category in productCategory.objects.all():
#             file_path = os.path.join(settings.BASE_DIR, 'static_files', 'categories', f'{category.name}.txt')
#             with open(file_path, "r") as f:
#                 file_contents = f.read()

#             lines = file_contents.split("\n")
#             filtered_lines = []
#             for line in lines:
#                 line = line.split(">")
#                 # if len(line)>1 and line[1].replace(" ", "") not in filtered_lines:
#                 filtered_lines.append(line[-1])
#             # catpage[category.name]=filtered_lines

#             with open(f'./subcats/{category.name}.txt', 'w') as file:
#                 for item in filtered_lines:
#                     file.write("%s\n" % item)


#         return Response({"status": True, "message": "Success", "data": catpage})

#     except Exception as e:
#         return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["POST"])
def send_image(request):
    imageee = request.FILES.get("image")

    test_obj = test_image(image=imageee)

    test_obj.save()

    return Response({"status": True, "message": "Success", "data": "success"})


@api_view(["GET"])
def buyer_returned_products_function(request):
    try:
        user_id = request.headers["User-id"]
        user = userDB.objects.filter(user_id=user_id)
        response_list = list()
        if user:
            all_returns = buyer_returned_products.objects.all()

            for each_return in all_returns:
                response_dict = dict()

                response_dict = {
                    "image": each_return.ddp_order.product_id.variant_id.main_image.url
                    if each_return.shipment_type == "DDP"
                    else each_return.Exwork_order.product_id.variant_id.main_image.url,
                    "product_name": each_return.ddp_order.product_id.variant_id.main_product_id.product_title
                    if each_return.shipment_type == "DDP"
                    else each_return.Exwork_order.product_id.variant_id.main_product_id.product_title,
                    "refund_amount": each_return.refund_amount,
                    "status": each_return.status,
                    "action": each_return.action,
                    "shipment_type": each_return.shipment_type,
                }
                response_list.append(response_dict)

        return Response({"status": True, "message": "Success", "data": response_list})
    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["POST"])
def buyer_submit_return_request(request):
    try:
        user_id = request.headers["User-id"]
        user = userDB.objects.get(user_id=user_id)
        shipment_type = request.data.get("shipment_type")
        order_id = request.data.get("order_id")
        reason = request.data.get("reason")
        breif_reason = request.data.get("breif_reason")
        images = request.FILES.getlist("images")

        response_list = list()
        if user:
            if shipment_type == "DDP":
                ddp_order_obj = DDP_Orders_DB.objects.get(id=order_id)

                db_obj = buyer_returned_products.objects.create(
                    buyer=user,
                    shipment_type=shipment_type,
                    ddp_order=ddp_order_obj,
                    status="pending",
                    action="Under Observation",
                    reason=reason,
                    breif_reason=breif_reason,
                )
            else:
                exwork_order_obj = Exwork_Orders_DB.objects.get(id=order_id)

                db_obj = buyer_returned_products.objects.create(
                    buyer=user,
                    shipment_type=shipment_type,
                    Exwork_order=exwork_order_obj,
                    status="pending",
                    action="Under Observation",
                    reason=reason,
                    breif_reason=breif_reason,
                )

            for each_image in images:
                image_obj = return_images.objects.create(return_item=db_obj)
                image_obj.image = each_image
                image_obj.save()

            return Response({"status": True, "message": "Success", "data": ""})
    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


# ----------------------------------------------------------------------------------------------------------------------------
# Contributed by Ashutosh Tiwari

# from rest_framework import status
from django.core.mail import send_mail

@api_view(["POST"])
def MsgToSupp(request):
    try:
        # if request.method == "POST":
        user_id = request.headers["User-id"]
        user = userDB.objects.get(user_id=user_id)

        message = request.POST.get("message")
        checkbox = request.POST.get("checkbox")
        my_model_instance = MesasgeToSupplier.objects.create(
            message=message, checkbox=checkbox, user_id=user
        )
        my_model_instance.save()
        subject = 'New Message to Supplier'
        message = f'User: {user.username}\nPhone Number: {user.phone}\nMessage: {message}\nCheckbox: {checkbox}'
        from_email = 'rfq@negbuy.com' 
        recipient_list = ['amirmohd233@gmail.com'] 

        send_mail(subject, message, from_email, recipient_list)
        return JsonResponse({"status": "success"})
    except Exception as e:
        return Response({"status": "Error", "message": str(e), "data": {}})


@api_view(["POST"])
def ClientReview(request):
    if request.method == "POST":
        try:
            # Get user_id from request header
            user_id = request.headers.get("User-id")
            user = userDB.objects.get(user_id=user_id)
            # Extract other data from request body
            data = request.data
            main_product_id = data.get("main_product_id")
            main_product_id = ProductDB.objects.get(id=main_product_id)
            review_title = data.get("review_title")
            review_description = data.get("review_description")
            rating = data.get("rating")
            images = request.FILES.getlist("file")
            
            # Create a new review instance
            review_instance = review_db.objects.create(
                user=user,
                main_product_id=main_product_id,
                review_title=review_title,
                review_description=review_description,
                rating=rating,
            )
            for each_image in images:
                image_obj = Review_Images_DB.objects.create(review_id=review_instance)
                image_obj.file = each_image
                image_obj.save()
            
            return Response(
                {"status": "success", "message": "saved successfully", "id": review_instance.id},status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response(
                {"status": "error", "message": str(e)},status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    return Response(
        {"status": "error", "message": "Invalid request method"},
        status=status.HTTP_400_BAD_REQUEST,
    )

    

from rest_framework import status


@api_view(["GET"])
def primary_tags(request):
    primary_tags = Tag.objects.filter(parent_tag=None)
    serializer = TagSerializer(primary_tags, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(["POST"])
def sub_tags(request):
    serializer = TagIdSerializer(data=request.data)
    if serializer.is_valid():
        try:
            primary_tag_id = serializer.validated_data["id"]
            primary_tag = Tag.objects.get(id=primary_tag_id)
        except (Tag.DoesNotExist, ValueError):
            return Response(
                {"error": "Invalid primary tag ID"}, status=status.HTTP_404_NOT_FOUND
            )

        if not primary_tag.have_branch:
            with open("response.json") as json_file:
                data = json.load(json_file)
                last_subtag_name = primary_tag.name
                response_data = data.get(last_subtag_name, {})
                marking = False
            return Response({'data': response_data, 'marking': marking}, status=status.HTTP_200_OK)

        else:
            # Fetch subtags and return their data
            sub_tags = Tag.objects.filter(parent_tag=primary_tag)
            serializer = TagSerializer(sub_tags, many=True)
            marking = True
            return Response({'data': serializer.data, 'marking': marking}, status=status.HTTP_200_OK)
    else:
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



from django.conf import settings
import hashlib
import base64
import time

def generate_custom_token(user):
    user_id = str(user.pk)
    current_time = int(time.time())
    current_time_floor = current_time // 300 * 300 
    token_data = f"{user_id}:{current_time_floor}"
    
    hash_object = hashlib.sha256(token_data.encode())
    token = base64.urlsafe_b64encode(hash_object.digest()).rstrip(b'=').decode()
    return token



from django.contrib.auth.tokens import default_token_generator
from django.utils.encoding import force_bytes, force_text
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from djoser.utils import encode_uid
from .models import userDB


 
@api_view(['POST'])
def reset_password_request(request):
    email = request.data.get('email', '')

    try:
        user = userDB.objects.get(email=email)
    except userDB.DoesNotExist:
        return Response({'status': False,'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

    # Generate a token for the user
    uid = encode_uid(user.pk)
    token = generate_custom_token(user)

    # Include the token in the reset link
    reset_link = f'https://seller.negbuy.com/reset-password/{uid}/{token}/'

    html_message = render(request, 'reset-pass.html', {'reset_link': reset_link})
    send_mail(
        'Password Reset',
        f'',
        settings.EMAIL_HOST_USER,
        [email],
        html_message=html_message.content.decode('utf-8'),
        fail_silently=False,
    )

    return Response({'uid': uid,'token': token, 'status': True}, status=status.HTTP_200_OK)

@api_view(['POST'])
def forgot_password(request, uid, token):
    new_password = request.data.get('new_password', '')
    confirm_password = request.data.get('confirm_password', '')

    if new_password != confirm_password:
        return Response({'status': False,'error': 'New password and confirm password do not match'}, status=status.HTTP_400_BAD_REQUEST)

    
    try:
        user_id = urlsafe_base64_decode(uid).decode()
        user = userDB.objects.get(pk=user_id)
    except (TypeError, ValueError, OverflowError, userDB.DoesNotExist):
        return Response({'status': False,'error': 'Invalid reset link1'}, status=status.HTTP_400_BAD_REQUEST)

    
    if token != generate_custom_token(user):
        return Response({'status': False, 'error': 'Invalid reset link2'}, status=status.HTTP_400_BAD_REQUEST)

    user.password =(new_password)
    user.save()

    return Response({'status': True, 'success': 'Password changed successfully'}, status=status.HTTP_200_OK)

@api_view(['GET'])
def gettopclients(request):
    try:
        top_clients = topclients.objects.all()
    
        data = []
        for client in top_clients:
            data.append({
                'client_name': client.client_name,
                'order_quantity': client.order_quantity,
                'order_location': client.order_location,
                'client_status': client.client_status,
            })

        return Response({'status': True, 'data':data})
    except Exception as e:
        return Response({'status':False, 'error':str(e) })
# @api_view(['POST'])
# def new_letter(request):
#     email = request.data.get('email')
    
#     if '@' in email and '.' in email:
#         send_mail(
#         'News Letter from Negbuy',
#         f'Please find the attachments',
#         settings.EMAIL_HOST_USER,
#         [email],
#         # html_message=html_message.content.decode('utf-8'),
#         fail_silently=False,       
#         )
#         return Response({'Message':'News Letter sent successfully'}, status=status.HTTP_200_OK)
#     else:
#         return Response({'Message' : 'Email not valid!'}, status=status.HTTP_400_BAD_REQUEST)




@api_view(['GET'])
def DealsOfTheDay(request):
    try:
        queryset = DealsProductsInfo.objects.all()
        serializer = DealsProductsInfoSerializer(queryset, many=True, context={'request': request})
        return Response({'status':True, 'data': serializer.data})
    except Exception as e:
        return Response({'status': False, 'error': str(e)})

@api_view(['GET'])
def get_ware_house_details(request):
    try:
        queryset = Negbuy_Warehouse.objects.all()
        serializer = warehouseinfoserializer(queryset, many=True, context={'request': request})
        return Response({'status':True, 'data': serializer.data})
    except Exception as e:
        return Response({'status':False, 'error':str(e)})

@api_view(["POST"])
def verify_gst(request):
    try:
        user_id = request.headers["User-id"]
        gst_number = request.data["gst_number"]

        usr = userDB.objects.get(user_id=user_id, role="Seller")
        usr.gst_number = gst_number
        usr.save(update_fields=['gst_number'])

        gst_pattern = "^[0-9A-Z]{15}$"
        if not re.match(gst_pattern, gst_number):
            return Response({"msg": "Invalid GST Number", "status": False}, status=400)

        addr = "https://irisgst.com/gstin-filing-detail/?gstinno=" + gst_number
        response = requests.get(addr)
        htmlPage = bs4.BeautifulSoup(response.text, "html.parser")
        divData = htmlPage.find_all('div', {'class': 'form-group'})
        response = getGstObject(gst_number, divData)

        usr.gst_number = response['gst_number']
        usr.save(update_fields=['gst_number'])

        return Response(
            {
                "status": True,
                'response': response
            },
            status=200,
        )

    except userDB.DoesNotExist:
        return Response({"msg": "User not found", "status": False}, status=404)

    except Exception as e:
        return Response({"msgt": str(e), "status": False}, status=500)


@api_view(['POST'])
def get_distinct_colors_by_category(request):
    if request.method == 'POST':
        category_name = request.data.get('category_name')

        if not category_name:
            return Response({'error': 'Category name is required'}, status=400)

        try:
            category = productCategory.objects.get(name=category_name)
        except productCategory.DoesNotExist:
            return Response({'error': f'Category "{category_name}" not found'}, status=404)

        # Get distinct colors for variations in the specified category
        distinct_colors = ProductColorVariations.objects.filter(main_product_id__category_id=category).values_list('color', flat=True).distinct()

        return Response({'distinct_colors': list(distinct_colors)})

    return Response({'error': 'Invalid request method'}, status=405)